from __future__ import annotations

import io
import os
import sqlite3
from datetime import date

import httpx
import openpyxl
import pytest

pytestmark = [pytest.mark.e2e, pytest.mark.smoke]


@pytest.fixture(scope="session")
def client():
    with httpx.Client(base_url=os.environ["OPERATOROS_E2E_BACKEND_URL"], timeout=15.0) as session:
        yield session


@pytest.fixture(scope="session")
def authenticated(client):
    response = client.post("/api/auth/login", json={
        "username": os.environ["OPERATOROS_E2E_ADMIN_USERNAME"],
        "password": os.environ["OPERATOROS_E2E_ADMIN_PASSWORD"],
    })
    assert response.status_code == 200, response.text
    return client


def database_count(table: str) -> int:
    with sqlite3.connect(os.environ["OPERATOROS_E2E_DATABASE"]) as connection:
        return connection.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]


def test_authentication_authorization_and_hierarchy(client, authenticated):
    anonymous = httpx.get(f"{os.environ['OPERATOROS_E2E_BACKEND_URL']}/api/academic-masters/academic-years", timeout=15.0)
    assert anonymous.status_code == 401
    assert authenticated.get("/api/auth/me").json()["role"] == "admin"
    programs = authenticated.get("/api/academic-masters/programs").json()
    grades = authenticated.get("/api/academic-masters/grades").json()
    classes = authenticated.get("/api/academic-masters/classes").json()
    assert {item["name"] for item in programs if item["active"]} == {"MAIN", "SECONDARY MAIN"}
    assert {item["name"] for item in grades if item["active"]} == {"Primary 1", "Primary 2", "Secondary 7"}
    assert {"Primary 1A", "Primary 1B", "Primary 2A", "Next Primary 1A", "Next Primary 2A", "Secondary 7A"}.issubset({item["class_name"] for item in classes if item["active"]})
    assert "Primary 1 / MAIN" not in [item["class_name"] for item in classes if item["active"]]


def test_class_allocation_preview_and_attendance_filter_are_non_mutating(authenticated):
    enrollment_before = database_count("student_enrollments")
    years = authenticated.get("/api/academic-masters/academic-years").json()
    jenjangs = authenticated.get("/api/academic-masters/jenjangs").json()
    classes = authenticated.get("/api/academic-masters/classes").json()
    year_id, jenjang_id = next(item for item in years if item["status"] == "active")["id"], next(item for item in jenjangs if item["name"] == "Primary")["id"]
    active_class = next(item for item in classes if item["active"] and item["academic_year_id"] == year_id and item["class_name"] == "Primary 1A")
    candidates = authenticated.get("/api/grades/enrollment/candidates", params={"academic_year_id": year_id, "jenjang_id": jenjang_id})
    assert candidates.status_code == 200
    assert [item["name"] for item in candidates.json()] == ["E2E Bima", "E2E Citra"]
    attendance = authenticated.get("/api/review/attendance", params={
        "date": date.today().isoformat(), "academic_year_id": year_id, "academic_class_id": active_class["id"],
    })
    assert attendance.status_code == 200
    assert attendance.json()["total"] >= 1
    assert any(item["student_name"] == "E2E Ada" for item in attendance.json()["items"])
    assert attendance.json()["items"][0]["student_name"] == "E2E Ada"
    assert database_count("student_enrollments") == enrollment_before


def test_upload_preview_validation_does_not_commit_attendance(authenticated):
    attendance_before = database_count("attendance")
    invalid = authenticated.post("/api/uploads/preview", files={"file": ("bad.txt", b"bad", "text/plain")})
    assert invalid.status_code == 400

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["No. ID", "Nama", "Tanggal", "Scan Masuk", "Scan Pulang", "Terlambat", "Lembur", "Pengecualian", "week"])
    sheet.append(["E2E-DEVICE-2", "E2E Bima", date.today(), "07:20", "14:00", "00:05", "00:00", "", "E2E-WEEK"])
    payload = io.BytesIO()
    workbook.save(payload)
    valid = authenticated.post(
        "/api/uploads/preview",
        files={"file": ("e2e-attendance.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert valid.status_code == 200, valid.text
    assert "batch_id" in valid.json()
    assert database_count("attendance") == attendance_before


def test_student_management_identity_enrollment_roster_and_xlsx_round_trip(authenticated):
    years = authenticated.get("/api/academic-masters/academic-years").json()
    classes = [row for row in authenticated.get("/api/academic-masters/classes").json() if row["active"]]
    year_id = next(row for row in years if row["status"] == "active")["id"]
    class_a = next(row for row in classes if row["class_name"] == "Primary 1A")
    class_b = next(row for row in classes if row["class_name"] == "Primary 1B")
    created = authenticated.post("/api/student-masters", json={
        "identity": {"full_name": "E2E Student Management", "nisn": "0000990110", "nik": "3201000000990110", "student_status": "active"},
        "contact": {"address": "E2E Synthetic Address"},
        "guardians": [{"guardian_type": "guardian", "name": "E2E Guardian"}],
        "device_identity": {"device_identifier": "990110", "device_source": "attendance_machine", "effective_from": "2026-07-20", "reason": "E2E synthetic assignment"},
    })
    assert created.status_code == 201, created.text
    student = created.json(); student_id = student["id"]

    student["identity"]["preferred_name"] = "E2E Managed"
    edited = authenticated.patch(f"/api/student-masters/{student_id}/profile", json={
        "record_version": student["record_version"], "identity": student["identity"],
        "contact": student["contact"], "guardians": student["guardians"], "reason": "E2E profile edit",
    })
    assert edited.status_code == 200, edited.text
    replaced = authenticated.post(f"/api/student-masters/{student_id}/device-identities", json={
        "device_identifier": "990111", "device_source": "attendance_machine", "effective_from": "2026-08-01",
        "reason": "E2E device replacement", "confirmation": "REPLACE_ATTENDANCE_DEVICE_ID",
    })
    assert replaced.status_code == 201, replaced.text
    detail = authenticated.get(f"/api/student-masters/{student_id}/profile").json()
    assert len(detail["device_identities"]) == 2
    assert next(row for row in detail["device_identities"] if row["is_active"])["device_identifier"] == "990111"

    enrollment = authenticated.post(f"/api/student-enrollments/student/{student_id}", json={
        "academic_year_id": year_id, "academic_class_id": class_a["id"], "effective_from": "2026-07-20",
    })
    assert enrollment.status_code == 201, enrollment.text
    enrollment_id = enrollment.json()["id"]
    for target, effective in ((class_b, "2026-09-01"), (class_a, "2026-10-01")):
        moved = authenticated.post(f"/api/student-enrollments/{enrollment_id}/transfer", json={
            "target_class_id": target["id"], "effective_date": effective,
            "reason": "E2E reversible class transfer", "confirmation": "TRANSFER_STUDENT_ENROLLMENT",
        })
        assert moved.status_code == 200, moved.text
    history = authenticated.get(f"/api/student-enrollments/student/{student_id}").json()
    assert len(history[0]["class_history"]) == 3

    exported = authenticated.get("/api/student-masters/management/export-template")
    assert exported.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(exported.content))
    sheet = workbook["Student Data"]
    headers = [cell.value for cell in sheet[1]]
    target_row = next(row for row in range(2, sheet.max_row + 1) if sheet.cell(row, headers.index("OperatorOS Student UUID") + 1).value == student_id)
    sheet.cell(target_row, headers.index("Preferred Name") + 1).value = "E2E Workbook"
    payload = io.BytesIO(); workbook.save(payload)
    update_preview = authenticated.post("/api/student-masters/management/update-preview", files={
        "file": ("e2e-student-update.xlsx", payload.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    })
    assert update_preview.status_code == 200, update_preview.text
    preview_json = update_preview.json()
    update_row = next(row for row in preview_json["rows"] if row["student_master_id"] == student_id)
    assert update_row["classification"] == "UPDATE_EXISTING_MASTER"
    committed = authenticated.post(f"/api/student-masters/management/update-commit/{preview_json['id']}", json={
        "selected_row_ids": [update_row["id"]], "confirmation": "COMMIT_STUDENT_DATA_UPDATE",
        "preview_checksum": preview_json["preview_checksum"],
    })
    assert committed.status_code == 200, committed.text
    assert authenticated.get(f"/api/student-masters/{student_id}/profile").json()["identity"]["preferred_name"] == "E2E Workbook"

    roster = openpyxl.Workbook(); roster_sheet = roster.active; roster_sheet.title = "Roster"
    roster_sheet.append(["student_identifier", "student_name", "academic_year", "jenjang", "class_name", "program", "status", "nisn", "nik", "start_date"])
    roster_sheet.append(["990120", "E2E Roster Student", "2026/2027", "Primary", "Primary 1A", "MAIN", "active", "0000990120", "3201000000990120", "2026-07-20"])
    roster_bytes = io.BytesIO(); roster.save(roster_bytes)
    roster_preview = authenticated.post("/api/student-enrollments/roster-preview", data={"source_owner": "E2E Registrar", "date_received": "2026-07-20"}, files={"file": ("e2e-roster.xlsx", roster_bytes.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert roster_preview.status_code == 200, roster_preview.text
    roster_json = roster_preview.json(); assert roster_json["rows"][0]["classification"] == "CREATE_NEW_MASTER"
    roster_commit = authenticated.post("/api/student-enrollments/roster-commit", json={
        "preview_id": roster_json["preview_id"], "selected_row_ids": [1], "confirmation": "COMMIT_ACADEMIC_ROSTER", "preview_checksum": roster_json["preview_checksum"],
    })
    assert roster_commit.status_code == 200, roster_commit.text
    assert roster_commit.json()["students_created"] == 1


def test_progression_promotion_retention_graduation_cross_jenjang_stale_and_rollback(authenticated):
    years = authenticated.get("/api/academic-masters/academic-years").json()
    classes = authenticated.get("/api/academic-masters/classes").json()
    grades = authenticated.get("/api/academic-masters/grades").json()
    programs = authenticated.get("/api/academic-masters/programs").json()
    jenjangs = authenticated.get("/api/academic-masters/jenjangs").json()
    source_year = next(row for row in years if row["label"] == "2026/2027")
    destination_year = next(row for row in years if row["label"] == "2027/2028")
    class_by_name = {row["class_name"]: row for row in classes}
    grade_by_id = {row["id"]: row for row in grades}
    program_by_id = {row["id"]: row for row in programs}
    jenjang_by_name = {row["name"]: row for row in jenjangs}

    sequence = iter(range(991201, 991220))

    def create_source(name: str, source_class_name: str):
        device_identifier = str(next(sequence))
        created = authenticated.post("/api/student-masters", json={
            "identity": {"full_name": name, "student_status": "active"},
            "device_identity": {
                "device_identifier": device_identifier,
                "device_source": "attendance_machine",
                "effective_from": "2026-07-01",
                "reason": "Synthetic progression E2E identity",
            },
        })
        assert created.status_code == 201, created.text
        master_id = created.json()["id"]
        enrollment = authenticated.post(f"/api/student-enrollments/student/{master_id}", json={
            "academic_year_id": source_year["id"],
            "academic_class_id": class_by_name[source_class_name]["id"],
            "effective_from": "2026-07-01",
        })
        assert enrollment.status_code == 201, enrollment.text
        return {"master_id": master_id, "enrollment_id": enrollment.json()["id"], "legacy_id": int(device_identifier)}

    promoted = create_source("E2E Progress Promote", "Primary 1A")
    retained = create_source("E2E Progress Retain", "Primary 1A")
    graduated = create_source("E2E Progress Graduate", "Primary 2A")
    crossed = create_source("E2E Progress Cross", "Primary 2A")

    database = os.environ["OPERATOROS_E2E_DATABASE"]
    with sqlite3.connect(database) as connection:
        subject_id = connection.execute("SELECT id FROM subjects WHERE name='E2E Progression Subject'").fetchone()[0]
        component_id = connection.execute("SELECT id FROM assessment_components WHERE name='E2E Progression Score'").fetchone()[0]
        connection.execute("INSERT INTO student_subject_grades(enrollment_id,subject_id,component_id,score) VALUES(?,?,?,88.0)", (promoted["enrollment_id"], subject_id, component_id))
        connection.execute("INSERT INTO attendance(student_id,date,late_duration,late_source,is_absent,status) VALUES(?,'2026-08-01',0,'fixture',0,'Hadir')", (promoted["legacy_id"],))

    retain_class = class_by_name["Next Primary 1A"]
    retain_grade = grade_by_id[retain_class["grade_id"]]
    retain_program = program_by_id[retain_grade["program_id"]]
    cross_class = class_by_name["Secondary 7A"]
    cross_grade = grade_by_id[cross_class["grade_id"]]
    cross_program = program_by_id[cross_grade["program_id"]]
    preview = authenticated.post("/api/student-progression/previews", json={
        "source_academic_year_id": source_year["id"],
        "destination_academic_year_id": destination_year["id"],
        "source_enrollment_ids": [promoted["enrollment_id"], retained["enrollment_id"], graduated["enrollment_id"], crossed["enrollment_id"]],
        "overrides": [
            {
                "source_enrollment_id": retained["enrollment_id"], "outcome": "RETAIN",
                "destination_jenjang_id": retain_program["jenjang_id"], "destination_program_id": retain_program["id"],
                "destination_grade_id": retain_grade["id"], "destination_class_id": retain_class["id"],
                "reason_code": "RETENTION_APPROVED", "reason": "Synthetic retention review",
            },
            {
                "source_enrollment_id": crossed["enrollment_id"], "outcome": "CROSS_JENJANG",
                "destination_jenjang_id": jenjang_by_name["Secondary"]["id"], "destination_program_id": cross_program["id"],
                "destination_grade_id": cross_grade["id"], "destination_class_id": cross_class["id"],
                "reason_code": "CROSS_JENJANG_APPROVED", "reason": "Synthetic cross-Jenjang review",
            },
        ],
    })
    assert preview.status_code == 201, preview.text
    preview_json = preview.json()
    assert {row["proposed_outcome"] for row in preview_json["rows"]} == {"PROMOTE", "RETAIN", "GRADUATE", "CROSS_JENJANG"}
    committed = authenticated.post(f"/api/student-progression/previews/{preview_json['batch_id']}/commit", json={
        "preview_version": preview_json["preview_version"],
        "effective_date": destination_year["start_date"],
        "confirmation": "COMMIT_CROSS_JENJANG_PROGRESSION",
    })
    assert committed.status_code == 200, committed.text
    assert committed.json()["destination_enrollments_created"] == 3
    assert committed.json()["graduated"] == committed.json()["retained"] == committed.json()["cross_jenjang"] == 1
    with sqlite3.connect(database) as connection:
        assert connection.execute("SELECT COUNT(*) FROM student_subject_grades WHERE enrollment_id=? AND score=88.0", (promoted["enrollment_id"],)).fetchone()[0] == 1
        assert connection.execute("SELECT COUNT(*) FROM attendance WHERE student_id=? AND date='2026-08-01'", (promoted["legacy_id"],)).fetchone()[0] == 1
        assert connection.execute("SELECT lifecycle_state FROM student_enrollments WHERE id=?", (graduated["enrollment_id"],)).fetchone()[0] == "GRADUATED"
        assert connection.execute("SELECT COUNT(*) FROM student_progression_audit WHERE batch_id=?", (preview_json["batch_id"],)).fetchone()[0] == 4

    stale_one = create_source("E2E Progress Stale", "Primary 1A")
    stale_preview = authenticated.post("/api/student-progression/previews", json={
        "source_academic_year_id": source_year["id"], "destination_academic_year_id": destination_year["id"],
        "source_enrollment_ids": [stale_one["enrollment_id"]], "overrides": [],
    }).json()
    revalidated = authenticated.post(f"/api/student-progression/previews/{stale_preview['batch_id']}/revalidate", json={"preview_version": 1})
    assert revalidated.status_code == 200 and revalidated.json()["preview_version"] == 2
    stale_commit = authenticated.post(f"/api/student-progression/previews/{stale_preview['batch_id']}/commit", json={
        "preview_version": 1, "effective_date": destination_year["start_date"], "confirmation": "COMMIT_STUDENT_PROGRESSION",
    })
    assert stale_commit.status_code == 409 and stale_commit.json()["detail"]["code"] == "PROGRESSION_PREVIEW_STALE"

    rollback_one = create_source("E2E Progress Rollback One", "Primary 1A")
    rollback_two = create_source("E2E Progress Rollback Two", "Primary 1A")
    rollback_preview = authenticated.post("/api/student-progression/previews", json={
        "source_academic_year_id": source_year["id"], "destination_academic_year_id": destination_year["id"],
        "source_enrollment_ids": [rollback_one["enrollment_id"], rollback_two["enrollment_id"]], "overrides": [],
    }).json()
    trigger_name = "e2e_inject_progression_failure"
    with sqlite3.connect(database) as connection:
        connection.execute(
            f"CREATE TRIGGER {trigger_name} BEFORE INSERT ON student_progression_audit "
            "WHEN NEW.batch_id=? AND NEW.preview_row_id=2 BEGIN SELECT RAISE(ABORT, 'synthetic progression failure'); END".replace("?", f"'{rollback_preview['batch_id']}'", 1)
        )
    failed = authenticated.post(f"/api/student-progression/previews/{rollback_preview['batch_id']}/commit", json={
        "preview_version": 1, "effective_date": destination_year["start_date"], "confirmation": "COMMIT_STUDENT_PROGRESSION",
    })
    with sqlite3.connect(database) as connection:
        connection.execute(f"DROP TRIGGER IF EXISTS {trigger_name}")
        states = connection.execute("SELECT lifecycle_state FROM student_enrollments WHERE id IN (?,?) ORDER BY id", (rollback_one["enrollment_id"], rollback_two["enrollment_id"])).fetchall()
        destination_count = connection.execute("SELECT COUNT(*) FROM student_enrollments WHERE student_master_id IN (?,?) AND academic_year_id=?", (rollback_one["master_id"], rollback_two["master_id"], destination_year["id"])).fetchone()[0]
    assert failed.status_code == 409 and failed.json()["detail"]["code"] == "PROGRESSION_TRANSACTION_FAILED"
    assert states == [("ACTIVE",), ("ACTIVE",)] and destination_count == 0
