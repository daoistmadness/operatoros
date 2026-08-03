"""Safe, auditable canonical student import for a single-school XLSX source."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import create_engine, func, inspect
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, sessionmaker

from core.student_import_schema_migration import (
    STUDENT_IMPORT_SCHEMA_VERSION,
    _safe_database_path,
    ensure_student_import_schema,
    student_import_schema_ready,
)
from models.academic_master import AcademicClass, AcademicGrade, AcademicProgram
from models.academic_roster import AcademicRosterImportBatch
from models.academic_year import AcademicYear
from models.jenjang import Jenjang
from models.student import Student
from models.student_enrollment import StudentEnrollment
from models.student_import_session import StudentImportAppliedAction, StudentImportSession
from models.student_master import (
    StudentDeviceIdentity,
    StudentEnrollmentClassHistory,
    StudentImportBatch,
    StudentMaster,
    StudentMasterChangeHistory,
)
from services.student_import_sessions import append_action, mark_committed, state_checksum
from services.student_management import _audit
from services.student_normalization import mask_identifier, normalize_name


IMPORT_CONFIRMATION = "APPLY_CANONICAL_STUDENT_IMPORT"
SOURCE_SYSTEM = "DAPODIK"
WORKBOOK_SCHEMA_VERSION = "1"

SHEET_REQUIRED = {
    "School": {"school_id", "name", "npsn"},
    "Classes": {"rombongan_belajar_id", "semester_id", "school_id", "jenjang", "program", "grade", "class_name"},
    "Students": {"peserta_didik_id", "name", "school_id", "rombongan_belajar_id", "status"},
}
IDENTIFIER_FIELDS = {"peserta_didik_id", "nis", "nipd", "nisn", "legacy_student_id", "rombongan_belajar_id", "registrasi_id", "anggota_rombel_id"}
DATE_FIELDS = {"birth_date", "admission_date", "source_last_update"}
OPTIONAL_NOT_IMPORTED = {
    "address", "phone", "email", "guardian_name", "guardian_phone", "parent_name", "parent_phone",
    "nik", "no_kk", "income", "occupation", "bank_account", "username", "password", "pengguna_id",
}
STATUS_VALUES = {"active", "inactive", "transferred", "withdrawn", "graduated"}


def _header(value: Any) -> str:
    text = str(value or "").strip().casefold()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _identifier(value: Any, field: str, issues: list[dict[str, str]]) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool) or isinstance(value, (int, float)):
        issues.append({"code": "SUSPICIOUS_NUMERIC_IDENTIFIER", "field": field})
        return None
    text = str(value).strip()
    if text.startswith("=") or re.search(r"[eE][+-]?\d+", text):
        issues.append({"code": "SUSPICIOUS_IDENTIFIER_FORMAT", "field": field})
        return None
    return text or None


def _date_value(value: Any, field: str, issues: list[dict[str, str]]) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if text.startswith("="):
        issues.append({"code": "FORMULA_CELL_REJECTED", "field": field})
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    issues.append({"code": "INVALID_DATE", "field": field})
    return None


def _read_sheet(workbook, sheet_name: str) -> tuple[list[dict[str, Any]], list[str]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"MISSING_REQUIRED_SHEET:{sheet_name}")
    sheet = workbook[sheet_name]
    if sheet.sheet_state != "visible":
        raise ValueError(f"REQUIRED_SHEET_HIDDEN:{sheet_name}")
    rows = list(sheet.iter_rows(values_only=False))
    if not rows:
        raise ValueError(f"EMPTY_SHEET:{sheet_name}")
    headers = [_header(cell.value) for cell in rows[0]]
    if len(headers) != len(set(headers)):
        raise ValueError(f"DUPLICATE_HEADERS:{sheet_name}")
    missing = sorted(SHEET_REQUIRED[sheet_name] - set(headers))
    if missing:
        raise ValueError(f"MISSING_COLUMNS:{sheet_name}:{','.join(missing)}")
    unknown_sensitive = sorted(set(headers) & OPTIONAL_NOT_IMPORTED)
    parsed = []
    for row_number, cells in enumerate(rows[1:], start=2):
        if not any(cell.value not in (None, "") for cell in cells):
            continue
        values = {header: cells[index].value for index, header in enumerate(headers)}
        parsed.append({"source_row": row_number, "values": values})
    return parsed, unknown_sensitive


def _read_workbook(path: Path) -> tuple[dict[str, Any], str]:
    if path.suffix.casefold() != ".xlsx":
        raise ValueError("STUDENT_IMPORT_SOURCE_MUST_BE_XLSX")
    if not path.is_file():
        raise ValueError("STUDENT_IMPORT_SOURCE_NOT_FOUND")
    checksum = hashlib.sha256(path.read_bytes()).hexdigest()
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        sheets = {}
        warnings: list[dict[str, str]] = []
        for name in ("School", "Classes", "Students"):
            rows, sensitive = _read_sheet(workbook, name)
            sheets[name] = rows
            for field in sensitive:
                warnings.append({"code": "OPTIONAL_FIELD_NOT_IMPORTED", "field": field})
        return {"sheets": sheets, "warnings": warnings, "source_filename": path.name}, checksum
    finally:
        workbook.close()


def _mask(value: str | None) -> str | None:
    return mask_identifier(value) if value else None


def _read_only_session(database: Path) -> tuple[Session, Any]:
    uri = f"sqlite:///file:{database.as_posix()}?mode=ro&immutable=1&uri=true"
    engine = create_engine(uri, connect_args={"check_same_thread": False})
    return sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)(), engine


def _schema_columns(database: Path) -> dict[str, set[str]]:
    session, engine = _read_only_session(database)
    try:
        inspector = inspect(engine)
        return {
            table: {column["name"] for column in inspector.get_columns(table)}
            for table in ("student_masters", "student_enrollments", "academic_classes")
            if table in inspector.get_table_names()
        }
    finally:
        session.close()
        engine.dispose()


def _resolve_class(db: Session, class_payload: dict[str, Any], year: AcademicYear):
    source_id = class_payload["rombongan_belajar_id"]
    if source_id:
        candidates = db.query(AcademicClass).filter(AcademicClass.dapodik_rombongan_belajar_id == source_id).all()
        if len(candidates) > 1:
            return None, "AMBIGUOUS_CLASS"
        if candidates:
            academic_class = candidates[0]
            if academic_class.academic_year_id != year.id:
                return None, "ACADEMIC_YEAR_CONTRADICTION"
            grade = db.get(AcademicGrade, academic_class.grade_id)
            program = db.get(AcademicProgram, grade.program_id) if grade else None
            jenjang = db.get(Jenjang, grade.jenjang_id) if grade else None
            if not grade or not program or not jenjang or any(
                (
                    grade.name.casefold() != class_payload["grade"].casefold(),
                    program.name.casefold() != class_payload["program"].casefold(),
                    jenjang.name.casefold() != class_payload["jenjang"].casefold(),
                    academic_class.class_name.casefold() != class_payload["class_name"].casefold(),
                )
            ):
                return None, "ACADEMIC_REFERENCE_CONTRADICTION"
            if academic_class.dapodik_sekolah_id not in (None, class_payload["school_id"]):
                return None, "SCHOOL_SCOPE_CONFLICT"
            return academic_class, None

    jenjang = db.query(Jenjang).filter(func.lower(Jenjang.name) == class_payload["jenjang"].casefold()).all()
    if len(jenjang) != 1:
        return None, "ACADEMIC_REFERENCE_UNRESOLVED"
    jenjang_row = jenjang[0]
    programs = db.query(AcademicProgram).filter(
        AcademicProgram.jenjang_id == jenjang_row.id,
        func.lower(AcademicProgram.name) == class_payload["program"].casefold(),
        AcademicProgram.active.is_(True),
    ).all()
    if len(programs) != 1:
        return None, "ACADEMIC_REFERENCE_UNRESOLVED"
    grades = db.query(AcademicGrade).filter(
        AcademicGrade.program_id == programs[0].id,
        func.lower(AcademicGrade.name) == class_payload["grade"].casefold(),
        AcademicGrade.active.is_(True),
    ).all()
    if len(grades) != 1:
        return None, "ACADEMIC_REFERENCE_UNRESOLVED"
    classes = db.query(AcademicClass).filter(
        AcademicClass.academic_year_id == year.id,
        AcademicClass.grade_id == grades[0].id,
        func.lower(AcademicClass.class_name) == class_payload["class_name"].casefold(),
        AcademicClass.active.is_(True),
    ).all()
    if len(classes) != 1:
        return None, "AMBIGUOUS_CLASS" if len(classes) > 1 else "ACADEMIC_REFERENCE_UNRESOLVED"
    return classes[0], None


def _resolve_identity(db: Session, payload: dict[str, Any]) -> tuple[StudentMaster | None, str | None, list[str]]:
    matches: dict[str, list[StudentMaster]] = {}
    if payload["peserta_didik_id"]:
        matches["dapodik_id"] = db.query(StudentMaster).filter(
            StudentMaster.dapodik_peserta_didik_id == payload["peserta_didik_id"]
        ).all()
    if payload["nisn"]:
        matches["nisn"] = db.query(StudentMaster).filter(StudentMaster.nisn == payload["nisn"]).all()
    if payload["nipd"]:
        matches["nipd"] = db.query(StudentMaster).filter(StudentMaster.nipd == payload["nipd"]).all()

    nonempty = {rule: rows for rule, rows in matches.items() if rows}
    if any(len(rows) > 1 for rows in nonempty.values()):
        return None, "AMBIGUOUS_IDENTITY", []
    identities = {rows[0].id for rows in nonempty.values()}
    if len(identities) > 1:
        return None, "CONFLICTING_IDENTIFIERS", []
    if identities:
        master_id = next(iter(identities))
        master = db.get(StudentMaster, master_id)
        rule = next(rule for rule, rows in nonempty.items() if rows)
        conflicts: list[str] = []
        if master.dapodik_peserta_didik_id and master.dapodik_peserta_didik_id != payload["peserta_didik_id"]:
            if payload["peserta_didik_id"]:
                conflicts.append("dapodik_peserta_didik_id")
        for field in ("full_name", "birth_date", "gender", "nisn", "nipd", "student_status"):
            incoming = payload.get(field)
            current = getattr(master, field)
            if incoming not in (None, "") and current not in (None, ""):
                same = normalize_name(incoming) == normalize_name(current) if field == "full_name" else incoming == current
                if not same:
                    conflicts.append(field)
        return master, rule, sorted(set(conflicts))

    normalized = normalize_name(payload["full_name"])
    candidate_query = db.query(StudentMaster).filter(StudentMaster.normalized_name == normalized)
    if payload["birth_date"]:
        candidate_query = candidate_query.filter(StudentMaster.birth_date == payload["birth_date"])
    candidates = candidate_query.order_by(StudentMaster.id).all()
    if candidates:
        return None, "POSSIBLE_DUPLICATE", [candidate.id for candidate in candidates]
    if not any(payload.get(field) for field in ("peserta_didik_id", "nisn", "nipd")):
        return None, "INVALID_REQUIRED_FIELD", []
    return None, "NEW_STUDENT", []


def _row_summary(row: dict[str, Any]) -> dict[str, Any]:
    payload = row["payload"]
    return {
        "source_row": row["source_row"],
        "classification": row["classification"],
        "match_rule": row.get("match_rule"),
        "student_master_id": row.get("student_master_id"),
        "identifiers": {
            "peserta_didik_id": _mask(payload.get("peserta_didik_id")),
            "nisn": _mask(payload.get("nisn")),
            "nipd": _mask(payload.get("nipd")),
        },
        "legacy_student_id": _mask(payload.get("legacy_student_id")),
        "errors": row.get("errors", []),
        "conflicting_fields": row.get("conflicting_fields", []),
        "proposed_actions": row.get("proposed_actions", []),
    }


def _report_checksum(report: dict[str, Any]) -> str:
    return hashlib.sha256(json.dumps(report, sort_keys=True, default=str, separators=(",", ":")).encode()).hexdigest()


def validate_import(
    file: str | Path, database: str | Path, academic_year: str, *, include_internal: bool = False,
) -> dict[str, Any]:
    path = Path(file)
    database_path = _safe_database_path(database)
    if not student_import_schema_ready(database_path):
        raise ValueError("STUDENT_IMPORT_SCHEMA_REQUIRED")
    workbook, file_checksum = _read_workbook(path)
    session, engine = _read_only_session(database_path)
    try:
        school_rows = workbook["sheets"]["School"]
        if len(school_rows) != 1:
            raise ValueError("SCHOOL_SHEET_MUST_CONTAIN_ONE_ROW")
        school = {key: _text(value) for key, value in school_rows[0]["values"].items()}
        school_id = school.get("school_id")
        if not school_id:
            raise ValueError("SCHOOL_ID_REQUIRED")
        year = session.query(AcademicYear).filter(AcademicYear.label == academic_year).one_or_none()
        if year is None:
            raise ValueError("ACADEMIC_YEAR_UNRESOLVED")
        classes: dict[str, dict[str, Any]] = {}
        class_errors: dict[str, str] = {}
        for source in workbook["sheets"]["Classes"]:
            values = source["values"]
            issues: list[dict[str, str]] = []
            class_payload = {
                "rombongan_belajar_id": _identifier(values.get("rombongan_belajar_id"), "rombongan_belajar_id", issues),
                "semester_id": _identifier(values.get("semester_id"), "semester_id", issues),
                "school_id": _identifier(values.get("school_id"), "school_id", issues),
                "jenjang": _text(values.get("jenjang")), "program": _text(values.get("program")),
                "grade": _text(values.get("grade")), "class_name": _text(values.get("class_name")),
                "source_row": source["source_row"], "issues": issues,
            }
            if class_payload["school_id"] != school_id:
                class_errors[class_payload["rombongan_belajar_id"] or f"row-{source['source_row']}"] = "SCHOOL_SCOPE_CONFLICT"
                continue
            if issues or not all(class_payload.get(field) for field in ("rombongan_belajar_id", "semester_id", "jenjang", "program", "grade", "class_name")):
                class_errors[class_payload["rombongan_belajar_id"] or f"row-{source['source_row']}"] = "INVALID_CLASS_ROW"
                continue
            if class_payload["rombongan_belajar_id"] in classes:
                class_errors[class_payload["rombongan_belajar_id"]] = "DUPLICATE_CLASS_SOURCE_ID"
                continue
            classes[class_payload["rombongan_belajar_id"]] = class_payload
        rows: list[dict[str, Any]] = []
        seen_source_identifiers: dict[tuple[str, str], int] = {}
        for source in workbook["sheets"]["Students"]:
            values = source["values"]
            issues: list[dict[str, str]] = []
            payload = {
                "peserta_didik_id": _identifier(values.get("peserta_didik_id"), "peserta_didik_id", issues),
                "full_name": _text(values.get("name")) or "",
                "school_id": _identifier(values.get("school_id"), "school_id", issues),
                "rombongan_belajar_id": _identifier(values.get("rombongan_belajar_id"), "rombongan_belajar_id", issues),
                "nisn": _identifier(values.get("nisn"), "nisn", issues),
                "nipd": _identifier(values.get("nipd") or values.get("nis"), "nipd", issues),
                "gender": _text(values.get("gender")),
                "birth_place": _text(values.get("birth_place")),
                "birth_date": _date_value(values.get("birth_date"), "birth_date", issues),
                "status": (_text(values.get("status")) or "").casefold(),
                "legacy_student_id": _identifier(values.get("legacy_student_id"), "legacy_student_id", issues),
                "registrasi_id": _identifier(values.get("registrasi_id"), "registrasi_id", issues),
                "anggota_rombel_id": _identifier(values.get("anggota_rombel_id"), "anggota_rombel_id", issues),
                "admission_date": _date_value(values.get("admission_date"), "admission_date", issues),
                "source_last_update": _date_value(values.get("source_last_update"), "source_last_update", issues),
            }
            row: dict[str, Any] = {"source_row": source["source_row"], "payload": payload, "errors": [], "proposed_actions": []}
            row["errors"].extend(issue["code"] for issue in issues)
            for field in ("peserta_didik_id", "nisn", "nipd"):
                value = payload.get(field)
                if not value:
                    continue
                key = (field, value)
                if key in seen_source_identifiers:
                    row["errors"].append("DUPLICATE_SOURCE_IDENTIFIER")
                else:
                    seen_source_identifiers[key] = source["source_row"]
            if payload["school_id"] != school_id:
                row["errors"].append("SCHOOL_SCOPE_CONFLICT")
            if not payload["full_name"] or not payload["peserta_didik_id"] or not payload["rombongan_belajar_id"]:
                row["errors"].append("MISSING_REQUIRED_FIELD")
            if payload["status"] not in STATUS_VALUES:
                row["errors"].append("UNKNOWN_STUDENT_STATUS")
            class_payload = classes.get(payload["rombongan_belajar_id"])
            if class_payload is None:
                row["errors"].append(class_errors.get(payload["rombongan_belajar_id"], "ACADEMIC_CLASS_UNRESOLVED"))
            else:
                academic_class, class_error = _resolve_class(session, class_payload, year)
                if class_error:
                    row["errors"].append(class_error)
                else:
                    row["academic_class_id"] = academic_class.id
                    row["academic_year_id"] = year.id
                    grade = session.get(AcademicGrade, academic_class.grade_id)
                    row["jenjang_id"] = grade.jenjang_id if grade else None
                    row["academic_class_source_id"] = class_payload["rombongan_belajar_id"]
                    row["academic_year_start"] = year.start_date
                    row["semester_id"] = class_payload["semester_id"]
            master, match_rule, conflicts = _resolve_identity(session, payload) if not row["errors"] else (None, None, [])
            row["student_master_id"] = master.id if master else None
            row["match_rule"] = match_rule
            row["conflicting_fields"] = conflicts
            if match_rule == "POSSIBLE_DUPLICATE":
                row["errors"].append("POSSIBLE_DUPLICATE")
            elif conflicts:
                row["errors"].append("CONFLICTING_CANONICAL_FIELD")
            if payload["legacy_student_id"]:
                try:
                    legacy = session.get(Student, int(payload["legacy_student_id"]))
                except ValueError:
                    legacy = None
                if legacy is None:
                    row["errors"].append("LEGACY_STUDENT_NOT_FOUND")
                else:
                    active_links = session.query(StudentDeviceIdentity).filter(
                        StudentDeviceIdentity.legacy_student_id == legacy.id,
                        StudentDeviceIdentity.is_active.is_(True),
                    ).all()
                    if any(link.student_master_id != row["student_master_id"] for link in active_links):
                        row["errors"].append("LEGACY_LINK_CONFLICT")
                    else:
                        row["proposed_actions"].append("LINK_LEGACY_STUDENT")
            if not row["errors"]:
                existing_enrollment = None
                if master:
                    existing_enrollment = session.query(StudentEnrollment).filter_by(
                        student_master_id=master.id, academic_year_id=year.id,
                    ).one_or_none()
                    if existing_enrollment:
                        row["existing_enrollment_id"] = existing_enrollment.id
                        if existing_enrollment.academic_class_id == row.get("academic_class_id"):
                            row["proposed_actions"].append("SKIP_EXISTING_ENROLLMENT")
                        else:
                            row["proposed_actions"].append("TRANSFER_ENROLLMENT")
                if match_rule == "NEW_STUDENT":
                    row["classification"] = "NEW_STUDENT"
                    row["proposed_actions"].extend(["CREATE_STUDENT_MASTER", "CREATE_ENROLLMENT"])
                else:
                    row["classification"] = "READY_TO_APPLY"
                    if not row["proposed_actions"]:
                        row["proposed_actions"].append("UPDATE_STUDENT_MASTER")
            else:
                row["classification"] = "REVIEW_REQUIRED"
            rows.append(row)

        counts = Counter(row["classification"] for row in rows)
        report = {
            "status": "READY" if not any(row["classification"] == "REVIEW_REQUIRED" for row in rows) else "REVIEW_REQUIRED",
            "source_system": SOURCE_SYSTEM,
            "schema_version": WORKBOOK_SCHEMA_VERSION,
            "source_filename": workbook["source_filename"],
            "source_checksum": file_checksum,
            "school_id": school_id,
            "academic_year": academic_year,
            "total_rows": len(rows),
            "counts": dict(sorted(counts.items())),
            "warnings": workbook["warnings"],
            "rows": [_row_summary(row) for row in rows],
        }
        report["report_checksum"] = _report_checksum(report)
        if include_internal:
            report["_internal_rows"] = rows
            report["_school"] = school
        return report
    finally:
        session.close()
        engine.dispose()


def _legacy_link(
    db: Session, master: StudentMaster, legacy_id: int, actor: str, effective: date,
    import_batch_id: str | None = None,
) -> StudentDeviceIdentity:
    legacy = db.get(Student, legacy_id)
    if legacy is None:
        raise ValueError("LEGACY_STUDENT_NOT_FOUND")
    active = db.query(StudentDeviceIdentity).filter(
        StudentDeviceIdentity.legacy_student_id == legacy_id,
        StudentDeviceIdentity.is_active.is_(True),
    ).all()
    if any(link.student_master_id != master.id for link in active):
        raise ValueError("LEGACY_LINK_CONFLICT")
    existing = next((link for link in active if link.student_master_id == master.id), None)
    if existing:
        return existing
    if normalize_name(legacy.name) != normalize_name(master.full_name):
        raise ValueError("LEGACY_IDENTITY_CONFLICT")
    mapping = StudentDeviceIdentity(
        student_master_id=master.id, legacy_student_id=legacy_id,
        device_identifier=str(legacy_id), device_source="legacy_students",
        effective_from=effective, is_active=True, created_by=actor,
    )
    db.add(mapping)
    db.flush()
    _audit(
        db, master.id, "legacy_identity_linked", actor, "dapodik_roster_import",
        "legacy_student_id", None, legacy_id, import_batch_id,
    )
    return mapping


def _apply_row(
    db: Session, row: dict[str, Any], actor: str, roster_batch_id: str,
    student_import_batch_id: str, session: StudentImportSession, sequence: int,
) -> tuple[int, StudentMaster]:
    payload = row["payload"]
    master = db.get(StudentMaster, row.get("student_master_id")) if row.get("student_master_id") else None
    if master is None:
        master = StudentMaster(
            full_name=payload["full_name"], normalized_name=normalize_name(payload["full_name"]),
            nipd=payload.get("nipd"), nisn=payload.get("nisn"), gender=payload.get("gender"),
            birth_place=payload.get("birth_place"), birth_date=payload.get("birth_date"),
            student_status=payload["status"], dapodik_peserta_didik_id=payload["peserta_didik_id"],
            dapodik_sekolah_id=payload["school_id"],
            dapodik_last_update_at=payload.get("source_last_update"), created_by=actor, updated_by=actor,
            admission_date=payload.get("admission_date"),
        )
        db.add(master)
        db.flush()
        _audit(
            db, master.id, "student_created", actor, "dapodik_roster_import",
            import_batch_id=student_import_batch_id,
        )
        append_action(
            db, session, source_row=row["source_row"], sequence=sequence,
            action_type="CREATE_STUDENT_MASTER", entity_type="STUDENT_MASTER", entity_id=master.id,
            actor=actor, before_state=None,
            after_state={"student_master_id": master.id, "source_id": _mask(payload["peserta_didik_id"])},
            compensation_type="DEACTIVATE_CREATED_MASTER", eligibility="ELIGIBLE", roster_batch_id=roster_batch_id,
        )
        sequence += 1
    else:
        changed: list[str] = []
        for field, value in (
            ("dapodik_peserta_didik_id", payload.get("peserta_didik_id")),
            ("dapodik_sekolah_id", payload.get("school_id")),
            ("nipd", payload.get("nipd")), ("nisn", payload.get("nisn")),
            ("gender", payload.get("gender")), ("birth_place", payload.get("birth_place")),
            ("birth_date", payload.get("birth_date")), ("student_status", payload.get("status")),
        ):
            if getattr(master, field) in (None, "") and value not in (None, ""):
                setattr(master, field, value); changed.append(field)
        master.updated_by = actor
        if master.admission_date is None and payload.get("admission_date"):
            master.admission_date = payload["admission_date"]
        if changed:
            _audit(
                db, master.id, "profile_updated", actor, "dapodik_roster_import", "fields", None,
                ",".join(sorted(changed)), student_import_batch_id,
            )
            append_action(
                db, session, source_row=row["source_row"], sequence=sequence,
                action_type="UPDATE_STUDENT_PROFILE", entity_type="STUDENT_MASTER", entity_id=master.id,
                actor=actor, before_state={"changed_fields": sorted(changed)},
                after_state={"changed_fields": sorted(changed)}, compensation_type="MANUAL_REVIEW_REQUIRED",
                eligibility="MANUAL_REVIEW_REQUIRED", roster_batch_id=roster_batch_id,
            )
            sequence += 1

    if payload.get("legacy_student_id"):
        _legacy_link(
            db, master, int(payload["legacy_student_id"]), actor,
            payload.get("admission_date") or date.today(), student_import_batch_id,
        )
        append_action(
            db, session, source_row=row["source_row"], sequence=sequence,
            action_type="LINK_LEGACY_STUDENT", entity_type="LEGACY_STUDENT", entity_id=payload["legacy_student_id"],
            actor=actor, before_state=None, after_state={"student_master_id": master.id},
            compensation_type="RETIRE_DEVICE_MAPPING", eligibility="MANUAL_REVIEW_REQUIRED", roster_batch_id=roster_batch_id,
        )
        sequence += 1

    academic_class = db.get(AcademicClass, row["academic_class_id"])
    if academic_class is None:
        raise ValueError("ACADEMIC_CLASS_UNRESOLVED")
    if academic_class.dapodik_rombongan_belajar_id in (None, payload.get("rombongan_belajar_id")):
        academic_class.dapodik_rombongan_belajar_id = academic_class.dapodik_rombongan_belajar_id or payload.get("rombongan_belajar_id")
        academic_class.dapodik_sekolah_id = academic_class.dapodik_sekolah_id or payload.get("school_id")
        academic_class.dapodik_semester_id = academic_class.dapodik_semester_id or row.get("semester_id")
        academic_class.dapodik_last_update_at = academic_class.dapodik_last_update_at or payload.get("source_last_update")

    enrollment = db.query(StudentEnrollment).filter_by(
        student_master_id=master.id, academic_year_id=row["academic_year_id"],
    ).one_or_none()
    legacy_id = int(payload["legacy_student_id"]) if payload.get("legacy_student_id") else None
    if enrollment is None:
        enrollment = StudentEnrollment(
            student_id=legacy_id, student_master_id=master.id,
            academic_year_id=row["academic_year_id"], jenjang_id=row["jenjang_id"],
            academic_class_id=row["academic_class_id"], class_name=None, class_assigned=True,
            effective_from=row["academic_year_start"], lifecycle_state="ACTIVE",
            lifecycle_effective_date=row["academic_year_start"],
            dapodik_registrasi_id=payload.get("registrasi_id"),
            dapodik_anggota_rombel_id=payload.get("anggota_rombel_id"),
            dapodik_sekolah_id=payload.get("school_id"), dapodik_semester_id=row.get("semester_id"),
        )
        enrollment.class_name = academic_class.class_name
        db.add(enrollment); db.flush()
        db.add(StudentEnrollmentClassHistory(
            enrollment_id=enrollment.id, class_name=enrollment.class_name,
            effective_from=enrollment.effective_from, changed_by=actor, source="dapodik_roster_import",
        ))
        _audit(
            db, master.id, "enrollment_created", actor, "dapodik_roster_import", "academic_class_id",
            None, enrollment.academic_class_id, student_import_batch_id,
        )
        action_type = "CREATE_ENROLLMENT"
    else:
        if enrollment.student_id not in (None, legacy_id):
            raise ValueError("ENROLLMENT_LEGACY_LINK_CONFLICT")
        if legacy_id:
            enrollment.student_id = legacy_id
        academic_class = db.get(AcademicClass, row["academic_class_id"])
        if enrollment.academic_class_id != academic_class.id:
            enrollment.academic_class_id = academic_class.id
            enrollment.jenjang_id = row["jenjang_id"]
            enrollment.class_name = academic_class.class_name
            db.add(StudentEnrollmentClassHistory(
                enrollment_id=enrollment.id, class_name=academic_class.class_name,
                effective_from=row["academic_year_start"], changed_by=actor,
                source="dapodik_roster_import",
            ))
            action_type = "TRANSFER_ENROLLMENT"
        else:
            action_type = "SKIP_EXISTING_ENROLLMENT"
        enrollment.dapodik_registrasi_id = enrollment.dapodik_registrasi_id or payload.get("registrasi_id")
        enrollment.dapodik_anggota_rombel_id = enrollment.dapodik_anggota_rombel_id or payload.get("anggota_rombel_id")
        enrollment.dapodik_sekolah_id = enrollment.dapodik_sekolah_id or payload.get("school_id")
        enrollment.dapodik_semester_id = enrollment.dapodik_semester_id or row.get("semester_id")

    if action_type != "SKIP_EXISTING_ENROLLMENT":
        append_action(
            db, session, source_row=row["source_row"], sequence=sequence,
            action_type=action_type, entity_type="STUDENT_ENROLLMENT", entity_id=enrollment.id,
            actor=actor, before_state=None, after_state={"enrollment_id": enrollment.id, "academic_class_id": enrollment.academic_class_id},
            compensation_type="END_ENROLLMENT" if action_type == "CREATE_ENROLLMENT" else "MANUAL_REVIEW_REQUIRED",
            eligibility="ELIGIBLE" if action_type == "CREATE_ENROLLMENT" else "MANUAL_REVIEW_REQUIRED", roster_batch_id=roster_batch_id,
        )
        sequence += 1
    return sequence, master


def apply_import(
    file: str | Path, database: str | Path, academic_year: str, approved_rows: list[int],
    actor: str, source_owner: str, date_received: date, confirmation: str,
) -> dict[str, Any]:
    if confirmation != IMPORT_CONFIRMATION:
        raise ValueError("IMPORT_CONFIRMATION_REQUIRED")
    report = validate_import(file, database, academic_year, include_internal=True)
    internal_rows = {row["source_row"]: row for row in report.pop("_internal_rows")}
    report.pop("_school", None)
    selected = [internal_rows[row_id] for row_id in dict.fromkeys(approved_rows) if row_id in internal_rows]
    if not selected or len(selected) != len(set(approved_rows)):
        raise ValueError("APPROVED_ROWS_INVALID")
    blocked = [row["source_row"] for row in selected if row["classification"] not in {"NEW_STUDENT", "READY_TO_APPLY"}]
    if blocked:
        raise ValueError("APPROVED_ROWS_BLOCKED:" + ",".join(map(str, blocked)))
    database_path = _safe_database_path(database)
    if not student_import_schema_ready(database_path):
        raise ValueError("STUDENT_IMPORT_SCHEMA_REQUIRED")
    engine = create_engine(f"sqlite:///{database_path}")
    Session = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    session = Session()
    batch_key = hashlib.sha256(
        f"{report['source_checksum']}|{report['school_id']}|{academic_year}|{STUDENT_IMPORT_SCHEMA_VERSION}".encode()
    ).hexdigest()
    try:
        prior = session.query(StudentImportSession).filter_by(idempotency_key=batch_key).one_or_none()
        if prior and prior.status == "COMMITTED":
            prior_result = prior.session_metadata.get("result", {"status": "committed"})
            return dict(prior_result, idempotent_replay=True)
        import_session = StudentImportSession(
            import_type="STUDENT_ROSTER", status="COMMIT_PENDING", provenance_status="PROVENANCE_FAILED",
            created_by=actor, expires_at=datetime.now(timezone.utc), source_filename=report["source_filename"],
            source_file_checksum=report["source_checksum"], preview_checksum=report["report_checksum"],
            idempotency_key=batch_key, row_count=len(internal_rows), selected_row_count=len(selected),
            session_metadata={
                "source_system": SOURCE_SYSTEM, "schema_version": WORKBOOK_SCHEMA_VERSION,
                "school_id": report["school_id"], "academic_year": academic_year,
                "snapshot_completeness": "partial", "source_owner": source_owner,
            },
        )
        session.add(import_session); session.flush()
        import_batch = StudentImportBatch(
            session_id=import_session.id, filename=report["source_filename"],
            file_checksum=report["source_checksum"], source_sheet="Students",
            status="committing", total_rows=len(internal_rows),
            new_count=sum(row["classification"] == "NEW_STUDENT" for row in internal_rows.values()),
            update_count=sum(row["classification"] == "READY_TO_APPLY" for row in internal_rows.values()),
            conflict_count=sum(row["classification"] == "REVIEW_REQUIRED" for row in internal_rows.values()),
            invalid_count=sum(bool(row.get("errors")) for row in internal_rows.values()),
            created_by=actor,
        )
        session.add(import_batch); session.flush()
        batch = AcademicRosterImportBatch(
            session_id=import_session.id, filename=report["source_filename"], checksum=report["source_checksum"],
            source_owner=source_owner, date_received=date_received, created_by=actor,
            status="preview", rows=[_row_summary(row) for row in internal_rows.values()],
            summary={"source_system": SOURCE_SYSTEM, "selected_rows": len(selected), "report_checksum": report["report_checksum"]},
        )
        session.add(batch); session.flush()
        sequence = 1
        for row in selected:
            sequence, _ = _apply_row(
                session, row, actor, batch.id, import_batch.id, import_session, sequence,
            )
        result = {
            "status": "committed", "batch_id": batch.id, "session_id": import_session.session_uuid,
            "selected_rows": len(selected), "created_or_updated": len(selected), "idempotent_replay": False,
        }
        import_batch.status = "committed"; import_batch.committed_at = datetime.now(timezone.utc)
        batch.status = "committed"; batch.committed_by = actor; batch.committed_at = datetime.now(timezone.utc); batch.commit_result = result
        mark_committed(import_session, actor=actor, selected_count=len(selected), action_count=sequence - 1)
        import_session.session_metadata = dict(import_session.session_metadata or {}) | {"result": result}
        session.commit()
        return result
    except (ValueError, IntegrityError):
        session.rollback()
        raise
    finally:
        session.close(); engine.dispose()


def _write_report(report: dict[str, Any], path: Path | None) -> None:
    output = dict(report)
    output.pop("_internal_rows", None); output.pop("_school", None)
    text = json.dumps(output, default=str, sort_keys=True)
    if path:
        path.write_text(text + "\n", encoding="utf-8")
    else:
        print(text)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="python -m core.student_master_import")
    commands = parser.add_subparsers(dest="command", required=True)
    validate = commands.add_parser("validate")
    apply = commands.add_parser("apply")
    upgrade = commands.add_parser("upgrade-schema")
    for command in (validate, apply):
        command.add_argument("--file", required=True, type=Path)
        command.add_argument("--database", required=True, type=Path)
        command.add_argument("--academic-year", required=True)
        command.add_argument("--report", type=Path)
    apply.add_argument("--approve-row", action="append", type=int, default=[])
    apply.add_argument("--confirm-import", required=True)
    apply.add_argument("--actor", default="student-import-cli")
    apply.add_argument("--source-owner", required=True)
    apply.add_argument("--date-received", required=True, type=date.fromisoformat)
    upgrade.add_argument("--database", required=True, type=Path)
    try:
        args = parser.parse_args(argv)
        if args.command == "upgrade-schema":
            print(json.dumps({"status": "ok", "schema": ensure_student_import_schema(args.database)}))
            return 0
        if args.command == "validate":
            _write_report(validate_import(args.file, args.database, args.academic_year), args.report)
            return 0
        result = apply_import(
            args.file, args.database, args.academic_year, args.approve_row, args.actor,
            args.source_owner, args.date_received, args.confirm_import,
        )
        _write_report(result, args.report)
        return 0
    except Exception as error:
        print(json.dumps({"status": "REJECTED", "code": str(error)}), file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
