"""Validate and apply the Edelweiss employee workbook without leaking PII."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy import create_engine, func
from sqlalchemy.orm import Session, sessionmaker

EXPECTED_HEADERS = (
    "Id Staff", "STATUS", "Nama", "NIP", "NUPTK", "DAPODIK", "Tempat Lahir",
    "Tanggal Lahir", "Umur", "Jabatan", "Mulai Kerja", "Masa Kerja", "NIK",
    "Alamat", "Email", "No Hp",
)
IGNORED_DERIVED_COLUMNS = {"Umur", "Masa Kerja"}
ISSUE_CODES = {
    "DUPLICATE_NIK", "DUPLICATE_NIP", "DUPLICATE_SOURCE_STAFF_ID", "POSSIBLE_DUPLICATE_PERSON",
    "INVALID_NIP_LENGTH", "PLACEHOLDER_NIP", "INVALID_NIK_LENGTH", "INVALID_NUPTK_LENGTH",
    "MISSING_NIK", "MISSING_BIRTH_DATE", "MISSING_EMPLOYMENT_START_DATE", "MISSING_EMAIL",
    "MISSING_PHONE", "INVALID_EMAIL", "INVALID_PHONE", "UNKNOWN_EMPLOYMENT_STATUS",
    "UNKNOWN_DAPODIK_STATUS", "UNMAPPED_JOB_TITLE", "EXCEL_IDENTIFIER_PRECISION_RISK",
}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PHONE_RE = re.compile(r"^[+0-9() .\-/]{6,32}$")


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _raw_json(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _identifier(value: Any) -> tuple[str | None, bool]:
    """Return normalized identifier and whether an Excel precision risk exists."""
    if value is None:
        return None, False
    raw = str(value).strip()
    normalized = raw[1:] if raw.startswith("'") else raw
    risk = isinstance(value, (int, float)) and not isinstance(value, bool) and len(normalized.replace(".", "").replace("-", "")) > 15
    if isinstance(value, float) and value.is_integer():
        normalized = str(int(value))
    return normalized or None, risk


def _date_value(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), pattern).date()
        except ValueError:
            pass
    return None


def _status(raw: str | None) -> tuple[str, str | None]:
    value = (raw or "").upper()
    if value == "AKTIF":
        return "ACTIVE", None
    if value == "KELUAR":
        return "FORMER", None
    return "REVIEW_REQUIRED", "UNKNOWN_EMPLOYMENT_STATUS"


def _dapodik(raw: str | None) -> tuple[str, str | None]:
    value = (raw or "").strip().upper()
    if value == "AKTIF":
        return "ACTIVE", None
    if value == "BELUM":
        return "NOT_REGISTERED", None
    if value == "SUDAH":
        return "SUBMITTED_OR_COMPLETED", None
    if value in {"", "TIDAK"}:
        return "UNKNOWN", "UNKNOWN_DAPODIK_STATUS"
    return "UNKNOWN", "UNKNOWN_DAPODIK_STATUS"


def _issue(code: str, field: str | None, severity: str = "WARNING") -> dict[str, str]:
    return {"code": code, "field": field or "", "severity": severity, "message": f"{code.replace('_', ' ').lower()} requires review"}


def _row_from_values(row_number: int, values: list[Any]) -> dict[str, Any]:
    raw = dict(zip(EXPECTED_HEADERS, (_raw_json(v) for v in values)))
    raw_status = _clean(values[1])
    source_staff_id, risk_staff_id = _identifier(values[0])
    nip, risk_nip = _identifier(values[3])
    nuptk, risk_nuptk = _identifier(values[4])
    nik, risk_nik = _identifier(values[12])
    employment_status, employment_issue = _status(raw_status)
    dapodik_status, dapodik_issue = _dapodik(_clean(values[5]))
    birth_date = _date_value(values[7])
    start_date = _date_value(values[10])
    full_name = _clean(values[2]) or ""
    job_title = _clean(values[9])
    email = _clean(values[14])
    phone = _clean(values[15])
    normalized = {
        "source_staff_id": source_staff_id,
        "full_name": full_name,
        "normalized_name": " ".join(full_name.casefold().split()),
        "employment_status": employment_status,
        "birth_place": _clean(values[6]),
        "birth_date": birth_date.isoformat() if birth_date else None,
        "job_title_raw": job_title,
        "job_title_normalized": job_title,
        "employment_start_date": start_date.isoformat() if start_date else None,
        "dapodik_status_raw": _clean(values[5]),
        "dapodik_status_normalized": dapodik_status,
        "nip": nip, "nuptk": nuptk, "nik": nik,
        "address": _clean(values[13]), "email": email, "phone": phone,
    }
    issues: list[dict[str, str]] = []
    if employment_issue: issues.append(_issue(employment_issue, "STATUS", "ERROR"))
    if dapodik_issue: issues.append(_issue(dapodik_issue, "DAPODIK"))
    if risk_staff_id or risk_nip or risk_nuptk or risk_nik:
        issues.append(_issue("EXCEL_IDENTIFIER_PRECISION_RISK", "identifier", "ERROR"))
    if nip in {"0", "202"}:
        issues.append(_issue("PLACEHOLDER_NIP", "NIP", "ERROR"))
    elif nip and len(re.sub(r"\D", "", nip)) != 18:
        issues.append(_issue("INVALID_NIP_LENGTH", "NIP", "ERROR"))
    if nuptk and len(re.sub(r"\D", "", nuptk)) != 16:
        issues.append(_issue("INVALID_NUPTK_LENGTH", "NUPTK", "ERROR"))
    if nik and len(re.sub(r"\D", "", nik)) != 16:
        issues.append(_issue("INVALID_NIK_LENGTH", "NIK", "ERROR"))
    if not nik: issues.append(_issue("MISSING_NIK", "NIK"))
    if not birth_date: issues.append(_issue("MISSING_BIRTH_DATE", "Tanggal Lahir"))
    if not start_date: issues.append(_issue("MISSING_EMPLOYMENT_START_DATE", "Mulai Kerja"))
    if not email: issues.append(_issue("MISSING_EMAIL", "Email"))
    elif not EMAIL_RE.match(email): issues.append(_issue("INVALID_EMAIL", "Email", "ERROR"))
    if not phone: issues.append(_issue("MISSING_PHONE", "No Hp"))
    elif not PHONE_RE.match(phone): issues.append(_issue("INVALID_PHONE", "No Hp", "ERROR"))
    if job_title: issues.append(_issue("UNMAPPED_JOB_TITLE", "Jabatan"))
    return {
        "source_row_number": row_number, "raw": raw, "normalized": normalized,
        "issues": issues, "row_status": "REVIEW_REQUIRED" if any(i["severity"] == "ERROR" for i in issues) else ("ACCEPTED_WITH_WARNINGS" if issues else "ACCEPTED"),
    }


def parse_workbook(file: str | Path, sheet: str) -> dict[str, Any]:
    path = Path(file)
    if path.suffix.lower() != ".xlsx":
        raise ValueError("STAFF_SOURCE_MUST_BE_XLSX")
    if not path.is_file():
        raise ValueError("STAFF_SOURCE_NOT_FOUND")
    file_sha256 = hashlib.sha256(path.read_bytes()).hexdigest()
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError("STAFF_SOURCE_SHEET_NOT_FOUND")
        worksheet = workbook[sheet]
        header = tuple(worksheet.cell(1, column).value for column in range(1, len(EXPECTED_HEADERS) + 1))
        if header != EXPECTED_HEADERS:
            raise ValueError("STAFF_SOURCE_HEADERS_INVALID")
        rows = []
        for row_number, cells in enumerate(worksheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True), start=2):
            if all(value is None for value in cells):
                continue
            rows.append(_row_from_values(row_number, list(cells)))
    finally:
        workbook.close()
    for field, code in (("nik", "DUPLICATE_NIK"), ("nip", "DUPLICATE_NIP"), ("source_staff_id", "DUPLICATE_SOURCE_STAFF_ID")):
        buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            value = row["normalized"].get(field)
            if value:
                buckets[value].append(row)
        for bucket in buckets.values():
            if len(bucket) > 1:
                for row in bucket:
                    row["issues"].append(_issue(code, field, "ERROR"))
                    row["row_status"] = "REVIEW_REQUIRED"
    names: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if row["normalized"]["normalized_name"]:
            names[row["normalized"]["normalized_name"]].append(row)
    for bucket in names.values():
        if len(bucket) > 1:
            for row in bucket:
                row["issues"].append(_issue("POSSIBLE_DUPLICATE_PERSON", "Nama"))
                if row["row_status"] == "ACCEPTED": row["row_status"] = "ACCEPTED_WITH_WARNINGS"
    counts = Counter(row["row_status"] for row in rows)
    issue_counts = Counter(issue["code"] for row in rows for issue in row["issues"])
    return {
        "file_sha256": file_sha256,
        "source_filename": path.name,
        "source_sheet": sheet,
        "total_rows": len(rows),
        "active_count": sum(row["normalized"]["employment_status"] == "ACTIVE" for row in rows),
        "former_count": sum(row["normalized"]["employment_status"] == "FORMER" for row in rows),
        "review_count": sum(row["row_status"] in {"REVIEW_REQUIRED", "CONFLICT"} for row in rows),
        "issue_count": sum(issue_counts.values()),
        "row_status_counts": dict(counts),
        "issue_counts": dict(sorted(issue_counts.items())),
        "rows": rows,
    }


def redacted_summary(result: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in result.items() if key != "rows"}


def _apply(result: dict[str, Any], database: str | Path, actor: str) -> dict[str, Any]:
    database_path = Path(database)
    if not database_path.is_absolute():
        raise ValueError("STAFF_DATABASE_PATH_MUST_BE_ABSOLUTE")
    # The target is explicit and authoritative for this one-shot process.  Do
    # not allow an inherited DATABASE_URL to select a different database.
    os.environ["DATABASE_URL"] = f"sqlite:///{database_path}"
    from core.staff_schema_migration import ensure_staff_schema
    from models.staff import (
        StaffContactDetail, StaffIdentifier, StaffImportBatch, StaffImportIssue,
        StaffImportRow, StaffMember,
    )

    ensure_staff_schema(database_path)
    engine = create_engine(f"sqlite:///{database_path}")
    SessionLocal = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    try:
        with SessionLocal.begin() as session:
            if session.query(StaffImportBatch).filter_by(file_sha256=result["file_sha256"]).first():
                raise ValueError("SKIP_DUPLICATE_BATCH")
            batch = StaffImportBatch(
                id=str(uuid.uuid4()), source_filename=result["source_filename"], source_sheet=result["source_sheet"],
                file_sha256=result["file_sha256"], actor=actor, total_rows=result["total_rows"],
                active_count=result["active_count"], former_count=result["former_count"],
                review_count=result["review_count"], issue_count=result["issue_count"],
                status="REVIEW_REQUIRED" if result["review_count"] else "APPLIED",
            )
            session.add(batch)
            session.flush()
            for item in result["rows"]:
                normalized = item["normalized"]
                member = StaffMember(
                    source_staff_id=normalized["source_staff_id"], full_name=normalized["full_name"], normalized_name=normalized["normalized_name"],
                    employment_status=normalized["employment_status"], birth_place=normalized["birth_place"],
                    birth_date=date.fromisoformat(normalized["birth_date"]) if normalized["birth_date"] else None,
                    job_title_raw=normalized["job_title_raw"], job_title_normalized=normalized["job_title_normalized"],
                    employment_start_date=date.fromisoformat(normalized["employment_start_date"]) if normalized["employment_start_date"] else None,
                    dapodik_status_raw=normalized["dapodik_status_raw"], dapodik_status_normalized=normalized["dapodik_status_normalized"],
                )
                session.add(member)
                session.flush()
                for kind in ("INTERNAL_STAFF_ID", "NIP", "NUPTK", "NIK"):
                    field = {"INTERNAL_STAFF_ID": "source_staff_id", "NIP": "nip", "NUPTK": "nuptk", "NIK": "nik"}[kind]
                    value = normalized.get(field)
                    if value:
                        session.add(StaffIdentifier(staff_member_id=member.id, identifier_type=kind, raw_value=item["raw"].get({"source_staff_id": "Id Staff", "nip": "NIP", "nuptk": "NUPTK", "nik": "NIK"}[field]), normalized_value=value, verification_status="REVIEW_REQUIRED" if any(i["code"] in {"EXCEL_IDENTIFIER_PRECISION_RISK", "INVALID_NIP_LENGTH", "INVALID_NUPTK_LENGTH", "INVALID_NIK_LENGTH", "PLACEHOLDER_NIP"} for i in item["issues"]) else "VALIDATED"))
                session.add(StaffContactDetail(staff_member_id=member.id, address=normalized["address"], email=normalized["email"], phone=normalized["phone"]))
                import_row = StaffImportRow(batch_id=batch.id, source_row_number=item["source_row_number"], source_staff_id=normalized["source_staff_id"], staff_member_id=member.id, raw_payload_json=item["raw"], normalized_payload_json=normalized, row_status=item["row_status"])
                session.add(import_row)
                session.flush()
                for problem in item["issues"]:
                    session.add(StaffImportIssue(batch_id=batch.id, import_row_id=import_row.id, issue_code=problem["code"], field_name=problem["field"], severity=problem["severity"], message=problem["message"]))
        return redacted_summary(result) | {"status": batch.status, "batch_id": batch.id}
    finally:
        engine.dispose()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="python -m core.staff_import")
    sub = parser.add_subparsers(dest="command", required=True)
    for command in ("validate", "apply"):
        command_parser = sub.add_parser(command)
        command_parser.add_argument("--file", required=True, type=Path)
        command_parser.add_argument("--sheet", required=True)
    sub.choices["apply"].add_argument("--database", required=True, type=Path)
    sub.choices["apply"].add_argument("--confirm-import", action="store_true")
    sub.choices["apply"].add_argument("--actor", default="staff-import-cli")
    args = parser.parse_args(argv)
    try:
        result = parse_workbook(args.file, args.sheet)
        if args.command == "validate":
            print(json.dumps({"EMPLOYEE_IMPORT_VALIDATION": redacted_summary(result)}, sort_keys=True))
            return 0
        if not args.confirm_import:
            raise ValueError("CONFIRM_IMPORT_REQUIRED")
        print(json.dumps({"EMPLOYEE_IMPORT_APPLY": _apply(result, args.database, args.actor)}, sort_keys=True))
        return 0
    except Exception as error:
        print(json.dumps({"status": "REJECTED", "code": str(error)}), file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
