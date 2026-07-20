from __future__ import annotations

import hashlib
import io
import re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from models.student_master import StudentAddress, StudentContact, StudentDeviceIdentity, StudentMaster, StudentParentGuardian
from services.operations_audit_service import increment_counter, log_operations_audit_event

APPROVED_SCOPES = {
    "SELECTED_STUDENTS",
    "FILTERED_RESULTS",
    "ACADEMIC_CLASS",
    "ACADEMIC_YEAR",
    "ALL_PERMITTED_STUDENTS",
}

APPROVED_FIELD_PROFILES = {
    "STANDARD_OPERATIONAL",
    "SENSITIVE_IDENTIFIERS",
    "CONTACT_AND_GUARDIAN",
}

MAX_EXPORT_ROWS = 5000


def sanitize_filename(filename: str) -> str:
    cleaned = re.sub(r"[^\w\-. ]", "_", filename)
    return cleaned[:128]


def build_student_export_query(
    db: Session,
    *,
    scope: str,
    filters: Optional[Dict[str, Any]] = None,
    selected_student_ids: Optional[List[str]] = None,
):
    query = db.query(StudentMaster)
    filters = filters or {}

    if scope == "SELECTED_STUDENTS":
        if not selected_student_ids:
            raise HTTPException(status_code=400, detail="SELECTED_STUDENTS scope requires selected_student_ids")
        query = query.filter(StudentMaster.id.in_(selected_student_ids))
    elif scope in ("FILTERED_RESULTS", "ACADEMIC_CLASS", "ACADEMIC_YEAR"):
        if filters.get("status"):
            query = query.filter(StudentMaster.student_status == filters["status"])
        if filters.get("gender"):
            query = query.filter(StudentMaster.gender == filters["gender"])
        if filters.get("search"):
            search_term = f"%{filters['search'].strip()}%"
            query = query.filter(
                (StudentMaster.full_name.ilike(search_term)) | (StudentMaster.normalized_name.ilike(search_term))
            )
    elif scope == "ALL_PERMITTED_STUDENTS":
        pass  # Query all students
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported export scope: {scope}")

    return query


def generate_export_preview(
    db: Session,
    *,
    scope: str,
    field_profile: str,
    filters: Optional[Dict[str, Any]] = None,
    selected_student_ids: Optional[List[str]] = None,
    actor: str,
    actor_capabilities: set[str],
) -> Dict[str, Any]:
    if scope not in APPROVED_SCOPES:
        increment_counter("export_preview_failures")
        raise HTTPException(status_code=400, detail=f"Unrecognized export scope: {scope}")
    if field_profile not in APPROVED_FIELD_PROFILES:
        increment_counter("export_preview_failures")
        raise HTTPException(status_code=400, detail=f"Unrecognized field profile: {field_profile}")

    query = build_student_export_query(db, scope=scope, filters=filters, selected_student_ids=selected_student_ids)
    estimated_count = query.count()

    is_sensitive = field_profile in ("SENSITIVE_IDENTIFIERS", "CONTACT_AND_GUARDIAN")
    required_cap = "export_sensitive_student_fields" if is_sensitive else "export_student_data"

    allowed = required_cap in actor_capabilities
    warnings = []

    if estimated_count == 0:
        warnings.append("No student records match the export criteria.")
    elif estimated_count > MAX_EXPORT_ROWS:
        allowed = False
        warnings.append(f"Export size ({estimated_count} rows) exceeds maximum allowed threshold of {MAX_EXPORT_ROWS} rows.")

    if not allowed and is_sensitive and required_cap not in actor_capabilities:
        warnings.append("Elevated capability export_sensitive_student_fields is required for sensitive fields.")

    payload = {
        "scope": scope,
        "field_profile": field_profile,
        "filters": filters or {},
        "estimated_count": estimated_count,
    }
    preview_checksum = hashlib.sha256(f"{scope}:{field_profile}:{estimated_count}:{hashlib.sha256(str(filters).encode()).hexdigest()[:16]}".encode()).hexdigest()

    log_operations_audit_event(
        db,
        actor_id=actor,
        actor_role="admin" if "export_sensitive_student_fields" in actor_capabilities else "staff",
        capability=required_cap,
        entity_type="STUDENT_EXPORT",
        entity_reference=f"EXPORT_PREVIEW_{scope}",
        operation="EXPORT_PREVIEW",
        risk_level="MEDIUM" if is_sensitive else "LOW",
        export_scope=scope,
        success=allowed,
        failure_code=None if allowed else "EXPORT_PREVIEW_DENIED",
        metadata={
            "estimated_count": estimated_count,
            "sensitive": is_sensitive,
            "field_profile": field_profile,
            "preview_checksum": preview_checksum,
        },
    )

    return {
        "normalized_scope": scope,
        "field_profile": field_profile,
        "estimated_row_count": estimated_count,
        "sensitive_field_indicator": is_sensitive,
        "required_capability": required_cap,
        "allowed": allowed,
        "warnings": warnings,
        "maximum_permitted_row_count": MAX_EXPORT_ROWS,
        "filter_summary": filters or {},
        "preview_checksum": preview_checksum,
        "expiration": (datetime.now() + timedelta(minutes=30)).isoformat(),
    }


def execute_student_export(
    db: Session,
    *,
    scope: str,
    field_profile: str,
    filters: Optional[Dict[str, Any]] = None,
    selected_student_ids: Optional[List[str]] = None,
    preview_checksum: Optional[str] = None,
    actor: str,
    actor_capabilities: set[str],
) -> StreamingResponse:
    if scope not in APPROVED_SCOPES or field_profile not in APPROVED_FIELD_PROFILES:
        increment_counter("export_failures")
        raise HTTPException(status_code=400, detail="Invalid export parameters")

    is_sensitive = field_profile in ("SENSITIVE_IDENTIFIERS", "CONTACT_AND_GUARDIAN")
    required_cap = "export_sensitive_student_fields" if is_sensitive else "export_student_data"

    if required_cap not in actor_capabilities:
        increment_counter("export_failures")
        log_operations_audit_event(
            db,
            actor_id=actor,
            actor_role="staff",
            capability=required_cap,
            entity_type="STUDENT_EXPORT",
            entity_reference=f"EXPORT_{scope}",
            operation="EXPORT_DOWNLOAD",
            risk_level="HIGH" if is_sensitive else "MEDIUM",
            export_scope=scope,
            success=False,
            failure_code="PERMISSION_DENIED",
            metadata={"reason": "Missing required capability for export"},
        )
        raise HTTPException(status_code=403, detail=f"Permission denied: missing capability {required_cap}")

    query = build_student_export_query(db, scope=scope, filters=filters, selected_student_ids=selected_student_ids)
    students = query.limit(MAX_EXPORT_ROWS + 1).all()

    if not students:
        increment_counter("export_failures")
        raise HTTPException(status_code=400, detail="Cannot generate empty export. No matching records found.")

    if len(students) > MAX_EXPORT_ROWS:
        increment_counter("export_failures")
        raise HTTPException(status_code=400, detail=f"Export row count exceeds maximum limit of {MAX_EXPORT_ROWS} rows.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Records"

    # Define backend-owned columns based on profile
    if field_profile == "STANDARD_OPERATIONAL":
        headers = ["ID", "Nama Lengkap", "Status", "Jenis Kelamin", "Tempat Lahir", "Tanggal Lahir", "Agama"]
    elif field_profile == "SENSITIVE_IDENTIFIERS":
        headers = ["ID", "Nama Lengkap", "Status", "NIK", "NISN", "NIPD", "Jenis Kelamin", "Device ID Active"]
    else:  # CONTACT_AND_GUARDIAN
        headers = ["ID", "Nama Lengkap", "Status", "Alamat", "No Telp / HP", "Nama Wali", "No HP Wali"]

    ws.append(headers)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for student in students:
        if field_profile == "STANDARD_OPERATIONAL":
            row = [
                student.id,
                student.full_name,
                student.student_status,
                student.gender or "-",
                student.birth_place or "-",
                student.birth_date.strftime("%Y-%m-%d") if student.birth_date else "-",
                student.religion or "-",
            ]
        elif field_profile == "SENSITIVE_IDENTIFIERS":
            active_device = (
                db.query(StudentDeviceIdentity)
                .filter(StudentDeviceIdentity.student_master_id == student.id, StudentDeviceIdentity.is_active.is_(True))
                .first()
            )
            device_str = active_device.device_identifier if active_device else "-"
            # Formatted as text explicitly by converting to string
            row = [
                student.id,
                student.full_name,
                student.student_status,
                str(student.nik or "-"),
                str(student.nisn or "-"),
                str(student.nipd or "-"),
                student.gender or "-",
                str(device_str),
            ]
        else:
            addr = db.query(StudentAddress).filter(StudentAddress.student_master_id == student.id).first()
            contact = db.query(StudentContact).filter(StudentContact.student_master_id == student.id).first()
            guardian = db.query(StudentParentGuardian).filter(StudentParentGuardian.student_master_id == student.id).first()
            row = [
                student.id,
                student.full_name,
                student.student_status,
                addr.street_address if addr else "-",
                str(contact.phone_number if contact else "-"),
                guardian.guardian_name if guardian else "-",
                str(guardian.guardian_phone if guardian else "-"),
            ]
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = sanitize_filename(f"student_export_{scope.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    log_operations_audit_event(
        db,
        actor_id=actor,
        actor_role="admin",
        capability=required_cap,
        entity_type="STUDENT_EXPORT",
        entity_reference=f"EXPORT_FILE_{filename}",
        operation="EXPORT_DOWNLOAD",
        risk_level="HIGH" if is_sensitive else "MEDIUM",
        export_scope=scope,
        success=True,
        metadata={
            "actual_row_count": len(students),
            "sensitive": is_sensitive,
            "field_profile": field_profile,
            "filename": filename,
        },
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
