from __future__ import annotations

import os
import tempfile
import uuid
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from services.report_export import DEFAULT_BRANDING


def _rate(value) -> str:
    return "N/A" if value is None else f"{value:.1f}%"


def _table(rows: list[list], widths=None) -> Table:
    table = Table(rows, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def build_monthly_management_pdf(report: dict, branding: dict | None = None) -> bytes:
    branding = branding or DEFAULT_BRANDING

    temp_dir_obj = tempfile.TemporaryDirectory()
    temp_pdf_path = os.path.join(temp_dir_obj.name, f"monthly_management_{uuid.uuid4().hex}.pdf")

    document = SimpleDocTemplate(
        temp_pdf_path, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title="Monthly Management Report", author=branding["school_name"], pageCompression=0,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle("ManagementTitle", parent=styles["Title"], fontSize=18, leading=22, textColor=colors.HexColor("#1E3A8A"))
    section = ParagraphStyle("ManagementSection", parent=styles["Heading2"], fontSize=12, leading=15, spaceBefore=9, spaceAfter=5)
    small = ParagraphStyle("ManagementSmall", parent=styles["BodyText"], fontSize=8, leading=11)
    meta = report["metadata"]
    period = report["report_period"]
    sections = period["sections"]
    story = [
        Paragraph(branding["school_name"], title),
        Paragraph("Monthly Management Report", styles["Heading1"]),
        Paragraph(
            f"Academic Year: {meta['academic_year']['name']} | Month: {period['selected_month']} | Scope: {meta['scope']} | Generated UTC: {meta['generated_at']}", small
        ),
        Spacer(1, 4 * mm),
        Paragraph("Reporting Basis", section),
        _table([["Section", "Basis", "Label"],
                ["Attendance", sections["attendance"]["basis"], sections["attendance"]["label"]],
                ["Population", sections["population"]["basis"], sections["population"]["label"]],
                ["Academics", sections["academics"]["basis"], sections["academics"]["label"]]], [34 * mm, 58 * mm, 88 * mm]),
        Paragraph("Executive Summary", section),
    ]
    summary = report["executive_summary"]
    story.append(_table([
        ["Students", "Classes", "Attendance", "Present", "Izin", "Sakit", "Alfa", "Late", "Below KKM"],
        [summary["total_students"], summary["total_classes"], _rate(summary["attendance_rate"]), summary["present_count"],
         summary["excused_absence_count"], summary["sick_count"], summary["unexcused_absence_count"],
         summary["late_count"], summary["students_below_kkm"]],
    ]))
    population = report["student_population"]
    level_rows = [["Jenjang", "Students", "% Eligible", "Classes"]] + [
        [row["jenjang"], row["student_count"], _rate(row["percentage_of_eligible"]), row["class_count"]]
        for row in population["by_jenjang"]
    ]
    story.extend([Paragraph("Academic-Year Enrollment Snapshot", section), _table(level_rows, [65 * mm, 30 * mm, 35 * mm, 30 * mm])])
    class_rows = [["Jenjang", "Class", "Students", "% Jenjang", "% Eligible"]] + [
        [row["jenjang"], row["class_name"], row["student_count"], _rate(row["percentage_within_jenjang"]), _rate(row["percentage_of_eligible"])]
        for row in population["by_class"]
    ]
    story.extend([Paragraph("Population by Class", section), _table(class_rows)])
    story.append(PageBreak())
    story.append(Paragraph("Demographics", section))
    for label, key in (("Religion", "religion"), ("Gender", "gender"), ("Residential Area (Kelurahan)", "residential_area")):
        group = report["demographics"][key]
        rows = [[label, "Count", "% Known", "% Eligible"]] + [
            [row["name"], row["count"], _rate(row["percentage_of_known"]), _rate(row["percentage_of_eligible"])]
            for row in group["rows"]
        ]
        if len(rows) == 1:
            rows.append(["No known values", 0, "N/A", "N/A"])
        story.extend([
            Paragraph(label, section),
            Paragraph(f"Eligible: {group['eligible_count']} | Known: {group['known_count']} | Unknown: {group['unknown_count']} | Denominator: {group['denominator_used']}", small),
            _table(rows),
        ])
    attendance = report["attendance"]["summary"]
    story.extend([
        Paragraph("Monthly Attendance", section),
        Paragraph(sections["attendance"]["label"], small),
        _table([["Present", "Sakit", "Izin", "Alfa", "Incomplete", "Late", "Attendance Rate"],
                [attendance["present"], attendance["sakit"], attendance["izin"], attendance["alfa"], attendance["incomplete"], attendance["late_days"], _rate(attendance["attendance_rate"])]]),
        Paragraph("Available Academic Performance Summary", section),
        Paragraph("Academic performance figures reflect available records for the selected academic year and are not restricted to the selected calendar month.", small),
    ])
    academic = report["academic_summary"]
    story.append(_table([["Available", "Sumatif Average", "Formatif Average", "Below KKM"],
                         ["Yes" if academic["availability"] else "No", academic["sumatif_average"] or "N/A", academic["formatif_average"] or "N/A", academic["below_kkm_count"]]]))
    story.extend([PageBreak(), Paragraph("Data Quality & Coverage", section)])
    quality_rows = [["Section", "Eligible", "Known", "Unknown", "Excluded", "Denominator", "Difference"]]
    for name, row in report["data_quality"]["sections"].items():
        quality_rows.append([name.replace("_", " ").title(), row["eligible_count"], row["known_count"], row["unknown_count"], row["excluded_count"], row["denominator_used"], row["reconciliation_difference"]])
    story.append(_table(quality_rows))
    for warning in report["data_quality"]["warnings"]:
        story.append(Paragraph(f"- {warning}", small))

    def footer(canvas, doc):
        canvas.saveState(); canvas.setFont("Helvetica", 7); canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(14 * mm, 8 * mm, branding["footer_text"])
        canvas.drawRightString(A4[0] - 14 * mm, 8 * mm, f"Page {doc.page}"); canvas.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    with open(temp_pdf_path, "rb") as f:
        data = f.read()
    temp_dir_obj.cleanup()
    return data


def _sheet(workbook: Workbook, name: str, headers: list[str], rows: list[list], primary: str) -> None:
    ws = workbook.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=primary.lstrip("#"))
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, ws.max_row)}"
    for index, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(index)].width = min(max(12, len(header) + 2, *(len(str(ws.cell(row, index).value or "")) + 2 for row in range(2, ws.max_row + 1))), 48)
        if "%" in header or "Rate" in header:
            for row in range(2, ws.max_row + 1):
                ws.cell(row, index).number_format = '0.0"%"'


def build_monthly_management_xlsx(report: dict, branding: dict | None = None) -> bytes:
    branding = branding or DEFAULT_BRANDING
    primary = branding["primary_color"]
    wb = Workbook(); wb.remove(wb.active)
    meta = report["metadata"]; period = report["report_period"]; summary = report["executive_summary"]
    _sheet(wb, "Executive Summary", ["Metric", "Value", "Basis"], [
        ["Report", meta["title"], period["selected_month"]], ["Academic Year", meta["academic_year"]["name"], period["sections"]["population"]["label"]],
        *[[key.replace("_", " ").title(), value, "Canonical management report model"] for key, value in summary.items()],
    ], primary)
    _sheet(wb, "Population by Jenjang", ["Jenjang", "Student Count", "% Eligible", "Class Count", "Classification"], [[r["jenjang"], r["student_count"], r["percentage_of_eligible"], r["class_count"], r["classification"]] for r in report["student_population"]["by_jenjang"]], primary)
    _sheet(wb, "Population by Class", ["Jenjang", "Class", "Student Count", "% Within Jenjang", "% Eligible"], [[r["jenjang"], r["class_name"], r["student_count"], r["percentage_within_jenjang"], r["percentage_of_eligible"]] for r in report["student_population"]["by_class"]], primary)
    for sheet_name, key in (("Religion", "religion"), ("Gender", "gender"), ("Residential Area", "residential_area")):
        group = report["demographics"][key]
        rows = [[r["name"], r["count"], r["percentage_of_known"], r["percentage_of_eligible"], group["eligible_count"], group["known_count"], group["unknown_count"], group["denominator_used"]] for r in group["rows"]]
        _sheet(wb, sheet_name, [sheet_name, "Count", "% Known", "% Eligible", "Eligible", "Known", "Unknown", "Denominator"], rows, primary)
    attendance = report["attendance"]["summary"]
    _sheet(wb, "Monthly Attendance", ["Level", "Present", "Sakit", "Izin", "Alfa", "Incomplete", "Late Days", "Late Minutes", "Attendance Rate", "Late Rate"], [["Overall", attendance["present"], attendance["sakit"], attendance["izin"], attendance["alfa"], attendance["incomplete"], attendance["late_days"], attendance["late_minutes"], attendance["attendance_rate"], attendance["late_rate"]], *[[r["level"], r["present"], r["sakit"], r["izin"], r["alfa"], r["incomplete"], r["late_days"], r["late_minutes"], r["attendance_rate"], r["late_rate"]] for r in report["attendance"]["by_jenjang"]]], primary)
    academic = report["academic_summary"]
    _sheet(wb, "Academic Summary", ["Subject", "Jenjang", "Sumatif Average", "Formatif Average", "Below KKM", "Time Basis"], [["Overall", "All", academic["sumatif_average"], academic["formatif_average"], academic["below_kkm_count"], period["sections"]["academics"]["label"]], *[[r["subject_name"], r["jenjang"], r["sumatif_average"], r["formatif_average"], r["below_kkm_count"], period["sections"]["academics"]["label"]] for r in academic["by_subject"]]], primary)
    _sheet(wb, "Data Quality", ["Section", "Eligible", "Known", "Unknown", "Excluded", "Denominator", "Percentage Basis", "Difference", "Reconciles"], [[name.replace("_", " ").title(), r["eligible_count"], r["known_count"], r["unknown_count"], r["excluded_count"], r["denominator_used"], r["percentage_basis"], r["reconciliation_difference"], r["reconciles"]] for name, r in report["data_quality"]["sections"].items()], primary)
    output = BytesIO(); wb.save(output); return output.getvalue()
