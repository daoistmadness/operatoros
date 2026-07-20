import hashlib
import json
from collections import Counter
from datetime import date, datetime, timedelta
from io import BytesIO

import pandas as pd
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from models.academic_master import AcademicClass, AcademicGrade, AcademicProgram
from models.academic_roster import AcademicRosterImportBatch
from models.academic_year import AcademicYear
from models.jenjang import Jenjang
from models.student_enrollment import StudentEnrollment
from models.student_master import StudentDeviceIdentity, StudentEnrollmentClassHistory, StudentMaster
from services.student_normalization import normalize_name
from services.student_management import _audit, _check_identifier_conflicts, _create_enrollment, _create_legacy_identity


ROSTER_CONFIRMATION = "COMMIT_ACADEMIC_ROSTER"
REQUIRED = {"student_identifier", "student_name", "academic_year", "jenjang", "class_name", "program", "status"}
OPTIONAL = {"student_master_id", "nipd", "nisn", "nik", "birth_date", "homeroom_teacher", "admission_type", "start_date"}


def roster_preview_checksum(rows: list[dict]) -> str:
    return hashlib.sha256(json.dumps(rows, sort_keys=True, separators=(",", ":")).encode()).hexdigest()


def roster_template() -> bytes:
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Roster"
    headers = sorted(REQUIRED | OPTIONAL, key=lambda value: (value not in ("student_identifier", "student_name"), value))
    sheet.append(headers); sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]: cell.font = Font(bold=True)
    for index, header in enumerate(headers, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = max(16, len(header) + 3)
        if header in {"student_identifier", "nipd", "nisn", "nik", "student_master_id"}:
            sheet.cell(2, index).number_format = "@"
    instructions = workbook.create_sheet("Instructions")
    instructions.append(["OperatorOS Student Roster"])
    instructions.append(["Required columns", ", ".join(sorted(REQUIRED))])
    instructions.append(["Workflow", "Upload creates a non-mutating preview. Select valid rows and confirm before commit."])
    output = BytesIO(); workbook.save(output); return output.getvalue()


def _text(value):
    return None if pd.isna(value) or not str(value).strip() else str(value).strip()


def _date(value):
    parsed = pd.to_datetime(value, errors="coerce")
    return parsed.date().isoformat() if not pd.isna(parsed) else None


def _match_master(db: Session, payload: dict) -> tuple[StudentMaster | None, str | None]:
    if payload.get("student_master_id"):
        master = db.get(StudentMaster, payload["student_master_id"])
        return master, "student_master_id" if master else None
    for key in ("nipd", "nisn", "nik"):
        value = payload.get(key)
        if value:
            matches = db.query(StudentMaster).filter(getattr(StudentMaster, key) == value).all()
            if len(matches) == 1:
                return matches[0], key
            if len(matches) > 1:
                return None, "ambiguous_identifier"
    identifier = payload.get("student_identifier")
    if identifier:
        mappings = db.query(StudentDeviceIdentity).filter(
            StudentDeviceIdentity.device_identifier == identifier,
            StudentDeviceIdentity.is_active.is_(True),
        ).all()
        if len(mappings) == 1:
            return db.get(StudentMaster, mappings[0].student_master_id), "approved_device_identity"
        if len(mappings) > 1:
            return None, "ambiguous_device_identity"
    if payload.get("birth_date"):
        matches = db.query(StudentMaster).filter(
            StudentMaster.normalized_name == normalize_name(payload["student_name"]),
            StudentMaster.birth_date == date.fromisoformat(payload["birth_date"]),
        ).all()
        if len(matches) == 1:
            return matches[0], "normalized_name_birth_date"
        if len(matches) > 1:
            return None, "ambiguous_name_birth_date"
    return None, None


def create_roster_preview(db: Session, file_bytes: bytes, filename: str, owner: str, received: date, username: str):
    workbook = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
    frames = []
    for sheet in workbook.sheet_names:
        frame = pd.read_excel(workbook, sheet_name=sheet, dtype=str)
        frame.columns = [str(column).strip().casefold().replace(" ", "_") for column in frame.columns]
        missing = REQUIRED - set(frame.columns)
        if missing:
            raise ValueError(f"Sheet '{sheet}' missing required columns: {', '.join(sorted(missing))}")
        frame["__sheet"] = sheet
        frame["__row"] = frame.index + 2
        frames.append(frame)
    source = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    canonical_jenjang = {row.name: row for row in db.query(Jenjang).filter(Jenjang.active.is_(True)).all()}
    rows = []
    seen_enrollment_keys = set()
    for _, raw in source.iterrows():
        payload = {key: _text(raw.get(key)) for key in REQUIRED | OPTIONAL}
        payload["birth_date"] = _date(raw.get("birth_date"))
        payload["start_date"] = _date(raw.get("start_date"))
        errors = []
        if not payload["student_name"] or not payload["student_identifier"]:
            classification = "INVALID"
            errors.append("Student identifier and name are required")
            master = None
            match_rule = None
        else:
            master, match_rule = _match_master(db, payload)
            year = db.query(AcademicYear).filter(AcademicYear.label == payload["academic_year"]).first()
            jenjang = canonical_jenjang.get(payload["jenjang"])
            program = db.query(AcademicProgram).filter_by(
                jenjang_id=jenjang.id if jenjang else None,
                name=payload["program"], active=True,
            ).first() if jenjang else None
            academic_class = (
                db.query(AcademicClass)
                .join(AcademicGrade, AcademicGrade.id == AcademicClass.grade_id)
                .filter(
                    AcademicClass.academic_year_id == year.id,
                    AcademicClass.class_name == payload["class_name"],
                    AcademicClass.active.is_(True),
                    AcademicGrade.jenjang_id == jenjang.id,
                    AcademicGrade.program_id == program.id,
                    AcademicGrade.active.is_(True),
                ).first()
            ) if year and jenjang and program else None
            if match_rule and match_rule.startswith("ambiguous"):
                classification = "POSSIBLE_DUPLICATE"; errors.append("Identity match is ambiguous")
            elif year is None:
                classification = "INVALID"; errors.append("Unknown academic year")
            elif jenjang is None:
                classification = "MISSING_JENJANG"; errors.append("Unknown canonical jenjang")
            elif program is None or academic_class is None:
                classification = "MISSING_CLASS"; errors.append("Program/class is not active approved master data")
            elif not payload["status"] or payload["status"].casefold() != "active":
                classification = "INVALID"; errors.append("Only active roster rows are committable")
            elif master is None and not payload["student_identifier"].isdigit():
                classification = "INVALID"; errors.append("New students require a numeric attendance device ID")
            else:
                key = ((master.id if master else f"new:{payload['student_identifier']}"), year.id)
                existing = db.query(StudentEnrollment).filter_by(student_master_id=master.id, academic_year_id=year.id).first() if master else None
                if existing or key in seen_enrollment_keys:
                    classification = "INVALID"; errors.append("Duplicate enrollment for academic year")
                else:
                    classification = "CREATE_ENROLLMENT" if master else "CREATE_NEW_MASTER"
                    seen_enrollment_keys.add(key)
                    payload.update({
                        "academic_year_id": year.id, "academic_year_start": year.start_date.isoformat(),
                        "jenjang_id": jenjang.id, "academic_class_id": academic_class.id,
                        "target_class": academic_class.class_name,
                    })
        rows.append({
            "preview_row_id": len(rows) + 1,
            "source_sheet": raw["__sheet"], "source_row": int(raw["__row"]),
            "classification": classification, "matched_student_master_id": master.id if master else None,
            "match_rule": match_rule, "payload": payload, "errors": errors,
        })
    counts = Counter(row["classification"] for row in rows)
    summary = {"total": len(rows), **{key.casefold(): counts[key] for key in ("CREATE_ENROLLMENT", "CREATE_NEW_MASTER", "POSSIBLE_DUPLICATE", "MISSING_JENJANG", "MISSING_CLASS", "INVALID")}}
    batch = AcademicRosterImportBatch(
        filename=filename, checksum=hashlib.sha256(file_bytes).hexdigest(), source_owner=owner,
        date_received=received, created_by=username, rows=rows, summary=summary,
    )
    db.add(batch); db.commit(); db.refresh(batch)
    return batch


def commit_roster_preview(db: Session, preview_id: str, selected_rows: list[int], confirmation: str, username: str, preview_checksum: str | None = None):
    if confirmation != ROSTER_CONFIRMATION:
        raise HTTPException(status_code=400, detail="Invalid confirmation token")
    batch = db.get(AcademicRosterImportBatch, preview_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Roster preview not found")
    if batch.status == "committed":
        return batch.commit_result
    if batch.created_at and datetime.utcnow() - batch.created_at > timedelta(hours=24):
        raise HTTPException(status_code=410, detail="Roster preview expired; upload the workbook again")
    if preview_checksum and preview_checksum != roster_preview_checksum(batch.rows):
        raise HTTPException(status_code=409, detail="Roster preview checksum changed")
    selected = [row for row in batch.rows if row["preview_row_id"] in set(selected_rows)]
    if not selected or len(selected) != len(set(selected_rows)):
        raise HTTPException(status_code=400, detail="Selected rows are not part of the preview")
    blocked = [row["preview_row_id"] for row in selected if row["classification"] not in {"CREATE_ENROLLMENT", "CREATE_NEW_MASTER"}]
    if blocked:
        raise HTTPException(status_code=409, detail=f"Selected rows are not committable: {blocked}")
    created = 0
    try:
        created_students = 0
        for row in selected:
            payload = row["payload"]
            effective_from = date.fromisoformat(
                payload.get("start_date") or payload["academic_year_start"]
            )
            master = db.get(StudentMaster, row["matched_student_master_id"]) if row["matched_student_master_id"] else None
            if row["classification"] == "CREATE_NEW_MASTER":
                identity = {"nipd": payload.get("nipd"), "nisn": payload.get("nisn"), "nik": payload.get("nik")}
                _check_identifier_conflicts(db, identity)
                if db.query(StudentDeviceIdentity).filter_by(device_identifier=payload["student_identifier"], is_active=True).first():
                    raise HTTPException(status_code=409, detail=f"Device identity changed after preview row {row['source_row']}")
                master = StudentMaster(
                    full_name=payload["student_name"], normalized_name=normalize_name(payload["student_name"]),
                    nipd=payload.get("nipd"), nisn=payload.get("nisn"), nik=payload.get("nik"),
                    birth_date=date.fromisoformat(payload["birth_date"]) if payload.get("birth_date") else None,
                    admission_type=payload.get("admission_type"), student_status="active",
                    created_by=username, updated_by=username,
                )
                db.add(master); db.flush()
                _create_legacy_identity(db, master, {
                    "device_identifier": payload["student_identifier"], "device_source": "attendance_machine",
                    "effective_from": effective_from, "actor": username,
                })
                db.flush()
                _audit(db, master.id, "student_created", username, "academic_roster_import")
                _audit(db, master.id, "device_identity_added", username, "academic_roster_import", "device_identifier", None, payload["student_identifier"])
                row["matched_student_master_id"] = master.id
                created_students += 1
            if db.query(StudentEnrollment).filter_by(student_master_id=master.id, academic_year_id=payload["academic_year_id"]).first():
                raise HTTPException(status_code=409, detail=f"Enrollment changed after preview row {row['source_row']}")
            _create_enrollment(db, master, {
                "academic_year_id": payload["academic_year_id"],
                "academic_class_id": payload["academic_class_id"],
                "effective_from": effective_from,
            }, username, "academic_roster_import")
            created += 1
        result = {"status": "committed", "preview_id": batch.id, "created": created, "students_created": created_students}
        batch.status = "committed"; batch.committed_by = username; batch.committed_at = datetime.now(); batch.commit_result = result
        db.commit(); return result
    except Exception:
        db.rollback(); raise
