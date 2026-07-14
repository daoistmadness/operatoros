from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from models.report_builder import ReportBrandingConfig


DEFAULT_BRANDING = {
    "school_name": "EDELWEISS SCHOOL",
    "report_header_title": "Executive Report",
    "report_subtitle": "Attendance and Academic Analytics",
    "primary_color": "#1E3A8A",
    "secondary_color": "#0F172A",
    "accent_color": "#D97706",
    "footer_text": "School Attendance Analytics",
}


def get_report_branding(db: Session) -> dict:
    row = (
        db.query(ReportBrandingConfig)
        .filter(ReportBrandingConfig.is_default.is_(True))
        .order_by(ReportBrandingConfig.id.asc())
        .first()
    )
    if row is None:
        return dict(DEFAULT_BRANDING)
    return {
        "school_name": row.school_name,
        "report_header_title": row.report_header_title,
        "report_subtitle": row.report_subtitle,
        "primary_color": row.primary_color,
        "secondary_color": row.secondary_color,
        "accent_color": row.accent_color,
        "footer_text": row.footer_text,
    }


def _display(value) -> str:
    if value is None:
        return "N/A"
    if isinstance(value, float):
        return f"{value:.1f}"
    return str(value)


def _rate(value) -> str:
    return "N/A" if value is None else f"{value:.1f}%"


def _table(rows: list[list], header: bool = True, widths=None) -> Table:
    table = Table(rows, colWidths=widths, repeatRows=1 if header else 0, hAlign="LEFT")
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1 if header else 0), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if header:
        commands.extend([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ])
    table.setStyle(TableStyle(commands))
    return table


def build_report_pdf(report: dict, branding: dict | None = None) -> bytes:
    branding = branding or DEFAULT_BRANDING
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"{report['meta']['report_type'].title()} Executive Report",
        author=branding.get("school_name", DEFAULT_BRANDING["school_name"]),
        pageCompression=0,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18,
        textColor=colors.HexColor(branding.get("primary_color", "#1E3A8A")), alignment=TA_CENTER,
        spaceAfter=4,
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12,
        textColor=colors.HexColor(branding.get("secondary_color", "#0F172A")), spaceBefore=8, spaceAfter=5,
    )
    small = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=8, leading=10)
    meta = report["meta"]
    period = meta["period"]
    story = [
        Paragraph(branding.get("school_name", "EDELWEISS SCHOOL"), title_style),
        Paragraph(f"{meta['report_type'].title()} Executive Report", styles["Heading1"]),
        Paragraph(
            f"Scope: {meta['scope']} | Academic Year: {meta['academic_year']['name']} | "
            f"Period: {period['start']} to {period['end']} | Generated UTC: {meta['generated_at']}",
            small,
        ),
        Spacer(1, 5 * mm),
        Paragraph("Executive Summary", section_style),
    ]
    executive = report["executive_summary"]
    story.append(_table([
        ["Total Students", "Attendance Rate", "Late Rate", "Late Minutes", "Below KKM", "Data Completeness"],
        [executive["total_students"], _rate(executive["attendance_rate"]), _rate(executive["late_rate"]),
         executive["late_minutes"], executive["below_kkm_count"], _rate(executive["data_completeness_rate"])],
    ], widths=[35 * mm] * 6))

    distribution = report["student_distribution"]
    story.extend([Paragraph("Student Distribution", section_style)])
    dist_rows = [["Dimension", "Name", "Count", "Percentage"]]
    for dimension, key in (("Level", "by_level"), ("Class", "by_class")):
        for row in distribution[key]:
            dist_rows.append([dimension, row["name"], row["count"], _rate(row["percentage"])])
    if len(dist_rows) == 1:
        dist_rows.append(["-", "No population data", 0, "N/A"])
    story.append(_table(dist_rows, widths=[35 * mm, 75 * mm, 30 * mm, 35 * mm]))
    story.append(Paragraph("Demographic distributions are unavailable in the current Student master schema.", small))

    attendance = report["attendance_summary"]
    story.extend([Paragraph("Attendance Summary", section_style), _table([
        ["Present", "Sakit", "Izin", "Alfa", "Incomplete", "Late Days", "Late Minutes", "Attendance Rate", "Late Rate"],
        [attendance["present"], attendance["sakit"], attendance["izin"], attendance["alfa"],
         attendance["incomplete"], attendance["late_days"], attendance["late_minutes"],
         _rate(attendance["attendance_rate"]), _rate(attendance["late_rate"])],
    ], widths=[25 * mm] * 9)])
    level_rows = [["Level", "Present", "Sakit", "Izin", "Alfa", "Incomplete", "Late Days", "Late Minutes", "Attendance Rate", "Late Rate"]]
    for row in report["attendance_by_level"]:
        level_rows.append([row["level"], row["present"], row["sakit"], row["izin"], row["alfa"], row["incomplete"],
                           row["late_days"], row["late_minutes"], _rate(row["attendance_rate"]), _rate(row["late_rate"])])
    if len(level_rows) == 1:
        level_rows.append(["No level data", 0, 0, 0, 0, 0, 0, 0, "N/A", "N/A"])
    story.extend([Paragraph("Attendance by Level", section_style), _table(level_rows)])

    academic = report["academic_summary"]
    story.extend([PageBreak(), Paragraph("Academic Summary", section_style), _table([
        ["Availability", "Sumatif Average", "Formatif Average", "Below KKM Count"],
        ["Available" if academic["availability"] else "Unavailable", _display(academic["sumatif_average"]),
         _display(academic["formatif_average"]), academic["below_kkm_count"]],
    ], widths=[50 * mm] * 4)])
    if academic.get("reason"):
        story.append(Paragraph(academic["reason"], small))
    subject_rows = [["Subject", "Level", "Sumatif Average", "Formatif Average", "Below KKM Count"]]
    for row in academic["by_subject"]:
        subject_rows.append([row["subject_name"], row["jenjang"], _display(row["sumatif_average"]),
                             _display(row["formatif_average"]), row["below_kkm_count"]])
    if len(subject_rows) == 1:
        subject_rows.append(["No subject data", "-", "N/A", "N/A", 0])
    story.append(_table(subject_rows, widths=[65 * mm, 45 * mm, 40 * mm, 40 * mm, 35 * mm]))

    if meta["report_type"] == "annual":
        story.extend([Paragraph("Monthly Trends", section_style)])
        trend_rows = [["Month", "Present", "Sakit", "Izin", "Alfa", "Incomplete", "Denominator", "Attendance Rate",
                       "Late Days", "Late Minutes", "Late Rate", "Sumatif", "Formatif", "Below KKM"]]
        for row in report["trends"]:
            trend_rows.append([row["month"], row["present"], row["sakit"], row["izin"], row["alfa"], row["incomplete"],
                               row["attendance_denominator"], _rate(row["attendance_rate"]), row["late_days"],
                               row["late_minutes"], _rate(row["late_rate"]), _display(row["sumatif_average"]),
                               _display(row["formatif_average"]), row["below_kkm_count"]])
        story.append(_table(trend_rows))
        comparisons = report["comparisons"]
        comparison_rows = [["Comparison", "Name", "Attendance Rate", "Denominator"]]
        for label, key in (("Highest attendance month", "highest_attendance_month"),
                           ("Lowest attendance month", "lowest_attendance_month"),
                           ("Highest attendance level", "highest_attendance_level"),
                           ("Lowest attendance level", "lowest_attendance_level")):
            row = comparisons[key]
            comparison_rows.append([label, row["name"] if row else "N/A", _rate(row["attendance_rate"]) if row else "N/A",
                                    row["attendance_denominator"] if row else "N/A"])
        story.append(KeepTogether([Paragraph("Annual Comparisons", section_style), _table(comparison_rows)]))

    quality = report["data_quality"]
    story.extend([Paragraph("Data Quality", section_style), _table([
        ["Missing Gender", "Missing Religion", "Missing Domicile", "Incomplete Attendance", "Empty Grade Cells", "Unmapped Levels"],
        [quality["missing_gender"], quality["missing_religion"], quality["missing_domicile"], quality["incomplete_attendance"],
         quality["empty_grade_cells"], ", ".join(quality["unmapped_levels"]) or "None"],
    ])])
    for warning in quality["warnings"]:
        story.append(Paragraph(f"- {warning}", small))

    def page_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(14 * mm, 8 * mm, branding.get("footer_text", "School Attendance Analytics"))
        canvas.drawRightString(landscape(A4)[0] - 14 * mm, 8 * mm, f"Page {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=page_footer, onLaterPages=page_footer)
    return output.getvalue()


def _write_sheet(ws, title: str, headers: list[str], rows: list[list], primary_color: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=primary_color.lstrip("#"))
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=primary_color.lstrip("#"))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(2, ws.max_row)}"
    for column_index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(ws.cell(row=row, column=column_index).value or "") for row in range(3, ws.max_row + 1)]
        ws.column_dimensions[get_column_letter(column_index)].width = min(max(max(map(len, values)) + 2, 12), 42)
        if "Rate" in header or "Percentage" in header:
            for row_index in range(3, ws.max_row + 1):
                ws.cell(row=row_index, column=column_index).number_format = '0.0"%"'
        elif "Average" in header:
            for row_index in range(3, ws.max_row + 1):
                ws.cell(row=row_index, column=column_index).number_format = "0.0"
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            if isinstance(cell.value, float) and cell.number_format == "General":
                cell.number_format = "0.0"


def build_report_xlsx(report: dict, branding: dict | None = None) -> bytes:
    branding = branding or DEFAULT_BRANDING
    primary = branding.get("primary_color", "#1E3A8A")
    workbook = Workbook()
    workbook.remove(workbook.active)

    executive = report["executive_summary"]
    summary_rows = [
        ["Report Type", report["meta"]["report_type"].title()],
        ["Scope", report["meta"]["scope"]],
        ["Academic Year", report["meta"]["academic_year"]["name"]],
        ["Period Start", report["meta"]["period"]["start"]],
        ["Period End", report["meta"]["period"]["end"]],
        ["Generated UTC", report["meta"]["generated_at"].isoformat()],
        ["Total Students", executive["total_students"]],
        ["Attendance Rate", executive["attendance_rate"]],
        ["Late Rate", executive["late_rate"]],
        ["Late Minutes", executive["late_minutes"]],
        ["Below KKM Count", executive["below_kkm_count"]],
        ["Data Completeness Rate", executive["data_completeness_rate"]],
    ]
    if report["meta"]["report_type"] == "annual":
        for label, key in (("Highest Attendance Month", "highest_attendance_month"),
                           ("Lowest Attendance Month", "lowest_attendance_month"),
                           ("Highest Attendance Level", "highest_attendance_level"),
                           ("Lowest Attendance Level", "lowest_attendance_level")):
            row = report["comparisons"][key]
            summary_rows.append([label, row["name"] if row else None])
    _write_sheet(workbook.create_sheet("Executive Summary"), "Executive Summary", ["Metric", "Value"], summary_rows, primary)

    attendance_rows = []
    overall = report["attendance_summary"]
    attendance_rows.append(["Overall", overall["present"], overall["sakit"], overall["izin"], overall["alfa"], overall["incomplete"],
                            overall["late_days"], overall["late_minutes"], overall["attendance_rate"], overall["late_rate"]])
    for row in report["attendance_by_level"]:
        attendance_rows.append([row["level"], row["present"], row["sakit"], row["izin"], row["alfa"], row["incomplete"],
                                row["late_days"], row["late_minutes"], row["attendance_rate"], row["late_rate"]])
    _write_sheet(workbook.create_sheet("Attendance"), "Attendance", ["Level", "Present", "Sakit", "Izin", "Alfa", "Incomplete",
                 "Late Days", "Late Minutes", "Attendance Rate", "Late Rate"], attendance_rows, primary)

    distribution_rows = []
    for dimension, key in (("Level", "by_level"), ("Class", "by_class"), ("Gender", "by_gender"),
                           ("Religion", "by_religion"), ("Domicile", "by_domicile")):
        for row in report["student_distribution"][key]:
            distribution_rows.append([dimension, row["name"], row["count"], row["percentage"]])
    _write_sheet(workbook.create_sheet("Student Distribution"), "Student Distribution", ["Dimension", "Name", "Count", "Percentage"],
                 distribution_rows, primary)

    academic = report["academic_summary"]
    academic_rows = [["Overall", None, academic["sumatif_average"], academic["formatif_average"], academic["below_kkm_count"],
                      academic["availability"], academic["reason"]]]
    for row in academic["by_subject"]:
        academic_rows.append([row["subject_name"], row["jenjang"], row["sumatif_average"], row["formatif_average"],
                              row["below_kkm_count"], True, None])
    _write_sheet(workbook.create_sheet("Academic Summary"), "Academic Summary", ["Subject", "Level", "Sumatif Average", "Formatif Average",
                 "Below KKM Count", "Available", "Reason"], academic_rows, primary)

    if report["meta"]["report_type"] == "annual":
        trend_rows = [[row["month"], row["label"], row["present"], row["sakit"], row["izin"], row["alfa"], row["incomplete"],
                       row["attendance_denominator"], row["attendance_rate"], row["late_days"], row["late_minutes"], row["late_rate"],
                       row["sumatif_average"], row["formatif_average"], row["below_kkm_count"]] for row in report["trends"]]
        _write_sheet(workbook.create_sheet("Annual Trends"), "Annual Trends", ["Month", "Label", "Present", "Sakit", "Izin", "Alfa",
                     "Incomplete", "Attendance Denominator", "Attendance Rate", "Late Days", "Late Minutes", "Late Rate",
                     "Sumatif Average", "Formatif Average", "Below KKM Count"], trend_rows, primary)

    quality = report["data_quality"]
    quality_rows = [
        ["Missing Gender", quality["missing_gender"]], ["Missing Religion", quality["missing_religion"]],
        ["Missing Domicile", quality["missing_domicile"]], ["Incomplete Attendance", quality["incomplete_attendance"]],
        ["Empty Grade Cells", quality["empty_grade_cells"]], ["Unmapped Levels", ", ".join(quality["unmapped_levels"])],
    ] + [["Warning", warning] for warning in quality["warnings"]]
    _write_sheet(workbook.create_sheet("Data Quality"), "Data Quality", ["Metric", "Value"], quality_rows, primary)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
