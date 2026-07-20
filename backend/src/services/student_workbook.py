import hashlib
import json
from collections import Counter
from datetime import date, datetime, timedelta
from io import BytesIO

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from models.academic_master import AcademicClass
from models.academic_year import AcademicYear
from models.student_enrollment import StudentEnrollment
from models.student_master import (
    StudentAddress,
    StudentContact,
    StudentDeviceIdentity,
    StudentEnrollmentClassHistory,
    StudentImportBatch,
    StudentImportRow,
    StudentMaster,
    StudentParentGuardian,
)
from services.student_management import (
    _audit,
    _check_identifier_conflicts,
    _create_legacy_identity,
    _resolve_class,
    record_version,
)
from services.student_normalization import normalize_name


UPDATE_CONFIRMATION = "COMMIT_STUDENT_DATA_UPDATE"
SHEET_NAME = "Student Data"
HEADERS = [
    "OperatorOS Student UUID", "Record Version", "Legal Name", "Preferred Name",
    "NIPD", "NISN", "NIK", "Birth Place", "Birth Date", "Gender", "Religion",
    "Student Status", "Address", "Kelurahan", "Kecamatan", "City", "Province",
    "Postal Code", "Phone", "Email", "Guardian Name", "Guardian Phone",
    "Attendance Device No. ID", "Device Source", "Academic Year ID", "Academic Year",
    "Academic Class ID", "Class",
]
IDENTIFIER_COLUMNS = {"NIPD", "NISN", "NIK", "Attendance Device No. ID"}
PROFILE_MAP = {
    "Legal Name": "full_name", "Preferred Name": "preferred_name", "NIPD": "nipd",
    "NISN": "nisn", "NIK": "nik", "Birth Place": "birth_place",
    "Birth Date": "birth_date", "Gender": "gender", "Religion": "religion",
    "Student Status": "student_status",
}
ADDRESS_MAP = {
    "Address": "address", "Kelurahan": "kelurahan", "Kecamatan": "kecamatan",
    "City": "city_regency", "Province": "province", "Postal Code": "postal_code",
}
CONTACT_MAP = {"Phone": "student_phone", "Email": "student_email"}


def _clean(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _current_rows(db: Session):
    for student in db.query(StudentMaster).order_by(StudentMaster.full_name, StudentMaster.id):
        address = db.query(StudentAddress).filter_by(student_master_id=student.id).first()
        contact = db.query(StudentContact).filter_by(student_master_id=student.id).first()
        guardian = db.query(StudentParentGuardian).filter_by(student_master_id=student.id).order_by(StudentParentGuardian.id).first()
        device = db.query(StudentDeviceIdentity).filter_by(student_master_id=student.id, is_active=True).order_by(StudentDeviceIdentity.id.desc()).first()
        enrollment = (
            db.query(StudentEnrollment, AcademicYear, AcademicClass)
            .join(AcademicYear, AcademicYear.id == StudentEnrollment.academic_year_id)
            .outerjoin(AcademicClass, AcademicClass.id == StudentEnrollment.academic_class_id)
            .filter(StudentEnrollment.student_master_id == student.id)
            .order_by(AcademicYear.start_date.desc()).first()
        )
        yield [
            student.id, record_version(student), student.full_name, student.preferred_name,
            student.nipd, student.nisn, student.nik, student.birth_place, student.birth_date,
            student.gender, student.religion, student.student_status,
            address.address if address else None, address.kelurahan if address else None,
            address.kecamatan if address else None, address.city_regency if address else None,
            address.province if address else None, address.postal_code if address else None,
            contact.student_phone if contact else None, contact.student_email if contact else None,
            guardian.name if guardian else None, guardian.phone if guardian else None,
            device.device_identifier if device else None, device.device_source if device else "attendance_machine",
            enrollment[0].academic_year_id if enrollment else None, enrollment[1].label if enrollment else None,
            enrollment[0].academic_class_id if enrollment else None,
            enrollment[2].class_name if enrollment and enrollment[2] else enrollment[0].class_name if enrollment else None,
        ]


def export_student_workbook(db: Session) -> bytes:
    workbook = Workbook()
    data = workbook.active
    data.title = SHEET_NAME
    data.append(HEADERS)
    for row in _current_rows(db):
        data.append(row)
    data.freeze_panes = "A2"
    data.auto_filter.ref = data.dimensions
    data.row_dimensions[1].height = 30
    for cell in data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    widths = {header: max(14, min(36, len(header) + 4)) for header in HEADERS}
    widths["Legal Name"] = 28
    widths["Address"] = 36
    for index, header in enumerate(HEADERS, 1):
        data.column_dimensions[data.cell(1, index).column_letter].width = widths[header]
        if header in IDENTIFIER_COLUMNS or header in {"OperatorOS Student UUID", "Record Version"}:
            for cell in data.iter_cols(min_col=index, max_col=index, min_row=2, max_row=max(2, data.max_row)):
                for item in cell:
                    item.number_format = "@"
        if header == "Birth Date":
            for cell in data.iter_cols(min_col=index, max_col=index, min_row=2, max_row=max(2, data.max_row)):
                for item in cell:
                    item.number_format = "yyyy-mm-dd"

    reference = workbook.create_sheet("Reference Values")
    reference.append(["Gender", "Religion", "Student Status"])
    genders = ["male", "female"]
    religions = ["Islam", "Christian", "Catholic", "Hindu", "Buddhist", "Other"]
    statuses = ["pending_review", "active", "inactive", "transferred", "withdrawn", "graduated", "archived"]
    for index in range(max(len(genders), len(religions), len(statuses))):
        reference.append([
            genders[index] if index < len(genders) else None,
            religions[index] if index < len(religions) else None,
            statuses[index] if index < len(statuses) else None,
        ])
    reference.sheet_state = "hidden"
    for formula, header in [
        ("'Reference Values'!$A$2:$A$3", "Gender"),
        ("'Reference Values'!$B$2:$B$7", "Religion"),
        ("'Reference Values'!$C$2:$C$8", "Student Status"),
    ]:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        data.add_data_validation(validation)
        column = HEADERS.index(header) + 1
        validation.add(f"{data.cell(2, column).column_letter}2:{data.cell(2, column).column_letter}1048576")

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["OperatorOS Student Data Update"])
    instructions.append(["Generated", datetime.now().replace(microsecond=0).isoformat()])
    instructions.append(["Match rule", "Rows are matched only by OperatorOS Student UUID."])
    instructions.append(["Concurrency", "Do not edit Record Version; stale rows are blocked during preview."])
    instructions.append(["Identifiers", "NIPD, NISN, NIK, and Attendance Device No. ID are stored as text."])
    instructions.append(["Commit", "Uploading creates a preview only. An administrator must select and confirm changes."])
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 92
    instructions["A1"].font = Font(bold=True, size=14)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _preview_checksum(rows: list[StudentImportRow]) -> str:
    payload = [{
        "source_row": row.source_row,
        "classification": row.classification,
        "student": row.matched_student_master_id,
        "payload": row.normalized_payload,
        "differences": row.differences,
        "errors": row.validation_errors,
    } for row in rows]
    return hashlib.sha256(json.dumps(payload, sort_keys=True, default=str, separators=(",", ":")).encode()).hexdigest()


def serialize_update_batch(db: Session, batch: StudentImportBatch) -> dict:
    rows = db.query(StudentImportRow).filter_by(batch_id=batch.id).order_by(StudentImportRow.source_row).all()
    return {
        "id": batch.id, "filename": batch.filename, "file_checksum": batch.file_checksum,
        "status": batch.status, "created_by": batch.created_by, "created_at": batch.created_at,
        "committed_at": batch.committed_at, "preview_checksum": _preview_checksum(rows),
        "summary": {
            "total": batch.total_rows, "updates": batch.update_count,
            "unchanged": batch.unchanged_count, "conflicts": batch.conflict_count,
            "invalid": batch.invalid_count,
        },
        "rows": [{
            "id": row.id, "source_row": row.source_row,
            "classification": row.classification,
            "student_master_id": row.matched_student_master_id,
            "payload": row.normalized_payload, "differences": row.differences,
            "errors": row.validation_errors, "selected": row.selected_for_commit,
        } for row in rows],
    }


def create_update_preview(db: Session, file_bytes: bytes, filename: str, username: str) -> StudentImportBatch:
    checksum = hashlib.sha256(file_bytes).hexdigest()
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="The uploaded file is not a readable XLSX workbook") from exc
    if SHEET_NAME not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"Workbook must contain the '{SHEET_NAME}' worksheet")
    sheet = workbook[SHEET_NAME]
    header_values = [_clean(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])]
    missing = [header for header in HEADERS if header not in header_values]
    if missing:
        raise HTTPException(status_code=400, detail={"code": "MISSING_COLUMNS", "columns": missing})
    indexes = {header: header_values.index(header) for header in HEADERS}
    batch = StudentImportBatch(
        filename=filename, file_checksum=checksum, source_sheet="student_update",
        status="preview", created_by=username,
    )
    db.add(batch); db.flush()
    counts = Counter()
    try:
        for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if not any(value is not None and str(value).strip() for value in values):
                continue
            payload = {header: _clean(values[indexes[header]]) if indexes[header] < len(values) else None for header in HEADERS}
            errors = []
            differences = {}
            student_id = payload["OperatorOS Student UUID"]
            student = db.get(StudentMaster, student_id) if student_id else None
            classification = "INVALID"
            if not student_id:
                errors.append({"code": "MISSING_UUID", "message": "OperatorOS Student UUID is required"})
            elif student is None:
                errors.append({"code": "UNKNOWN_UUID", "message": "Student UUID was not found"})
            elif payload["Record Version"] != record_version(student):
                classification = "CONFLICT"
                errors.append({"code": "STALE_RECORD", "message": "Student changed after this workbook was exported"})
            else:
                for column, field in PROFILE_MAP.items():
                    uploaded = payload[column]
                    if field == "birth_date":
                        try:
                            uploaded = date.fromisoformat(uploaded) if uploaded else None
                        except ValueError:
                            errors.append({"code": "INVALID_DATE", "field": column, "message": "Use YYYY-MM-DD"})
                            continue
                    current = getattr(student, field)
                    if column == "Legal Name" and not uploaded:
                        errors.append({"code": "REQUIRED", "field": column, "message": "Legal Name is required"})
                    if field == "nisn" and uploaded and (not uploaded.isdigit() or len(uploaded) != 10):
                        errors.append({"code": "INVALID_NISN", "field": column, "message": "NISN must contain 10 digits"})
                    if field == "nik" and uploaded and (not uploaded.isdigit() or len(uploaded) != 16):
                        errors.append({"code": "INVALID_NIK", "field": column, "message": "NIK must contain 16 digits"})
                    if current != uploaded:
                        differences[field] = {"current": _clean(current), "uploaded": _clean(uploaded), "action": "update_profile"}
                address = db.query(StudentAddress).filter_by(student_master_id=student.id).first()
                contact = db.query(StudentContact).filter_by(student_master_id=student.id).first()
                for column, field in {**ADDRESS_MAP, **CONTACT_MAP}.items():
                    current = getattr(address if field in ADDRESS_MAP.values() else contact, field) if (address if field in ADDRESS_MAP.values() else contact) else None
                    if current != payload[column]:
                        differences[field] = {"current": current, "uploaded": payload[column], "action": "update_profile"}
                guardian = db.query(StudentParentGuardian).filter_by(student_master_id=student.id).order_by(StudentParentGuardian.id).first()
                for column, field in (("Guardian Name", "guardian_name"), ("Guardian Phone", "guardian_phone")):
                    current = getattr(guardian, "name" if field == "guardian_name" else "phone") if guardian else None
                    if current != payload[column]:
                        differences[field] = {"current": current, "uploaded": payload[column], "action": "update_guardian"}
                device = db.query(StudentDeviceIdentity).filter_by(student_master_id=student.id, is_active=True).order_by(StudentDeviceIdentity.id.desc()).first()
                uploaded_device = payload["Attendance Device No. ID"]
                if uploaded_device and not uploaded_device.isdigit():
                    errors.append({"code": "INVALID_DEVICE_ID", "field": "Attendance Device No. ID", "message": "Device ID must contain digits only"})
                if (device.device_identifier if device else None) != uploaded_device:
                    differences["device_identifier"] = {"current": device.device_identifier if device else None, "uploaded": uploaded_device, "action": "replace_device"}
                enrollment = db.query(StudentEnrollment).filter_by(student_master_id=student.id, academic_year_id=int(payload["Academic Year ID"]) if payload["Academic Year ID"] and payload["Academic Year ID"].isdigit() else -1).first()
                class_id = int(payload["Academic Class ID"]) if payload["Academic Class ID"] and payload["Academic Class ID"].isdigit() else None
                if class_id and (not enrollment or enrollment.academic_class_id != class_id):
                    differences["academic_class_id"] = {"current": enrollment.academic_class_id if enrollment else None, "uploaded": class_id, "action": "transfer_or_enroll"}
                for field in ("nipd", "nisn", "nik"):
                    uploaded = differences.get(field, {}).get("uploaded")
                    if uploaded:
                        conflict = db.query(StudentMaster).filter(getattr(StudentMaster, field) == uploaded, StudentMaster.id != student.id).first()
                        if conflict:
                            errors.append({"code": f"DUPLICATE_{field.upper()}", "field": field, "message": f"{field.upper()} belongs to another student"})
                if uploaded_device:
                    conflict = db.query(StudentDeviceIdentity).filter(StudentDeviceIdentity.device_identifier == uploaded_device, StudentDeviceIdentity.is_active.is_(True), StudentDeviceIdentity.student_master_id != student.id).first()
                    if conflict:
                        errors.append({"code": "DEVICE_ID_IN_USE", "field": "device_identifier", "message": "Attendance Device ID belongs to another student"})
                classification = "INVALID" if errors else "UPDATE_EXISTING_MASTER" if differences else "NO_CHANGE"
            counts[classification] += 1
            db.add(StudentImportRow(
                batch_id=batch.id, source_row=source_row, classification=classification,
                matched_student_master_id=student.id if student else None,
                normalized_payload=payload, differences=differences,
                validation_errors=errors, selected_for_commit=False,
            ))
        batch.total_rows = sum(counts.values())
        batch.update_count = counts["UPDATE_EXISTING_MASTER"]
        batch.unchanged_count = counts["NO_CHANGE"]
        batch.conflict_count = counts["CONFLICT"]
        batch.invalid_count = counts["INVALID"]
        db.commit(); db.refresh(batch); return batch
    except Exception:
        db.rollback(); raise


def commit_update_preview(db: Session, batch_id: str, row_ids: list[int], confirmation: str, preview_checksum: str, username: str) -> dict:
    if confirmation != UPDATE_CONFIRMATION:
        raise HTTPException(status_code=400, detail="Invalid confirmation token")
    batch = db.get(StudentImportBatch, batch_id)
    if batch is None or batch.source_sheet != "student_update":
        raise HTTPException(status_code=404, detail="Student update preview not found")
    rows = db.query(StudentImportRow).filter_by(batch_id=batch.id).order_by(StudentImportRow.source_row).all()
    if preview_checksum != _preview_checksum(rows):
        raise HTTPException(status_code=409, detail={"code": "PREVIEW_CHECKSUM_MISMATCH", "message": "Preview changed; upload the workbook again"})
    if batch.status == "committed":
        return {"status": "committed", "batch_id": batch.id, "updated": batch.update_count, "idempotent_replay": True}
    if batch.created_at and datetime.now() - batch.created_at > timedelta(hours=24):
        batch.status = "expired"; db.commit()
        raise HTTPException(status_code=409, detail={"code": "PREVIEW_EXPIRED", "message": "Preview expired; upload the workbook again"})
    selected = [row for row in rows if row.id in set(row_ids)]
    if not selected or len(selected) != len(set(row_ids)):
        raise HTTPException(status_code=400, detail="Selected rows are not part of this preview")
    blocked = [row.source_row for row in selected if row.classification != "UPDATE_EXISTING_MASTER"]
    if blocked:
        raise HTTPException(status_code=409, detail={"code": "ROWS_NOT_COMMITTABLE", "rows": blocked})
    try:
        for row in selected:
            student = db.get(StudentMaster, row.matched_student_master_id)
            if student is None or row.normalized_payload["Record Version"] != record_version(student):
                raise HTTPException(status_code=409, detail={"code": "STALE_RECORD", "row": row.source_row})
            payload = row.normalized_payload
            for field, change in row.differences.items():
                if field in PROFILE_MAP.values():
                    new = change["uploaded"]
                    if field == "birth_date":
                        new = date.fromisoformat(new) if new else None
                    _audit(db, student.id, "profile_updated", username, "student_update_workbook", field, getattr(student, field), new, batch.id)
                    setattr(student, field, new)
            _check_identifier_conflicts(db, {field: getattr(student, field) for field in ("nipd", "nisn", "nik")}, student.id)
            student.normalized_name = normalize_name(student.full_name)
            student.updated_by = username
            student.updated_at = datetime.now()
            address = db.query(StudentAddress).filter_by(student_master_id=student.id).first()
            contact = db.query(StudentContact).filter_by(student_master_id=student.id).first()
            for field in ADDRESS_MAP.values():
                if field in row.differences:
                    if address is None: address = StudentAddress(student_master_id=student.id); db.add(address)
                    change = row.differences[field]; _audit(db, student.id, "profile_updated", username, "student_update_workbook", field, change["current"], change["uploaded"], batch.id); setattr(address, field, change["uploaded"])
            for field in CONTACT_MAP.values():
                if field in row.differences:
                    if contact is None: contact = StudentContact(student_master_id=student.id); db.add(contact)
                    change = row.differences[field]; _audit(db, student.id, "profile_updated", username, "student_update_workbook", field, change["current"], change["uploaded"], batch.id); setattr(contact, field, change["uploaded"])
            if "guardian_name" in row.differences or "guardian_phone" in row.differences:
                guardian = db.query(StudentParentGuardian).filter_by(student_master_id=student.id).order_by(StudentParentGuardian.id).first()
                if guardian is None:
                    guardian = StudentParentGuardian(student_master_id=student.id, guardian_type="guardian", name=payload["Guardian Name"] or "Not provided")
                    db.add(guardian)
                else:
                    guardian.name = payload["Guardian Name"] or guardian.name
                guardian.phone = payload["Guardian Phone"]
                _audit(db, student.id, "guardians_updated", username, "student_update_workbook", "guardian", None, guardian.name, batch.id)
            if "device_identifier" in row.differences:
                change = row.differences["device_identifier"]
                old = db.query(StudentDeviceIdentity).filter_by(student_master_id=student.id, is_active=True).order_by(StudentDeviceIdentity.id.desc()).first()
                if change["uploaded"]:
                    conflict = db.query(StudentDeviceIdentity).filter(StudentDeviceIdentity.device_identifier == change["uploaded"], StudentDeviceIdentity.is_active.is_(True), StudentDeviceIdentity.student_master_id != student.id).first()
                    if conflict: raise HTTPException(status_code=409, detail={"code": "DEVICE_ID_IN_USE", "row": row.source_row})
                    if old: old.is_active = False; old.effective_to = date.today()
                    _create_legacy_identity(db, student, {"device_identifier": change["uploaded"], "device_source": payload["Device Source"] or "attendance_machine", "effective_from": date.today(), "actor": username})
                elif old:
                    old.is_active = False; old.effective_to = date.today()
                _audit(db, student.id, "device_identity_replaced", username, "student_update_workbook", "device_identifier", change["current"], change["uploaded"], batch.id)
            if "academic_class_id" in row.differences:
                class_id = int(row.differences["academic_class_id"]["uploaded"])
                year_id = int(payload["Academic Year ID"])
                target, _grade, _program, jenjang, year = _resolve_class(db, year_id, class_id)
                enrollment = db.query(StudentEnrollment).filter_by(student_master_id=student.id, academic_year_id=year_id).first()
                mapping = db.query(StudentDeviceIdentity).filter_by(student_master_id=student.id, is_active=True).first()
                if mapping is None: raise HTTPException(status_code=409, detail={"code": "DEVICE_ID_REQUIRED", "row": row.source_row})
                if enrollment:
                    old_class = enrollment.class_name
                    enrollment.academic_class_id = target.id; enrollment.jenjang_id = jenjang.id; enrollment.class_name = target.class_name; enrollment.class_assigned = True
                else:
                    old_class = None
                    enrollment = StudentEnrollment(student_id=mapping.legacy_student_id, student_master_id=student.id, academic_year_id=year.id, jenjang_id=jenjang.id, academic_class_id=target.id, class_name=target.class_name, class_assigned=True, effective_from=max(date.today(), year.start_date))
                    db.add(enrollment); db.flush()
                db.add(StudentEnrollmentClassHistory(enrollment_id=enrollment.id, class_name=target.class_name, effective_from=max(date.today(), year.start_date), changed_by=username, source="student_update_workbook", import_batch_id=batch.id))
                _audit(db, student.id, "enrollment_transferred" if old_class else "enrollment_created", username, "student_update_workbook", "class_name", old_class, target.class_name, batch.id)
            row.selected_for_commit = True
        batch.status = "committed"; batch.committed_at = datetime.now(); batch.update_count = len(selected)
        db.commit()
        return {"status": "committed", "batch_id": batch.id, "updated": len(selected), "idempotent_replay": False}
    except HTTPException:
        db.rollback(); raise
    except IntegrityError as exc:
        db.rollback(); raise HTTPException(status_code=409, detail="Student update conflicts with current data") from exc


def result_workbook(db: Session, batch: StudentImportBatch) -> bytes:
    rows = db.query(StudentImportRow).filter_by(batch_id=batch.id).order_by(StudentImportRow.source_row).all()
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Validation Summary"
    sheet.append(["Original Row", "Student UUID", "Name", "Status", "Proposed Action", "Result", "Error Code", "Error Message", "Suggested Resolution"])
    for row in rows:
        errors = row.validation_errors or []
        error = errors[0] if errors else {}
        sheet.append([
            row.source_row, row.matched_student_master_id, row.normalized_payload.get("Legal Name"),
            row.classification, ", ".join(sorted({value["action"] for value in (row.differences or {}).values()})),
            "Committed" if row.selected_for_commit and batch.status == "committed" else "Not committed",
            error.get("code"), error.get("message"), "Re-export current data and resolve the indicated row" if error else None,
        ])
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    output = BytesIO(); workbook.save(output); return output.getvalue()
