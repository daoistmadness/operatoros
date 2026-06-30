import calendar
from collections import defaultdict
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import logging
import re
import time

import pandas as pd

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import Float, Integer, String, and_, case, cast, desc, func, literal, or_, select, true
from sqlalchemy.orm import Session

from core.database import get_db
from models.absence_reason import AbsenceReason
from models.absence_reason_class_entry import AbsenceReasonClassEntry
from models.attendance import Attendance
from models.attendance_review import AttendanceOverride
from models.student import Student
from services.attendance_metrics import (
    calculate_auto_heb,
    calculate_heb,
    get_heb_override,
    month_year_filters,
    month_bucket_string_expression,
)
from services.management_analytics import build_management_summary

from services.management_report_export import (
    PDF_MIME,
    XLSX_MIME,
    build_management_report_filename,
    build_management_summary_excel,
    build_management_summary_pdf,
)

router = APIRouter()

logger = logging.getLogger(__name__)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_REPORT_TITLE = "Student Tardiness Report"
_REKAP_ABSENSI_TITLE = "Rekap Absensi Siswa SD"
_SCHOOL_NAME = "EDELWEISS SCHOOL"
_INDONESIAN_MONTHS = [
    "Januari",
    "Februari",
    "Maret",
    "April",
    "Mei",
    "Juni",
    "Juli",
    "Agustus",
    "September",
    "Oktober",
    "November",
    "Desember",
]
_TERM_METADATA = {
    1: {"months": (7, 9), "label": "July–September"},
    2: {"months": (10, 12), "label": "October–December"},
    3: {"months": (1, 3), "label": "January–March"},
    4: {"months": (4, 6), "label": "April–June"},
}
_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _month_bucket_expression(db: Session):
    dialect = db.bind.dialect.name if db.bind is not None else ""
    if dialect == "sqlite":
        return func.strftime("%Y-%m-01", Attendance.date)
    return func.date_trunc("month", Attendance.date)


def _serialize_rows(rows, keys):
    serialized = []
    for row in rows:
        values = tuple(row)
        serialized.append(dict(zip(keys, values)))
    return serialized


def _parse_month_key(month_key: str) -> tuple[int, int]:
    year_text, month_text = month_key.split("-")
    return int(year_text), int(month_text)


def _safe_rate(numerator: float, denominator: float):
    if denominator == 0:
        return None
    return round(numerator / denominator, 3)


def _format_duration(td: int | float | None) -> str:
    if td is None:
        return "—"

    if isinstance(td, (int, float)):
        total_minutes = int(td)
    else:
        total_minutes = int(td.total_seconds() // 60)

    if total_minutes <= 0:
        return "—"

    hours = total_minutes // 60
    minutes = total_minutes % 60
    if hours > 0:
        return f"{hours}h {minutes}m" if minutes > 0 else f"{hours}h"
    return f"{minutes}m"


def _class_name_label_expression():
    normalized = func.trim(func.coalesce(Student.class_name, literal("")))
    return case(
        (func.length(normalized) == 0, literal("Belum Diatur")),
        else_=normalized,
    )


def _report_jenjang_expression(db: Session):
    return func.upper(func.coalesce(Student.jenjang, literal("Unassigned")))


@router.get("/jenjangs")
def get_all_jenjangs(db: Session = Depends(get_db)):
    """
    Returns a sorted distinct list of all assigned jenjangs.
    Used for filtering reports.
    """
    rows = db.execute(select(Student.jenjang).filter(Student.jenjang.isnot(None)).distinct()).scalars().all()
    return sorted(list(set(r.strip() for r in rows if r and r.strip())))


def _minutes_to_hhmm_expression(db: Session, minutes_expression):
    dialect = db.bind.dialect.name if db.bind is not None else ""
    total_minutes = func.coalesce(cast(minutes_expression, Integer), 0)
    hours = cast(total_minutes / 60, Integer)
    minutes = total_minutes % 60

    if dialect == "sqlite":
        return func.printf("%02d:%02d", hours, minutes)

    return (
        func.lpad(cast(hours, String), 2, literal("0"))
        + literal(":")
        + func.lpad(cast(minutes, String), 2, literal("0"))
    )


def _percentage_expression(numerator_expression, denominator_expression):
    denominator = func.coalesce(denominator_expression, 0)
    return case(
        (denominator == 0, literal(0.0)),
        else_=cast(
            func.round((cast(numerator_expression, Float) * 100.0) / denominator_expression, 1),
            Float,
        ),
    )


def _academic_year_label(anchor_date: date) -> str:
    academic_year_start = anchor_date.year if anchor_date.month >= 7 else anchor_date.year - 1
    return f"TA {academic_year_start}/{academic_year_start + 1}"


def _month_pairs_in_range(start_date: date, end_date: date) -> list[tuple[int, int]]:
    pairs = []
    year = start_date.year
    month = start_date.month

    while (year, month) <= (end_date.year, end_date.month):
        pairs.append((year, month))
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1

    return pairs


def _get_absence_reason_map_for_range(db: Session, start_date: date, end_date: date) -> dict[str, dict]:
    month_pairs = _month_pairs_in_range(start_date, end_date)
    if not month_pairs:
        return {}

    pair_set = set(month_pairs)
    grouped: dict[str, dict] = {}

    class_entries = (
        db.query(AbsenceReasonClassEntry)
        .filter(AbsenceReasonClassEntry.year >= start_date.year, AbsenceReasonClassEntry.year <= end_date.year)
        .all()
    )
    class_entry_map = {
        (row.class_name, row.year, row.month): row
        for row in class_entries
        if (row.year, row.month) in pair_set
    }

    student_rows = db.query(AbsenceReason).filter(AbsenceReason.year >= start_date.year, AbsenceReason.year <= end_date.year).all()
    student_grouped: dict[tuple[str, int, int], dict] = {}
    for row in student_rows:
        if (row.year, row.month) not in pair_set:
            continue
        key = (row.class_name, row.year, row.month)
        if key not in student_grouped:
            student_grouped[key] = {"sakit": 0, "izin": 0, "alfa": 0}
        student_grouped[key]["sakit"] += int(row.sakit or 0)
        student_grouped[key]["izin"] += int(row.izin or 0)
        student_grouped[key]["alfa"] += int(row.alfa or 0)

    monthly_keys = set(student_grouped.keys()) | set(class_entry_map.keys())
    for class_name, year, month in monthly_keys:
        if (year, month) not in pair_set:
            continue
        source = class_entry_map.get((class_name, year, month))
        values = (
            {
                "sakit": int(source.sakit or 0),
                "izin": int(source.izin or 0),
                "alfa": int(source.alfa or 0),
            }
            if source is not None
            else student_grouped.get((class_name, year, month), {"sakit": 0, "izin": 0, "alfa": 0})
        )
        if class_name not in grouped:
            grouped[class_name] = {"sakit": 0, "izin": 0, "alfa": 0, "total_absence_reasons": 0}

        grouped[class_name]["sakit"] += values["sakit"]
        grouped[class_name]["izin"] += values["izin"]
        grouped[class_name]["alfa"] += values["alfa"]
        grouped[class_name]["total_absence_reasons"] = (
            grouped[class_name]["sakit"] + grouped[class_name]["izin"] + grouped[class_name]["alfa"]
        )
    return grouped


def _get_student_absence_reason_map_for_range(db: Session, start_date: date, end_date: date) -> dict[int, dict]:
    month_pairs = _month_pairs_in_range(start_date, end_date)
    if not month_pairs:
        return {}
    pair_set = set(month_pairs)

    rows = db.query(AbsenceReason).filter(AbsenceReason.year >= start_date.year, AbsenceReason.year <= end_date.year).all()
    grouped: dict[int, dict] = {}
    for row in rows:
        if (row.year, row.month) not in pair_set:
            continue
        if row.student_id not in grouped:
            grouped[row.student_id] = {"sakit": 0, "izin": 0, "alfa": 0}

        grouped[row.student_id]["sakit"] += int(row.sakit or 0)
        grouped[row.student_id]["izin"] += int(row.izin or 0)
        grouped[row.student_id]["alfa"] += int(row.alfa or 0)
    return grouped


def _resolve_tardiness_period(
    month: int | None,
    year: int | None,
    date_from: date | None,
    date_to: date | None,
    term: int | None,
):
    today = date.today()

    if (date_from is None) != (date_to is None):
        raise HTTPException(status_code=400, detail="date_from and date_to must be provided together")

    if date_from is not None and date_to is not None:
        if date_from > date_to:
            raise HTTPException(status_code=400, detail="date_from must be before or equal to date_to")
        return {
            "date_from": date_from,
            "date_to": date_to,
            "label": f"{date_from.strftime('%d/%m/%Y')} - {date_to.strftime('%d/%m/%Y')}",
            "mode": "date_range",
        }

    if term is not None and year is None:
        raise HTTPException(status_code=400, detail="year is required when term is provided")

    if month is not None and year is None:
        raise HTTPException(status_code=400, detail="year is required when month is provided")

    if term is not None and year is not None:
        term_metadata = _TERM_METADATA.get(term)
        if term_metadata is None:
            raise HTTPException(status_code=400, detail="invalid term")

        start_month, end_month = term_metadata["months"]
        start_date = date(year, start_month, 1)
        end_date = date(year, end_month, calendar.monthrange(year, end_month)[1])
        return {
            "date_from": start_date,
            "date_to": end_date,
            "label": f"Term {term} ({term_metadata['label']}) - {_academic_year_label(start_date)}",
            "mode": "term",
        }

    if month is not None and year is not None:
        start_date = date(year, month, 1)
        end_date = date(year, month, calendar.monthrange(year, month)[1])
        return {
            "date_from": start_date,
            "date_to": end_date,
            "label": f"{_INDONESIAN_MONTHS[month - 1]} {year}",
            "mode": "month",
        }

    if year is not None:
        return {
            "date_from": date(year, 1, 1),
            "date_to": date(year, 12, 31),
            "label": f"Tahun {year}",
            "mode": "year",
        }

    start_date = date(today.year, today.month, 1)
    end_date = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    return {
        "date_from": start_date,
        "date_to": end_date,
        "label": f"{_INDONESIAN_MONTHS[today.month - 1]} {today.year}",
        "mode": "current_month",
    }


def _report_period_slug(label: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", label.strip().lower())
    return slug.strip("_") or "periode"


def _resolve_rekap_absensi_period(
    month: int | None,
    year: int | None,
    date_from: date | None,
    date_to: date | None,
    term: int | None,
):
    period = _resolve_tardiness_period(month, year, date_from, date_to, term)

    period_year = year
    if period_year is None:
        if date_to is not None:
            period_year = date_to.year
        elif date_from is not None:
            period_year = date_from.year
        else:
            period_year = period["date_to"].year

    period_label = period["label"]
    if period["mode"] == "month":
        period_label = f"{period['label']} - {_academic_year_label(period['date_from'])}"
    elif period["mode"] == "date_range":
        period_label = f"{period['label']} - {_academic_year_label(period['date_to'])}"

    return {
        **period,
        "label": period_label,
        "term": term if period["mode"] == "term" else None,
        "year": period_year,
        "month": month if period["mode"] == "month" else None,
    }


def _round_half_up(value: float | int | Decimal) -> int:
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _round_percentage_int(numerator: int | float, denominator: int | float) -> float | None:
    if not denominator:
        return None
    val = (Decimal(str(numerator)) * Decimal("100")) / Decimal(str(denominator))
    return float(val.quantize(Decimal("0.01"), rounding="ROUND_HALF_UP"))


def _normalize_percentage_row(hadir_pct: float | None, sakit_pct: float | None, izin_pct: float | None, alfa_pct: float | None, lain2_pct: float | None):
    values = [hadir_pct, sakit_pct, izin_pct, alfa_pct, lain2_pct]
    if any(value is None for value in values):
        return {
            "hadir_pct": None,
            "sakit_pct": None,
            "izin_pct": None,
            "alfa_pct": None,
            "lain2_pct": None,
            "total_pct": None,
        }

    total_pct = sum(values)
    # Absorb the rounding remainder into hadir, which is expected to be
    # the largest bucket in normal attendance data. Guard against going negative
    # if a future dataset violates that assumption.
    if abs(total_pct - 100.0) > 0.001:
        adjustment = 100.0 - total_pct
        hadir_pct = max(0.0, hadir_pct + adjustment)

    total_pct = 100.0

    return {
        "hadir_pct": hadir_pct,
        "sakit_pct": sakit_pct,
        "izin_pct": izin_pct,
        "alfa_pct": alfa_pct,
        "lain2_pct": lain2_pct,
        "total_pct": total_pct,
    }


def _month_pair_filters(model, month_pairs: list[tuple[int, int]]):
    if not month_pairs:
        return literal(False)
    return or_(*[and_(model.month == month, model.year == year) for year, month in month_pairs])


def _valid_student_jenjang_filter():
    normalized = func.trim(func.coalesce(Student.jenjang, literal("")))
    return func.length(normalized) > 0


def _valid_student_class_filter():
    normalized = func.trim(func.coalesce(Student.class_name, literal("")))
    return func.length(normalized) > 0


def _format_rekap_excel_pct(value: int | float | None) -> str:
    return "-" if value is None else f"{value:.2f}%"


def _collect_rekap_absensi_report_data(db: Session, period: dict):
    month_pairs = _month_pairs_in_range(period["date_from"], period["date_to"])
    jenjang_expr = _report_jenjang_expression(db)
    raw_jenjang_expr = func.trim(Student.jenjang)
    effective_status = func.coalesce(AttendanceOverride.override_status, Attendance.status)

    student_count_stmt = (
        select(
            jenjang_expr.label("jenjang"),
            raw_jenjang_expr.label("raw_jenjang"),
            func.count(Student.id).label("student_count"),
        )
        .select_from(Student)
        .where(_valid_student_jenjang_filter())
        .group_by(jenjang_expr, raw_jenjang_expr)
        .order_by(jenjang_expr.asc())
    )
    student_count_rows = db.execute(student_count_stmt).mappings().all()
    student_counts: dict[str, int] = defaultdict(int)
    jenjang_source_map: dict[str, str] = {}
    for row in student_count_rows:
        count = int(row["student_count"] or 0)
        if count <= 0:
            continue
        display_jenjang = row["jenjang"]
        student_counts[display_jenjang] += count
        if display_jenjang not in jenjang_source_map:
            jenjang_source_map[display_jenjang] = row["raw_jenjang"]

    attendance_stmt = (
        select(
            jenjang_expr.label("jenjang"),
            func.count(Attendance.id).label("hadir_days"),
        )
        .select_from(Attendance)
        .join(Student, Student.id == Attendance.student_id)
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .where(
            Attendance.date >= period["date_from"],
            Attendance.date <= period["date_to"],
            _valid_student_jenjang_filter(),
            effective_status.in_(("on-time", "late")),
        )
        .group_by(jenjang_expr)
    )
    attendance_rows = db.execute(attendance_stmt).mappings().all()
    hadir_days_map = {row["jenjang"]: int(row["hadir_days"] or 0) for row in attendance_rows}

    class_jenjang_cte = (
        select(
            func.trim(Student.class_name).label("class_name"),
            jenjang_expr.label("jenjang"),
        )
        .select_from(Student)
        .where(_valid_student_jenjang_filter(), _valid_student_class_filter())
        .group_by(func.trim(Student.class_name), jenjang_expr)
        .cte("class_jenjang_map")
    )

    absence_stmt = (
        select(
            class_jenjang_cte.c.jenjang.label("jenjang"),
            func.coalesce(func.sum(AbsenceReasonClassEntry.sakit), 0).label("sakit"),
            func.coalesce(func.sum(AbsenceReasonClassEntry.izin), 0).label("izin"),
            func.coalesce(func.sum(AbsenceReasonClassEntry.alfa), 0).label("alfa"),
            func.count(func.distinct(class_jenjang_cte.c.class_name)).label("classes_with_sia"),
        )
        .select_from(AbsenceReasonClassEntry)
        .join(class_jenjang_cte, func.trim(AbsenceReasonClassEntry.class_name) == class_jenjang_cte.c.class_name)
        .where(_month_pair_filters(AbsenceReasonClassEntry, month_pairs))
        .group_by(class_jenjang_cte.c.jenjang)
    )
    absence_rows = db.execute(absence_stmt).mappings().all()
    absence_map = {
        row["jenjang"]: {
            "sakit": int(row["sakit"] or 0),
            "izin": int(row["izin"] or 0),
            "alfa": int(row["alfa"] or 0),
            "classes_with_sia": int(row["classes_with_sia"] or 0),
        }
        for row in absence_rows
    }

    class_count_stmt = (
        select(
            jenjang_expr.label("jenjang"),
            func.count(func.distinct(func.trim(Student.class_name))).label("class_count"),
        )
        .select_from(Student)
        .where(_valid_student_jenjang_filter(), _valid_student_class_filter())
        .group_by(jenjang_expr)
    )
    class_count_rows = db.execute(class_count_stmt).mappings().all()
    class_count_map = {row["jenjang"]: int(row["class_count"] or 0) for row in class_count_rows}

    period_sia_entry_count = (
        db.execute(
            select(func.count(AbsenceReasonClassEntry.id)).where(_month_pair_filters(AbsenceReasonClassEntry, month_pairs))
        ).scalar()
        or 0
    )

    heb_cache: dict[tuple[str, int, int], int] = {}
    warnings: list[str] = []
    rows: list[dict] = []
    heb_zero_jenjangs: list[str] = []

    for jenjang in sorted(student_counts.keys()):
        student_count = student_counts[jenjang]
        source_jenjang = jenjang_source_map.get(jenjang, jenjang)
        heb_total = 0
        for pair_year, pair_month in month_pairs:
            cache_key = (source_jenjang, pair_year, pair_month)
            if cache_key not in heb_cache:
                heb_cache[cache_key] = int(calculate_heb(db, source_jenjang, pair_month, pair_year)["heb"] or 0)
            heb_total += heb_cache[cache_key]

        hadir_days = int(hadir_days_map.get(jenjang, 0))
        absence_values = absence_map.get(jenjang, {"sakit": 0, "izin": 0, "alfa": 0, "classes_with_sia": 0})
        sakit = int(absence_values["sakit"])
        izin = int(absence_values["izin"])
        alfa = int(absence_values["alfa"])
        denominator = student_count * heb_total
        lain2 = max(denominator - hadir_days - sakit - izin - alfa, 0)

        # Asumsikan data kosong (lain2) sebagai HADIR
        if lain2 > 0:
            hadir_days += lain2
            lain2 = 0

        if heb_total == 0 or denominator == 0:
            heb_zero_jenjangs.append(jenjang)
            percentage_values = {
                "hadir_pct": None,
                "sakit_pct": None,
                "izin_pct": None,
                "alfa_pct": None,
                "lain2_pct": None,
                "total_pct": None,
            }
        else:
            percentage_values = _normalize_percentage_row(
                _round_percentage_int(hadir_days, denominator),
                _round_percentage_int(sakit, denominator),
                _round_percentage_int(izin, denominator),
                _round_percentage_int(alfa, denominator),
                _round_percentage_int(lain2, denominator),
            )

        rows.append(
            {
                "jenjang": jenjang,
                "student_count": student_count,
                "heb": heb_total,
                "hadir_days": hadir_days,
                "hadir_pct": percentage_values["hadir_pct"],
                "sakit": sakit,
                "sakit_pct": percentage_values["sakit_pct"],
                "izin": izin,
                "izin_pct": percentage_values["izin_pct"],
                "alfa": alfa,
                "alfa_pct": percentage_values["alfa_pct"],
                "lain2": lain2,
                "lain2_pct": percentage_values["lain2_pct"],
                "total_pct": percentage_values["total_pct"],
            }
        )

    if heb_zero_jenjangs:
        warnings.append(
            "HEB belum tersedia untuk beberapa jenjang: " + ", ".join(heb_zero_jenjangs) + "."
        )

    if period_sia_entry_count == 0:
        warnings.append("Data Sakit/Izin/Alfa belum diisi untuk periode ini.")

    average_fields = ["hadir_pct", "sakit_pct", "izin_pct", "alfa_pct", "lain2_pct"]
    rata2_base = {}
    for field in average_fields:
        values = [row[field] for row in rows if row[field] is not None]
        rata2_base[field] = _round_half_up(sum(values) / len(values)) if values else None

    rata2 = _normalize_percentage_row(
        rata2_base["hadir_pct"],
        rata2_base["sakit_pct"],
        rata2_base["izin_pct"],
        rata2_base["alfa_pct"],
        rata2_base["lain2_pct"],
    )

    chart_data = [
        {"label": "Hadir", "value": rata2["hadir_pct"] if rata2["hadir_pct"] is not None else 0},
        {"label": "Sakit", "value": rata2["sakit_pct"] if rata2["sakit_pct"] is not None else 0},
        {"label": "Izin", "value": rata2["izin_pct"] if rata2["izin_pct"] is not None else 0},
        {"label": "Alfa", "value": rata2["alfa_pct"] if rata2["alfa_pct"] is not None else 0},
        {"label": "Lain2", "value": rata2["lain2_pct"] if rata2["lain2_pct"] is not None else 0},
    ]

    return {
        "report_title": _REKAP_ABSENSI_TITLE,
        "school_name": _SCHOOL_NAME,
        "period": {
            "label": period["label"],
            "date_from": period["date_from"].isoformat(),
            "date_to": period["date_to"].isoformat(),
            "term": period.get("term"),
            "year": period.get("year"),
        },
        "rows": rows,
        "heb_by_jenjang": {row["jenjang"]: row["heb"] for row in rows},
        "rata2": rata2,
        "chart_data": chart_data,
        "warnings": warnings,
        "warning_flags": {
            "heb_missing": bool(heb_zero_jenjangs),
            "sia_missing": period_sia_entry_count == 0,
        },
        "meta": {
            "classes_with_sia": sum(item.get("classes_with_sia", 0) for item in absence_map.values()),
            "classes_total": sum(class_count_map.values()),
        },
    }


def _style_rekap_sheet_title_row(ws, row_number: int, title: str, total_columns: int, *, bold: bool = False, size: int = 12):
    ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=total_columns)
    cell = ws.cell(row=row_number, column=1, value=title)
    cell.font = Font(bold=bold, size=size, color="0F172A")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_rekap_row(ws, row_number: int, total_columns: int, fill_color: str | None = None, bold: bool = False):
    for column in range(1, total_columns + 1):
        cell = ws.cell(row=row_number, column=column)
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill_color:
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
        if bold:
            cell.font = Font(bold=True, color="FFFFFF" if fill_color in {"2E7D32", "1B5E20"} else "0F172A")


def _build_rekap_absensi_workbook(report_data: dict):
    workbook = Workbook()

    summary_sheet = workbook.active
    if summary_sheet is None:
        summary_sheet = workbook.create_sheet()
    summary_sheet.title = "Rekap Absensi"

    summary_headers = ["KELAS", "HEB", "HADIR", "SAKIT", "IZIN", "ALFA", "LAIN2", "TOTAL"]
    total_columns = len(summary_headers)

    _style_rekap_sheet_title_row(summary_sheet, 1, report_data["report_title"].upper(), total_columns, bold=True, size=14)
    _style_rekap_sheet_title_row(summary_sheet, 2, report_data["period"]["label"], total_columns)
    _style_rekap_sheet_title_row(summary_sheet, 3, report_data["school_name"], total_columns)
    summary_sheet.append([])
    summary_sheet.append(summary_headers)

    header_row = 5
    for col_index, header in enumerate(summary_headers, start=1):
        cell = summary_sheet.cell(row=header_row, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="2E7D32")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    current_row = 6
    alternating_fills = ["FFFFFF", "E8F5E9"]
    for index, row in enumerate(report_data["rows"]):
        values = [
            row["jenjang"],
            row["heb"],
            _format_rekap_excel_pct(row["hadir_pct"]),
            _format_rekap_excel_pct(row["sakit_pct"]),
            _format_rekap_excel_pct(row["izin_pct"]),
            _format_rekap_excel_pct(row["alfa_pct"]),
            _format_rekap_excel_pct(row["lain2_pct"]),
            _format_rekap_excel_pct(row["total_pct"]),
        ]
        summary_sheet.append(values)
        fill = alternating_fills[index % 2]
        _style_rekap_row(summary_sheet, current_row, total_columns, fill_color=fill)
        summary_sheet.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        for col_index in range(2, total_columns + 1):
            summary_sheet.cell(row=current_row, column=col_index).alignment = Alignment(horizontal="right", vertical="center")
        current_row += 1

    summary_sheet.append(
        [
            "RATA2",
            "-",
            _format_rekap_excel_pct(report_data["rata2"]["hadir_pct"]),
            _format_rekap_excel_pct(report_data["rata2"]["sakit_pct"]),
            _format_rekap_excel_pct(report_data["rata2"]["izin_pct"]),
            _format_rekap_excel_pct(report_data["rata2"]["alfa_pct"]),
            _format_rekap_excel_pct(report_data["rata2"]["lain2_pct"]),
            _format_rekap_excel_pct(report_data["rata2"]["total_pct"]),
        ]
    )
    rata2_row = summary_sheet.max_row
    _style_rekap_row(summary_sheet, rata2_row, total_columns, fill_color="1B5E20", bold=True)
    summary_sheet.freeze_panes = "A6"
    _auto_size_worksheet_columns(summary_sheet)

    detail_sheet = workbook.create_sheet("Detail")
    detail_headers = ["KELAS", "SISWA", "HEB", "HADIR (hari)", "SAKIT", "IZIN", "ALFA", "LAIN2"]
    detail_columns = len(detail_headers)
    _style_rekap_sheet_title_row(detail_sheet, 1, report_data["report_title"].upper(), detail_columns, bold=True, size=14)
    _style_rekap_sheet_title_row(detail_sheet, 2, report_data["period"]["label"], detail_columns)
    _style_rekap_sheet_title_row(detail_sheet, 3, report_data["school_name"], detail_columns)
    detail_sheet.append([])
    detail_sheet.append(detail_headers)

    for col_index, header in enumerate(detail_headers, start=1):
        cell = detail_sheet.cell(row=5, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="2E7D32")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    detail_row = 6
    for index, row in enumerate(report_data["rows"]):
        detail_sheet.append(
            [
                row["jenjang"],
                row["student_count"],
                row["heb"],
                row["hadir_days"],
                row["sakit"],
                row["izin"],
                row["alfa"],
                row["lain2"],
            ]
        )
        _style_rekap_row(detail_sheet, detail_row, detail_columns, fill_color=alternating_fills[index % 2])
        detail_row += 1

    detail_sheet.freeze_panes = "A6"
    _auto_size_worksheet_columns(detail_sheet)
    return workbook


def _build_tardiness_base_cte(db: Session, start_date: date, end_date: date, jenjang: str | None = None):
    effective_status = func.coalesce(AttendanceOverride.override_status, Attendance.status)
    stmt = (
        select(
            Attendance.student_id.label("student_id"),
            Attendance.date.label("attendance_date"),
            Student.name.label("name"),
            _class_name_label_expression().label("class_name"),
            _report_jenjang_expression(db).label("jenjang"),
            func.coalesce(Attendance.late_duration, 0).label("late_duration"),
        )
        .select_from(Attendance)
        .join(Student, Student.id == Attendance.student_id)
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .where(
            effective_status == "late",
            Attendance.date >= start_date,
            Attendance.date <= end_date,
        )
    )

    if jenjang and jenjang.strip().lower() != "all":
        stmt = stmt.where(_report_jenjang_expression(db) == jenjang.strip().upper())

    return stmt.cte("tardiness_base")


def _collect_tardiness_report_data(db: Session, period: dict, jenjang: str | None = None, include_student_detail: bool = False):
    base_cte = _build_tardiness_base_cte(db, period["date_from"], period["date_to"], jenjang)
    absence_reason_map = _get_absence_reason_map_for_range(db, period["date_from"], period["date_to"])
    tracked_school_days_stmt = select(func.count(func.distinct(Attendance.date))).where(
        Attendance.date >= period["date_from"],
        Attendance.date <= period["date_to"],
        Attendance.status != "skipped",
    )

    totals_minutes = func.coalesce(func.sum(base_cte.c.late_duration), 0)
    totals_days = func.count()
    unique_late_days = func.count(func.distinct(base_cte.c.attendance_date))
    totals_students = func.count(func.distinct(base_cte.c.student_id))
    totals_avg_minutes = cast(
        func.round((cast(totals_minutes, Float) / func.nullif(cast(totals_days, Float), 0)), 0),
        Integer,
    )

    totals_stmt = select(
        totals_minutes.label("total_late_duration_minutes"),
        _minutes_to_hhmm_expression(db, totals_minutes).label("total_late_duration_str"),
        totals_days.label("total_days_late"),
        unique_late_days.label("unique_late_days"),
        totals_students.label("total_students_ever_late"),
        _minutes_to_hhmm_expression(db, totals_avg_minutes).label("average_late_duration_str"),
    ).select_from(base_cte)

    jenjang_group_cte = (
        select(
            base_cte.c.jenjang.label("jenjang"),
            func.sum(base_cte.c.late_duration).label("total_late_duration_minutes"),
            func.count().label("total_days_late"),
            func.count(func.distinct(base_cte.c.student_id)).label("late_student_count"),
        )
        .group_by(base_cte.c.jenjang)
        .cte("jenjang_group")
    )
    jenjang_totals_cte = select(
        func.coalesce(func.sum(jenjang_group_cte.c.total_late_duration_minutes), 0).label("grand_total_minutes"),
        func.coalesce(func.sum(jenjang_group_cte.c.total_days_late), 0).label("grand_total_days"),
    ).cte("jenjang_totals")

    jenjang_stmt = (
        select(
            jenjang_group_cte.c.jenjang,
            jenjang_group_cte.c.total_late_duration_minutes,
            _minutes_to_hhmm_expression(db, jenjang_group_cte.c.total_late_duration_minutes).label("total_late_duration_str"),
            _percentage_expression(
                jenjang_group_cte.c.total_late_duration_minutes,
                jenjang_totals_cte.c.grand_total_minutes,
            ).label("late_duration_pct"),
            jenjang_group_cte.c.total_days_late,
            _percentage_expression(
                jenjang_group_cte.c.total_days_late,
                jenjang_totals_cte.c.grand_total_days,
            ).label("days_late_pct"),
            jenjang_group_cte.c.late_student_count,
        )
        .select_from(jenjang_group_cte)
        .join(jenjang_totals_cte, true())
        .order_by(jenjang_group_cte.c.jenjang.asc())
    )

    class_group_cte = (
        select(
            base_cte.c.class_name.label("class_name"),
            base_cte.c.jenjang.label("jenjang"),
            func.sum(base_cte.c.late_duration).label("total_late_duration_minutes"),
            func.count(func.distinct(base_cte.c.attendance_date)).label("total_days_late"),
            func.count(func.distinct(base_cte.c.student_id)).label("late_student_count"),
        )
        .group_by(base_cte.c.class_name, base_cte.c.jenjang)
        .cte("class_group")
    )
    class_totals_cte = select(
        func.coalesce(func.sum(class_group_cte.c.total_late_duration_minutes), 0).label("grand_total_minutes"),
        func.coalesce(func.sum(class_group_cte.c.total_days_late), 0).label("grand_total_days"),
    ).cte("class_totals")

    class_stmt = (
        select(
            class_group_cte.c.class_name,
            class_group_cte.c.jenjang,
            class_group_cte.c.total_late_duration_minutes,
            _minutes_to_hhmm_expression(db, class_group_cte.c.total_late_duration_minutes).label("total_late_duration_str"),
            _percentage_expression(
                class_group_cte.c.total_late_duration_minutes,
                class_totals_cte.c.grand_total_minutes,
            ).label("late_duration_pct"),
            class_group_cte.c.total_days_late,
            _percentage_expression(
                class_group_cte.c.total_days_late,
                class_totals_cte.c.grand_total_days,
            ).label("days_late_pct"),
            class_group_cte.c.late_student_count,
        )
        .select_from(class_group_cte)
        .join(class_totals_cte, true())
        .order_by(class_group_cte.c.jenjang.asc(), class_group_cte.c.class_name.asc())
    )

    totals_row = db.execute(totals_stmt).mappings().one()
    tracked_school_days = int(db.execute(tracked_school_days_stmt).scalar() or 0)
    total_late_incidents = int(totals_row["total_days_late"] or 0)
    total_unique_late_days = int(totals_row["unique_late_days"] or 0)
    school_impact_rate_pct = round((total_unique_late_days / tracked_school_days) * 100, 1) if tracked_school_days > 0 else 0.0
    average_lateness_density = round(total_late_incidents / total_unique_late_days, 2) if total_unique_late_days > 0 else 0.0
    jenjang_rows = [dict(row) for row in db.execute(jenjang_stmt).mappings().all()]
    class_rows = [dict(row) for row in db.execute(class_stmt).mappings().all()]
    for row in class_rows:
        absence_values = absence_reason_map.get(
            row["class_name"],
            {"sakit": 0, "izin": 0, "alfa": 0, "total_absence_reasons": 0},
        )
        row.update(absence_values)

    month_pairs = _month_pairs_in_range(period["date_from"], period["date_to"])
    jenjang_expr = _report_jenjang_expression(db)
    jenjang_mapping = db.execute(
        select(jenjang_expr.label("normalized"), func.trim(Student.jenjang).label("original"))
        .filter(Student.jenjang.isnot(None))
        .distinct()
    ).mappings().all()

    heb_by_jenjang = {}
    for mapping in jenjang_mapping:
        norm_j = mapping["normalized"]
        orig_j = mapping["original"]
        heb_total = sum(int(calculate_heb(db, orig_j, pm, py)["heb"] or 0) for py, pm in month_pairs)
        if heb_total > 0:
            heb_by_jenjang[norm_j] = heb_total

    payload = {
        "report_title": _REPORT_TITLE,
        "school_name": _SCHOOL_NAME,
        "period": {
            "label": period["label"],
            "date_from": period["date_from"].isoformat(),
            "date_to": period["date_to"].isoformat(),
        },
        "heb_by_jenjang": heb_by_jenjang,
        "summary_by_jenjang": jenjang_rows,
        "breakdown_by_class": class_rows,
        "totals": {
            "total_late_duration_minutes": totals_row["total_late_duration_minutes"],
            "total_late_duration_str": totals_row["total_late_duration_str"],
            "total_days_late": totals_row["total_days_late"],
            "total_late_incidents": total_late_incidents,
            "unique_late_days": total_unique_late_days,
            "tracked_school_days": tracked_school_days,
            "school_impact_rate_pct": school_impact_rate_pct,
            "average_lateness_density": average_lateness_density,
            "total_students_ever_late": totals_row["total_students_ever_late"],
        },
        "management_summary": {
            "total_late_incidents": total_late_incidents,
            "unique_late_days": total_unique_late_days,
            "tracked_school_days": tracked_school_days,
            "school_impact_rate_pct": school_impact_rate_pct,
            "average_lateness_density": average_lateness_density,
        },
    }

    if include_student_detail:
        student_group_cte = (
            select(
                base_cte.c.student_id.label("student_id"),
                base_cte.c.name.label("name"),
                base_cte.c.class_name.label("class_name"),
                base_cte.c.jenjang.label("jenjang"),
                func.count().label("total_days_late"),
                func.sum(base_cte.c.late_duration).label("total_late_duration_minutes"),
            )
            .group_by(base_cte.c.student_id, base_cte.c.name, base_cte.c.class_name, base_cte.c.jenjang)
            .cte("student_group")
        )
        average_duration_minutes = cast(
            func.round(
                cast(student_group_cte.c.total_late_duration_minutes, Float)
                / func.nullif(cast(student_group_cte.c.total_days_late, Float), 0),
                0,
            ),
            Integer,
        )
        student_stmt = (
            select(
                student_group_cte.c.student_id.label("no_id"),
                student_group_cte.c.name.label("nama"),
                student_group_cte.c.class_name.label("kelas"),
                student_group_cte.c.jenjang,
                student_group_cte.c.total_days_late,
                _minutes_to_hhmm_expression(db, student_group_cte.c.total_late_duration_minutes).label("total_durasi"),
                _minutes_to_hhmm_expression(db, average_duration_minutes).label("rata_rata_durasi"),
            )
            .order_by(student_group_cte.c.jenjang.asc(), student_group_cte.c.class_name.asc(), student_group_cte.c.name.asc())
        )
        student_sia_map = _get_student_absence_reason_map_for_range(db, period["date_from"], period["date_to"])
        student_details = []
        for row in db.execute(student_stmt).mappings().all():
            d = dict(row)
            sia = student_sia_map.get(d["no_id"], {"sakit": 0, "izin": 0, "alfa": 0})
            d.update(sia)
            student_details.append(d)
        payload["student_details"] = student_details
        payload["detail_summary"] = {
            "average_late_duration_str": totals_row["average_late_duration_str"],
        }

    return payload


def _collect_tardiness_summary_by_jenjang(db: Session, period: dict, jenjang: str | None = None) -> dict:
    base_cte = _build_tardiness_base_cte(db, period["date_from"], period["date_to"], jenjang)
    rows = db.execute(
        select(
            base_cte.c.attendance_date.label("Tanggal"),
            base_cte.c.jenjang.label("Jenjang"),
        )
        .order_by(base_cte.c.jenjang.asc(), base_cte.c.attendance_date.asc())
    ).mappings().all()

    dataframe = pd.DataFrame(rows, columns=["Tanggal", "Jenjang"])
    if dataframe.empty:
        summary_rows: list[dict] = []
    else:
        summary = (
            dataframe.groupby("Jenjang", dropna=False)
            .agg(
                total_kejadian=("Jenjang", "size"),
                hari_efektif_terlambat=("Tanggal", pd.Series.nunique),
            )
            .reset_index()
        )
        summary["rata_rata_siswa_terlambat_per_hari"] = (
            summary["total_kejadian"]
            .div(summary["hari_efektif_terlambat"].replace({0: pd.NA}))
            .round(1)
            .fillna(0.0)
        )
        grand_total_kejadian = summary["total_kejadian"].sum()
        summary["percentage_of_total"] = (
            summary["total_kejadian"]
            .div(grand_total_kejadian if grand_total_kejadian else pd.NA)
            .mul(100)
            .round(1)
            .fillna(0.0)
        )
        summary_rows = [
            {
                "jenjang": row["Jenjang"],
                "total_kejadian": int(row["total_kejadian"]),
                "hari_efektif_terlambat": int(row["hari_efektif_terlambat"]),
                "rata_rata_siswa_terlambat_per_hari": float(row["rata_rata_siswa_terlambat_per_hari"]),
                "percentage_of_total": float(row["percentage_of_total"]),
            }
            for row in summary.sort_values("Jenjang", kind="stable").to_dict(orient="records")
        ]

    return {
        "period": {
            "label": period["label"],
            "date_from": period["date_from"].isoformat(),
            "date_to": period["date_to"].isoformat(),
        },
        "rows": summary_rows,
    }


def _apply_excel_title_and_headers(ws, title: str, headers: list[str]):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="DCEAFE")
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A3"


def _auto_size_worksheet_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)


def _append_total_row(ws, values: list):
    ws.append(values)
    total_row_index = ws.max_row
    for cell in ws[total_row_index]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="FEF3C7")


def _format_excel_percentage(value: float) -> str:
    return f"{value:.1f}%".replace(".", ",")


def _format_excel_float(value: float, decimals: int = 1) -> str:
    return f"{value:.{decimals}f}".replace(".", ",")


def _append_jenjang_summary_notes(ws):
    ws.append([])
    ws.append(["How to Read This Table", "Formula / Explanation"])
    notes_start_row = ws.max_row
    ws.cell(row=notes_start_row, column=1).font = Font(bold=True)
    ws.cell(row=notes_start_row, column=2).font = Font(bold=True)
    ws.cell(row=notes_start_row, column=1).fill = PatternFill(fill_type="solid", fgColor="DCEAFE")
    ws.cell(row=notes_start_row, column=2).fill = PatternFill(fill_type="solid", fgColor="DCEAFE")

    note_rows = [
        [
            "Total Late Incidents",
            "Count of all late records for the level. If 10 students are late on the same day, that counts as 10 incidents.",
        ],
        [
            "Percentage of Total",
            "(Level Total Late Incidents / Grand Total Late Incidents) x 100",
        ],
        [
            "Effective Late Days",
            "Count of unique school dates where at least one student in the level arrived late.",
        ],
        [
            "Average Late Students/Day",
            "Total Late Incidents / Effective Late Days",
        ],
    ]
    for row in note_rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=notes_start_row + 1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _append_class_summary_notes(ws):
    ws.append([])
    ws.append(["How to Read This Table", "Formula / Explanation"])
    notes_start_row = ws.max_row
    ws.cell(row=notes_start_row, column=1).font = Font(bold=True)
    ws.cell(row=notes_start_row, column=2).font = Font(bold=True)
    ws.cell(row=notes_start_row, column=1).fill = PatternFill(fill_type="solid", fgColor="DCEAFE")
    ws.cell(row=notes_start_row, column=2).fill = PatternFill(fill_type="solid", fgColor="DCEAFE")

    note_rows = [
        [
            "HEB (Hari Efektif Belajar)",
            "The total number of school/learning days in the selected reporting period (45 days in Term 4).",
        ],
        [
            "Total Late Duration",
            "The accumulated amount of late time (in hours and minutes, HH:MM) for all students in that class during the term.",
        ],
        [
            "% Duration",
            "(Class Total Late Duration in minutes / Grand Total Late Duration in minutes) x 100",
        ],
        [
            "Unique Late Days",
            "The number of unique calendar dates on which at least one student from that class arrived late.",
        ],
        [
            "% Late Days",
            "(Class Unique Late Days / Sum of Unique Late Days across all individual classes) x 100",
        ],
        [
            "Late Students",
            "The number of distinct (unique) students in that class who were late at least once during the term.",
        ],
        [
            "TOTAL - Unique Late Days",
            "The count of unique dates on which at least one student in the entire school was late. Note: This is not a sum of the individual class columns, as multiple classes can have late students on the same calendar days.",
        ],
    ]
    for row in note_rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=notes_start_row + 1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_tardiness_management_workbook(report_data: dict, jenjang_summary_rows: list[dict]):
    workbook = Workbook()

    title = f"Executive Tardiness Report - {report_data['period']['label']} - {_SCHOOL_NAME}"
    management_summary = report_data["management_summary"]

    management_sheet = workbook.active
    if management_sheet is None:
        management_sheet = workbook.create_sheet()
    management_sheet.title = "Executive Summary"
    management_headers = ["Metric", "Value", "Definition"]
    _apply_excel_title_and_headers(management_sheet, title, management_headers)

    management_rows = [
        [
            "Total Incidents",
            management_summary["total_late_incidents"],
            "Total number of individual late entries in the selected period.",
        ],
        [
            "Unique Late Days",
            management_summary["unique_late_days"],
            "Number of unique dates where at least one student arrived late.",
        ],
        [
            "Tracked School Days",
            management_summary["tracked_school_days"],
            "Number of recorded school days in the selected reporting period.",
        ],
        [
            "School Impact Rate",
            _format_excel_percentage(management_summary["school_impact_rate_pct"]),
            "Percentage of school days affected by at least one late incident.",
        ],
        [
            "Average Lateness Density",
            _format_excel_float(management_summary["average_lateness_density"], 2),
            "Average number of late students on each affected day.",
        ],
    ]
    for row in management_rows:
        management_sheet.append(row)
    _auto_size_worksheet_columns(management_sheet)

    jenjang_sheet = workbook.create_sheet("Jenjang Late Summary")
    jenjang_headers = [
        "Level",
        "HEB",
        "Total Late Incidents",
        "Percentage of Total",
        "Effective Late Days",
        "Average Late Students/Day",
    ]
    _apply_excel_title_and_headers(jenjang_sheet, title, jenjang_headers)
    for row in jenjang_summary_rows:
        jenjang_sheet.append(
            [
                row["jenjang"],
                report_data.get("heb_by_jenjang", {}).get(row["jenjang"], "-"),
                row["total_kejadian"],
                _format_excel_percentage(row["percentage_of_total"]),
                row["hari_efektif_terlambat"],
                _format_excel_float(row["rata_rata_siswa_terlambat_per_hari"]),
            ]
        )
    if jenjang_summary_rows:
        grand_total_incidents = sum(r["total_kejadian"] for r in jenjang_summary_rows)
        grand_effective_days = sum(r["hari_efektif_terlambat"] for r in jenjang_summary_rows)
        grand_avg = round(grand_total_incidents / grand_effective_days, 1) if grand_effective_days else 0.0
        _append_total_row(
            jenjang_sheet,
            ["TOTAL", "-", grand_total_incidents, "100,0%", grand_effective_days, _format_excel_float(grand_avg)],
        )
    _append_jenjang_summary_notes(jenjang_sheet)
    _auto_size_worksheet_columns(jenjang_sheet)

    return workbook


def _build_tardiness_report_workbook(report_data: dict, jenjang_summary_rows: list[dict]):
    workbook = Workbook()

    title = f"{_REPORT_TITLE} - {report_data['period']['label']} - {_SCHOOL_NAME}"
    total_minutes = report_data["totals"]["total_late_duration_minutes"]
    total_days = report_data["totals"]["total_late_incidents"]
    total_students = report_data["totals"]["total_students_ever_late"]
    management_summary = report_data["management_summary"]

    management_sheet = workbook.active
    if management_sheet is None:
        management_sheet = workbook.create_sheet()
    management_sheet.title = "Management Summary"
    management_headers = ["Metric", "Value", "Definition"]
    _apply_excel_title_and_headers(management_sheet, title, management_headers)
    management_rows = [
        [
            "Total Incidents",
            management_summary["total_late_incidents"],
            "Total number of individual late entries in the selected period.",
        ],
        [
            "Unique Late Days",
            management_summary["unique_late_days"],
            "Number of unique dates where at least one student arrived late.",
        ],
        [
            "Tracked School Days",
            management_summary["tracked_school_days"],
            "Number of recorded school days in the selected reporting period.",
        ],
        [
            "School Impact Rate",
            _format_excel_percentage(management_summary["school_impact_rate_pct"]),
            "Percentage of school days affected by at least one late incident.",
        ],
        [
            "Average Lateness Density",
            _format_excel_float(management_summary["average_lateness_density"], 2),
            "Average number of late students on each affected day.",
        ],
    ]
    for row in management_rows:
        management_sheet.append(row)
    _auto_size_worksheet_columns(management_sheet)

    jenjang_sheet = workbook.create_sheet("Summary by Jenjang")
    jenjang_headers = [
        "Level",
        "HEB",
        "Total Late Incidents",
        "Percentage of Total",
        "Effective Late Days",
        "Average Late Students/Day",
    ]
    _apply_excel_title_and_headers(jenjang_sheet, title, jenjang_headers)
    for row in jenjang_summary_rows:
        jenjang_sheet.append(
            [
                row["jenjang"],
                report_data.get("heb_by_jenjang", {}).get(row["jenjang"], "-"),
                row["total_kejadian"],
                _format_excel_percentage(row["percentage_of_total"]),
                row["hari_efektif_terlambat"],
                _format_excel_float(row["rata_rata_siswa_terlambat_per_hari"]),
            ]
        )
    if jenjang_summary_rows:
        grand_total_incidents = sum(r["total_kejadian"] for r in jenjang_summary_rows)
        grand_effective_days = sum(r["hari_efektif_terlambat"] for r in jenjang_summary_rows)
        grand_avg = round(grand_total_incidents / grand_effective_days, 1) if grand_effective_days else 0.0
        _append_total_row(
            jenjang_sheet,
            ["TOTAL", "-", grand_total_incidents, "100,0%", grand_effective_days, _format_excel_float(grand_avg)],
        )
    _append_jenjang_summary_notes(jenjang_sheet)
    _auto_size_worksheet_columns(jenjang_sheet)

    class_sheet = workbook.create_sheet("Class Breakdown")
    class_headers = [
        "Class",
        "Level",
        "HEB",
        "Total Late Duration",
        "% Duration",
        "Unique Late Days",
        "% Late Days",
        "Late Students",
    ]
    _apply_excel_title_and_headers(class_sheet, title, class_headers)
    for row in report_data["breakdown_by_class"]:
        class_sheet.append(
            [
                row["class_name"],
                row["jenjang"],
                report_data.get("heb_by_jenjang", {}).get(row["jenjang"], "-"),
                row["total_late_duration_str"],
                _format_excel_percentage(row["late_duration_pct"]),
                row["total_days_late"],
                _format_excel_percentage(row["days_late_pct"]),
                row["late_student_count"],
            ]
        )
    _append_total_row(
        class_sheet,
        [
            "TOTAL",
            "",
            "-",
            report_data["totals"]["total_late_duration_str"],
            _format_excel_percentage(100.0 if total_minutes > 0 else 0.0),
            report_data["totals"]["unique_late_days"],
            _format_excel_percentage(100.0 if total_minutes > 0 else 0.0),
            total_students,
        ],
    )
    _append_class_summary_notes(class_sheet)
    _auto_size_worksheet_columns(class_sheet)

    detail_sheet = workbook.create_sheet("Student Details")
    detail_headers = [
        "Student ID",
        "Name",
        "Class",
        "Level",
        "Total Incidents",
        "Total Duration",
        "Average Duration",
    ]
    _apply_excel_title_and_headers(detail_sheet, title, detail_headers)
    for row in report_data.get("student_details", []):
        detail_sheet.append(
            [
                row["no_id"],
                row["nama"],
                row["kelas"],
                row["jenjang"],
                row["total_days_late"],
                row["total_durasi"],
                row["rata_rata_durasi"],
            ]
        )
    _append_total_row(
        detail_sheet,
        [
            "TOTAL",
            "",
            "",
            "",
            total_days,
            report_data["totals"]["total_late_duration_str"],
            "-",
        ],
    )
    _auto_size_worksheet_columns(detail_sheet)

    return workbook


@router.get("/tardiness-report")
def get_tardiness_report(
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=1900),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    term: int | None = Query(None, ge=1, le=4),
    jenjang: str | None = Query(None),
    db: Session = Depends(get_db),
):
    period = _resolve_tardiness_period(month, year, date_from, date_to, term)
    return _collect_tardiness_report_data(db, period, jenjang, include_student_detail=True)


@router.get("/tardiness/summary-by-jenjang")
@router.get("/tardiness-report/summary-by-jenjang")
def get_tardiness_summary_by_jenjang(
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=1900),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    term: int | None = Query(None, ge=1, le=4),
    jenjang: str | None = Query(None),
    db: Session = Depends(get_db),
):
    period = _resolve_tardiness_period(month, year, date_from, date_to, term)
    return _collect_tardiness_summary_by_jenjang(db, period, jenjang)


@router.get("/tardiness-report/export-excel")
def export_tardiness_report_excel(
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=1900),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    term: int | None = Query(None, ge=1, le=4),
    jenjang: str | None = Query(None),
    db: Session = Depends(get_db),
):
    period = _resolve_tardiness_period(month, year, date_from, date_to, term)
    report_data = _collect_tardiness_report_data(db, period, jenjang, include_student_detail=True)
    jenjang_summary_payload = _collect_tardiness_summary_by_jenjang(db, period, jenjang)
    workbook = _build_tardiness_report_workbook(report_data, jenjang_summary_payload["rows"])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    filename = f"tardiness_report_{_report_period_slug(report_data['period']['label'])}.xlsx"
    return StreamingResponse(
        output,
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/tardiness-report/export-management-excel")
def export_tardiness_management_excel(
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=1900),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    term: int | None = Query(None, ge=1, le=4),
    jenjang: str | None = Query(None),
    db: Session = Depends(get_db),
):
    period = _resolve_tardiness_period(month, year, date_from, date_to, term)
    report_data = _collect_tardiness_report_data(db, period, jenjang, include_student_detail=False)
    jenjang_summary_payload = _collect_tardiness_summary_by_jenjang(db, period, jenjang)
    workbook = _build_tardiness_management_workbook(report_data, jenjang_summary_payload["rows"])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    filename = f"executive_tardiness_summary_{_report_period_slug(report_data['period']['label'])}.xlsx"
    return StreamingResponse(
        output,
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



def _normalize_v2_percentage_dict(pcts: dict[str, float | None]) -> dict[str, float | None]:
    if any(v is None for v in pcts.values()):
        return {k: None for k in pcts}

    total_pct = sum(pcts.values())
    if abs(total_pct - 100) > 0.001:
        diff = 100.0 - total_pct
        max_key = max(pcts, key=pcts.get)
        pcts[max_key] = max(0.0, pcts[max_key] + diff)

    pcts["total_pct"] = 100.0
    return pcts

def _collect_v2_rekap_absensi_report_data(db: Session, period: dict):
    month_pairs = _month_pairs_in_range(period["date_from"], period["date_to"])
    jenjang_expr = _report_jenjang_expression(db)
    raw_jenjang_expr = func.trim(Student.jenjang)
    class_expr = func.trim(Student.class_name)
    effective_status = func.coalesce(AttendanceOverride.override_status, Attendance.status)

    student_count_stmt = (
        select(
            jenjang_expr.label("jenjang"),
            raw_jenjang_expr.label("raw_jenjang"),
            class_expr.label("class_name"),
            func.count(Student.id).label("student_count"),
        )
        .select_from(Student)
        .where(_valid_student_jenjang_filter(), _valid_student_class_filter())
        .group_by(jenjang_expr, raw_jenjang_expr, class_expr)
    )
    student_count_rows = db.execute(student_count_stmt).mappings().all()

    classes_data = {}
    jenjang_source_map = {}
    for row in student_count_rows:
        count = int(row["student_count"] or 0)
        if count <= 0:
            continue
        jenjang = row["jenjang"]
        cls_name = row["class_name"]

        if jenjang not in classes_data:
            classes_data[jenjang] = {}
        classes_data[jenjang][cls_name] = {"student_count": count}
        jenjang_source_map[jenjang] = row["raw_jenjang"]

    attendance_stmt = (
        select(
            jenjang_expr.label("jenjang"),
            class_expr.label("class_name"),
            func.count(Attendance.id).label("hadir_days"),
        )
        .select_from(Attendance)
        .join(Student, Student.id == Attendance.student_id)
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .where(
            Attendance.date >= period["date_from"],
            Attendance.date <= period["date_to"],
            _valid_student_jenjang_filter(),
            _valid_student_class_filter(),
            effective_status.in_(("on-time", "late")),
        )
        .group_by(jenjang_expr, class_expr)
    )
    attendance_rows = db.execute(attendance_stmt).mappings().all()
    for row in attendance_rows:
        j = row["jenjang"]
        c = row["class_name"]
        if j in classes_data and c in classes_data[j]:
            classes_data[j][c]["hadir_days"] = int(row["hadir_days"] or 0)

    absence_stmt = (
        select(
            func.trim(AbsenceReasonClassEntry.class_name).label("class_name"),
            func.coalesce(func.sum(AbsenceReasonClassEntry.sakit), 0).label("sakit"),
            func.coalesce(func.sum(AbsenceReasonClassEntry.izin), 0).label("izin"),
            func.coalesce(func.sum(AbsenceReasonClassEntry.alfa), 0).label("alfa"),
        )
        .select_from(AbsenceReasonClassEntry)
        .where(_month_pair_filters(AbsenceReasonClassEntry, month_pairs))
        .group_by(func.trim(AbsenceReasonClassEntry.class_name))
    )
    absence_rows = db.execute(absence_stmt).mappings().all()
    sia_map = {row["class_name"]: {"sakit": int(row["sakit"]), "izin": int(row["izin"]), "alfa": int(row["alfa"])} for row in absence_rows}

    period_sia_entry_count = (
        db.execute(
            select(func.count(AbsenceReasonClassEntry.id)).where(_month_pair_filters(AbsenceReasonClassEntry, month_pairs))
        ).scalar() or 0
    )

    heb_cache = {}
    heb_zero_jenjangs = set()
    warnings = []
    has_data_quality_issue = False
    affected_classes = 0

    jenjang_results = []

    for jenjang in sorted(classes_data.keys()):
        raw_jenjang = jenjang_source_map[jenjang]
        heb_total = 0
        for pair_year, pair_month in month_pairs:
            cache_key = (raw_jenjang, pair_year, pair_month)
            if cache_key not in heb_cache:
                heb_cache[cache_key] = int(calculate_heb(db, raw_jenjang, pair_month, pair_year)["heb"] or 0)
            heb_total += heb_cache[cache_key]

        if heb_total == 0:
            heb_zero_jenjangs.add(jenjang)

        jenjang_classes = []

        sum_h = sum_s = sum_i = sum_a = sum_lain2 = sum_total = 0

        for cls_name in sorted(classes_data[jenjang].keys()):
            student_count = classes_data[jenjang][cls_name]["student_count"]
            hadir = classes_data[jenjang][cls_name].get("hadir_days", 0)

            sia = sia_map.get(cls_name, {"sakit": 0, "izin": 0, "alfa": 0})
            sakit = sia["sakit"]
            izin = sia["izin"]
            alfa = sia["alfa"]

            valid_total = hadir + sakit + izin + alfa
            expected_total = student_count * heb_total

            lain2 = 0
            flags = {}
            if heb_total > 0:
                lain2 = max(0, expected_total - valid_total)

                # Asumsikan data kosong (lain2) sebagai HADIR
                if lain2 > 0:
                    hadir += lain2
                    valid_total += lain2
                    lain2 = 0
            else:
                flags["expected_total_missing"] = True

            if valid_total == 0:
                flags["no_valid_data"] = True
                flags["data_quality_issue"] = True
                pcts = {"hadir_pct": None, "sakit_pct": None, "izin_pct": None, "alfa_pct": None, "total_pct": None}
            else:
                pcts = _normalize_v2_percentage_dict({
                    "hadir_pct": _round_percentage_int(hadir, valid_total),
                    "sakit_pct": _round_percentage_int(sakit, valid_total),
                    "izin_pct": _round_percentage_int(izin, valid_total),
                    "alfa_pct": _round_percentage_int(alfa, valid_total),
                })

            if lain2 > 0:
                flags["excluded_unclassified"] = True
                flags["lain2_count"] = lain2
                ratio = lain2 / (valid_total + lain2)
                if ratio > 0.1:
                    flags["data_quality_issue"] = True

            if flags.get("data_quality_issue"):
                has_data_quality_issue = True
                affected_classes += 1

            sum_h += hadir
            sum_s += sakit
            sum_i += izin
            sum_a += alfa
            sum_lain2 += lain2
            sum_total += valid_total

            jenjang_classes.append({
                "class_name": cls_name,
                "student_count": student_count,
                "hadir": hadir,
                "sakit": sakit,
                "izin": izin,
                "alfa": alfa,
                "lain2": lain2,
                "total": valid_total,
                "percentages": pcts,
                "warning_flags": flags,
            })

        j_pcts = {"hadir_pct": None, "sakit_pct": None, "izin_pct": None, "alfa_pct": None, "total_pct": None}
        if sum_total > 0:
            j_pcts = _normalize_v2_percentage_dict({
                "hadir_pct": _round_percentage_int(sum_h, sum_total),
                "sakit_pct": _round_percentage_int(sum_s, sum_total),
                "izin_pct": _round_percentage_int(sum_i, sum_total),
                "alfa_pct": _round_percentage_int(sum_a, sum_total),
            })

        jenjang_results.append({
            "name": jenjang,
            "classes": jenjang_classes,
            "summary": {
                "hadir": sum_h,
                "sakit": sum_s,
                "izin": sum_i,
                "alfa": sum_a,
                "lain2": sum_lain2,
                "total": sum_total,
                "heb": heb_total,
                "percentages": j_pcts,
            }
        })

    global_h = sum(j["summary"]["hadir"] for j in jenjang_results)
    global_s = sum(j["summary"]["sakit"] for j in jenjang_results)
    global_i = sum(j["summary"]["izin"] for j in jenjang_results)
    global_a = sum(j["summary"]["alfa"] for j in jenjang_results)
    global_lain2 = sum(j["summary"]["lain2"] for j in jenjang_results)
    global_total = sum(j["summary"]["total"] for j in jenjang_results)

    global_pcts = {"hadir_pct": None, "sakit_pct": None, "izin_pct": None, "alfa_pct": None, "total_pct": None}
    if global_total > 0:
        global_pcts = _normalize_v2_percentage_dict({
            "hadir_pct": _round_percentage_int(global_h, global_total),
            "sakit_pct": _round_percentage_int(global_s, global_total),
            "izin_pct": _round_percentage_int(global_i, global_total),
            "alfa_pct": _round_percentage_int(global_a, global_total),
        })

    global_summary = {
        "hadir": global_h,
        "sakit": global_s,
        "izin": global_i,
        "alfa": global_a,
        "lain2": global_lain2,
        "total": global_total,
        "percentages": global_pcts,
    }

    if heb_zero_jenjangs:
        warnings.append("HEB belum tersedia untuk beberapa jenjang: " + ", ".join(heb_zero_jenjangs) + ".")
    if period_sia_entry_count == 0:
        warnings.append("Data Sakit/Izin/Alfa belum diisi untuk periode ini.")

    chart_data = [
        {"label": "Hadir", "value": global_pcts["hadir_pct"] if global_pcts["hadir_pct"] is not None else 0},
        {"label": "Sakit", "value": global_pcts["sakit_pct"] if global_pcts["sakit_pct"] is not None else 0},
        {"label": "Izin", "value": global_pcts["izin_pct"] if global_pcts["izin_pct"] is not None else 0},
        {"label": "Alfa", "value": global_pcts["alfa_pct"] if global_pcts["alfa_pct"] is not None else 0},
    ]

    return {
        "report_title": _REKAP_ABSENSI_TITLE,
        "school_name": _SCHOOL_NAME,
        "period": {
            "label": period["label"],
            "date_from": period["date_from"].isoformat(),
            "date_to": period["date_to"].isoformat(),
            "term": period.get("term"),
            "year": period.get("year"),
        },
        "jenjang": jenjang_results,
        "heb_by_jenjang": {j["name"]: j["summary"]["heb"] for j in jenjang_results},
        "global_summary": global_summary,
        "chart_data": chart_data,
        "warnings": warnings,
        "global_flags": {
            "has_data_quality_issue": has_data_quality_issue,
            "affected_classes": affected_classes,
            "heb_missing": bool(heb_zero_jenjangs),
            "sia_missing": period_sia_entry_count == 0,
        }
    }

def _build_v2_rekap_absensi_workbook(report_data: dict):
    workbook = Workbook()

    summary_sheet = workbook.active
    if summary_sheet is None:
        summary_sheet = workbook.create_sheet()
    summary_sheet.title = "Rekap Absensi"

    summary_headers = ["JENJANG", "KELAS", "HEB", "HADIR", "SAKIT", "IZIN", "ALFA", "TOTAL"]
    total_columns = len(summary_headers)

    _style_rekap_sheet_title_row(summary_sheet, 1, report_data["report_title"].upper(), total_columns, bold=True, size=14)
    _style_rekap_sheet_title_row(summary_sheet, 2, report_data["period"]["label"], total_columns)
    _style_rekap_sheet_title_row(summary_sheet, 3, report_data["school_name"], total_columns)
    summary_sheet.append([])
    summary_sheet.append(summary_headers)

    header_row = 5
    for col_index, header in enumerate(summary_headers, start=1):
        cell = summary_sheet.cell(row=header_row, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="2E7D32")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    current_row = 6
    alternating_fills = ["FFFFFF", "E8F5E9"]

    for jenjang_idx, jenjang in enumerate(report_data["jenjang"]):
        for cls in jenjang["classes"]:
            pcts = cls["percentages"]
            values = [
                jenjang["name"],
                cls["class_name"],
                jenjang["summary"]["heb"],
                _format_rekap_excel_pct(pcts["hadir_pct"]),
                _format_rekap_excel_pct(pcts["sakit_pct"]),
                _format_rekap_excel_pct(pcts["izin_pct"]),
                _format_rekap_excel_pct(pcts["alfa_pct"]),
                _format_rekap_excel_pct(pcts["total_pct"]),
            ]
            summary_sheet.append(values)
            fill = alternating_fills[current_row % 2]
            _style_rekap_row(summary_sheet, current_row, total_columns, fill_color=fill)
            for col_index in range(1, 4):
                summary_sheet.cell(row=current_row, column=col_index).alignment = Alignment(horizontal="center", vertical="center")
            for col_index in range(4, total_columns + 1):
                summary_sheet.cell(row=current_row, column=col_index).alignment = Alignment(horizontal="right", vertical="center")
            current_row += 1

        # Subtotal row explicitly styled
        j_pcts = jenjang["summary"]["percentages"]
        summary_sheet.append([
            jenjang["name"],
            "SUBTOTAL",
            jenjang["summary"]["heb"],
            _format_rekap_excel_pct(j_pcts["hadir_pct"]),
            _format_rekap_excel_pct(j_pcts["sakit_pct"]),
            _format_rekap_excel_pct(j_pcts["izin_pct"]),
            _format_rekap_excel_pct(j_pcts["alfa_pct"]),
            _format_rekap_excel_pct(j_pcts["total_pct"]),
        ])
        _style_rekap_row(summary_sheet, current_row, total_columns, fill_color="C8E6C9", bold=True)
        current_row += 1

    gs_pcts = report_data["global_summary"]["percentages"]
    summary_sheet.append(
        [
            "GLOBAL",
            "RATA-RATA",
            "-",
            _format_rekap_excel_pct(gs_pcts["hadir_pct"]),
            _format_rekap_excel_pct(gs_pcts["sakit_pct"]),
            _format_rekap_excel_pct(gs_pcts["izin_pct"]),
            _format_rekap_excel_pct(gs_pcts["alfa_pct"]),
            _format_rekap_excel_pct(gs_pcts["total_pct"]),
        ]
    )
    rata2_row = summary_sheet.max_row
    _style_rekap_row(summary_sheet, rata2_row, total_columns, fill_color="1B5E20", bold=True)

    # Footnote
    summary_sheet.append([])
    summary_sheet.append(["*Data tidak terklasifikasi (LAIN2) tidak dimasukkan dalam perhitungan"])
    f_cell = summary_sheet.cell(row=summary_sheet.max_row, column=1)
    f_cell.font = Font(italic=True, size=10, color="64748b")

    summary_sheet.freeze_panes = "A6"
    _auto_size_worksheet_columns(summary_sheet)

    # Detail Sheet (with LAIN2 for audit)
    detail_sheet = workbook.create_sheet("Detail")
    detail_headers = ["JENJANG", "KELAS", "SISWA", "HEB", "HADIR (hari)", "SAKIT", "IZIN", "ALFA", "LAIN2", "TOTAL"]
    detail_columns = len(detail_headers)
    _style_rekap_sheet_title_row(detail_sheet, 1, report_data["report_title"].upper(), detail_columns, bold=True, size=14)
    _style_rekap_sheet_title_row(detail_sheet, 2, report_data["period"]["label"], detail_columns)
    _style_rekap_sheet_title_row(detail_sheet, 3, report_data["school_name"], detail_columns)
    detail_sheet.append([])
    detail_sheet.append(detail_headers)

    for col_index, header in enumerate(detail_headers, start=1):
        cell = detail_sheet.cell(row=5, column=col_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="2E7D32")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    detail_row = 6
    for jenjang in report_data["jenjang"]:
        for row in jenjang["classes"]:
            detail_sheet.append(
                [
                    jenjang["name"],
                    row["class_name"],
                    row["student_count"],
                    row.get("heb", "-"), # HEB isn't returned per class in dict right now, but let's just output -
                    row["hadir"],
                    row["sakit"],
                    row["izin"],
                    row["alfa"],
                    row["lain2"],
                    row["total"],
                ]
            )
            _style_rekap_row(detail_sheet, detail_row, detail_columns, fill_color=alternating_fills[detail_row % 2])
            detail_row += 1

    detail_sheet.append([])
    detail_sheet.append(["*Data tidak terklasifikasi (LAIN2) dihitung sebagai: (SISWA * HEB) - TOTAL (di mana TOTAL = HADIR + SAKIT + IZIN + ALFA)"])
    d_f_cell = detail_sheet.cell(row=detail_sheet.max_row, column=1)
    d_f_cell.font = Font(italic=True, size=10, color="64748b")

    detail_sheet.freeze_panes = "A6"
    _auto_size_worksheet_columns(detail_sheet)
    return workbook

@router.get("/v2/rekap-absensi")
def get_v2_rekap_absensi(
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=1900),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    term: int | None = Query(None, ge=1, le=4),
    db: Session = Depends(get_db),
):
    period = _resolve_rekap_absensi_period(month, year, date_from, date_to, term)
    return _collect_v2_rekap_absensi_report_data(db, period)


@router.get("/v2/rekap-absensi/export-excel")
def export_v2_rekap_absensi_excel(
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=1900),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    term: int | None = Query(None, ge=1, le=4),
    db: Session = Depends(get_db),
):
    period = _resolve_rekap_absensi_period(month, year, date_from, date_to, term)
    report_data = _collect_v2_rekap_absensi_report_data(db, period)
    workbook = _build_v2_rekap_absensi_workbook(report_data)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    filename = f"rekap_absensi_v2_{_report_period_slug(report_data['period']['label'])}.xlsx"
    return StreamingResponse(
        output,
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/rekap-absensi")
def get_rekap_absensi(
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=1900),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    term: int | None = Query(None, ge=1, le=4),
    db: Session = Depends(get_db),
):
    period = _resolve_rekap_absensi_period(month, year, date_from, date_to, term)
    return _collect_rekap_absensi_report_data(db, period)


@router.get("/rekap-absensi/export-excel")
def export_rekap_absensi_excel(
    month: int | None = Query(None, ge=1, le=12),
    year: int | None = Query(None, ge=1900),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    term: int | None = Query(None, ge=1, le=4),
    db: Session = Depends(get_db),
):
    period = _resolve_rekap_absensi_period(month, year, date_from, date_to, term)
    report_data = _collect_rekap_absensi_report_data(db, period)
    workbook = _build_rekap_absensi_workbook(report_data)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    filename = f"rekap_absensi_{_report_period_slug(report_data['period']['label'])}.xlsx"
    return StreamingResponse(
        output,
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/heb")
def get_heb_visibility(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=1900),
    db: Session = Depends(get_db),
):
    filters = month_year_filters(db, Attendance.date, month, year)
    student_rows = (
        db.query(
            Student.id,
            Student.jenjang,
            func.count(Attendance.id).label("present_days"),
        )
        .join(Attendance, Student.id == Attendance.student_id)
        .filter(Attendance.check_in.isnot(None), Student.jenjang.isnot(None), *filters)
        .group_by(Student.id, Student.jenjang)
        .all()
    )

    counts_by_jenjang: dict[str, list[int]] = defaultdict(list)
    for row in student_rows:
        counts_by_jenjang[row.jenjang].append(int(row.present_days))

    jenjang_rows = db.query(Student.jenjang).filter(Student.jenjang.isnot(None)).distinct().all()
    jenjang_list = sorted({row[0] for row in jenjang_rows if row[0]})
    student_count_rows = (
        db.query(Student.jenjang, func.count(Student.id).label("student_count"))
        .filter(Student.jenjang.isnot(None))
        .group_by(Student.jenjang)
        .all()
    )
    student_counts = {row.jenjang: int(row.student_count) for row in student_count_rows}

    heb_items = []
    for jenjang in jenjang_list:
        heb_result = calculate_heb(db, jenjang, month, year)
        auto_result = calculate_auto_heb(db, jenjang, month, year)
        override_row = get_heb_override(db, jenjang, month, year)

        heb_items.append(
            {
                "jenjang": jenjang,
                "heb": heb_result["heb"],
                "source": heb_result["source"],
                "note": heb_result["note"],
                "derived_from": heb_result["derived_from"],
                "median": heb_result["median"],
                "student_count": student_counts.get(jenjang, len(counts_by_jenjang.get(jenjang, []))),
                "auto_heb": auto_result["heb"],
                "auto_derived_from": auto_result["derived_from"],
                "auto_median": auto_result["median"],
                "override_heb": int(override_row.heb_value) if override_row is not None else None,
                "override_note": override_row.note if override_row is not None else None,
                "override_set_by": override_row.set_by if override_row is not None else None,
                "override_set_at": override_row.set_at.isoformat() if override_row is not None and override_row.set_at else None,
            }
        )

    return {
        "month": f"{year}-{month:02d}",
        "heb_by_jenjang": heb_items,
    }



@router.get("/monthly")
def get_monthly_late_trends(db: Session = Depends(get_db)):
    start = time.perf_counter()
    month_bucket = _month_bucket_expression(db).label("month")
    rows = (
        db.query(
            month_bucket,
            func.count().label("late_count"),
        )
        .filter(Attendance.status == "late")
        .group_by(month_bucket)
        .order_by(month_bucket)
        .all()
    )
    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/monthly: {elapsed:.3f}s")
    return _serialize_rows(rows, ["month", "late_count"])


@router.get("/monthly-by-class")
def get_monthly_late_by_class(db: Session = Depends(get_db)):
    start = time.perf_counter()
    month_bucket = _month_bucket_expression(db).label("month")
    rows = (
        db.query(
            Student.class_name,
            month_bucket,
            func.count().label("late_count"),
        )
        .join(Student)
        .filter(Attendance.status == "late")
        .group_by(Student.class_name, month_bucket)
        .all()
    )
    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/monthly-by-class: {elapsed:.3f}s")
    return _serialize_rows(rows, ["class_name", "month", "late_count"])


@router.get("/frequent-offenders")
def get_frequent_offenders(db: Session = Depends(get_db)):
    start = time.perf_counter()
    month_trunc = _month_bucket_expression(db).label("month")

    rows = (
        db.query(
            Student.name,
            Student.class_name,
            month_trunc,
            func.count().label("late_count"),
        )
        .join(Attendance)
        .filter(Attendance.status == "late")
        .group_by(Student.id, Student.name, Student.class_name, month_trunc)
        .having(func.count() >= 3)
        .order_by(desc("late_count"))
        .limit(20)
        .all()
    )
    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/frequent-offenders: {elapsed:.3f}s")
    return _serialize_rows(rows, ["name", "class_name", "month", "late_count"])


@router.get("/class-leaderboard")
def get_class_leaderboard(db: Session = Depends(get_db)):
    start = time.perf_counter()
    rows = (
        db.query(
            Student.class_name,
            func.count(Attendance.id).label("total_records"),
            func.count(case((Attendance.status == "late", 1))).label("late_count"),
            (
                100.0
                * func.count(case((Attendance.status == "on-time", 1)))
                / func.count(Attendance.id)
            ).label("punctuality_score"),
        )
        .join(Student)
        .group_by(Student.class_name)
        .order_by(desc("punctuality_score"))
        .all()
    )
    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/class-leaderboard: {elapsed:.3f}s")
    return _serialize_rows(rows, ["class_name", "total_records", "late_count", "punctuality_score"])


@router.get("/summary")
def get_analytics_summary(db: Session = Depends(get_db)):
    start = time.perf_counter()
    effective_status = func.coalesce(AttendanceOverride.override_status, Attendance.status)

    total_late = (
        db.query(func.count(Attendance.id))
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .filter(effective_status == "late")
        .scalar() or 0
    )
    total_incomplete = (
        db.query(func.count(Attendance.id))
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .filter(effective_status == "incomplete")
        .scalar()
        or 0
    )
    total_offenders = (
        db.query(Student.id)
        .join(Attendance)
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .filter(effective_status == "late")
        .group_by(Student.id)
        .having(func.count(Attendance.id) >= 3)
        .count()
    )

    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/summary: {elapsed:.3f}s")

    return {
        "total_late": total_late,
        "total_incomplete": total_incomplete,
        "total_offenders": total_offenders,
    }


@router.get("/attendance-date-range")
def get_attendance_date_range(db: Session = Depends(get_db)):
    row = db.query(func.min(Attendance.date), func.max(Attendance.date)).one()
    earliest_date, latest_date = row
    return {
        "earliest_date": earliest_date.isoformat() if earliest_date else None,
        "latest_date": latest_date.isoformat() if latest_date else None,
    }


@router.get("/incomplete-summary")
def get_incomplete_summary(db: Session = Depends(get_db)):
    start = time.perf_counter()
    effective_status = func.coalesce(AttendanceOverride.override_status, Attendance.status)

    rows = (
        db.query(Attendance.student_id, Attendance.date)
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .filter(effective_status == "incomplete", Attendance.check_in.isnot(None))
        .all()
    )

    total_incomplete = len(rows)
    affected_students = len({row.student_id for row in rows})

    dates = [row.date for row in rows]
    earliest_date = min(dates).isoformat() if dates else None
    latest_date = max(dates).isoformat() if dates else None

    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/incomplete-summary: {elapsed:.3f}s")

    return {
        "total_incomplete": total_incomplete,
        "affected_students": affected_students,
        "earliest_date": earliest_date,
        "latest_date": latest_date,
    }



@router.get("/pending-categorization")
def get_pending_categorization(db: Session = Depends(get_db)):
    start = time.perf_counter()
    rows = (
        db.query(Student)
        .filter(or_(Student.class_name.is_(None), Student.class_name == "Unknown Class"))
        .all()
    )
    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/pending-categorization: {elapsed:.3f}s")
    return rows


@router.get("/attendance-report")
def get_attendance_report(
    start_date: date = Query(...),
    end_date: date = Query(...),
    jenjang: str | None = None,
    class_name: str | None = None,
    db: Session = Depends(get_db),
):
    start = time.perf_counter()

    query = (
        db.query(
            Student.id.label("student_id"),
            Student.name,
            Student.class_name,
            Student.jenjang,
            func.count(case((func.coalesce(AttendanceOverride.override_status, Attendance.status) == "on-time", 1))).label("present_count"),
            func.count(case((func.coalesce(AttendanceOverride.override_status, Attendance.status) == "late", 1))).label("late_count"),
            func.count(case((func.coalesce(AttendanceOverride.override_status, Attendance.status) == "absent", 1))).label("absent_count"),
            func.count(case((func.coalesce(AttendanceOverride.override_status, Attendance.status) == "incomplete", 1))).label("incomplete_count"),
            func.sum(
                case(
                    (
                        func.coalesce(AttendanceOverride.override_status, Attendance.status) == "late",
                        func.coalesce(Attendance.late_duration, 0),
                    ),
                    else_=0,
                )
            ).label("total_late_duration"),
            func.count(Attendance.id).label("total_days"),
        )
        .join(Attendance, Student.id == Attendance.student_id)
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .filter(Attendance.date >= start_date, Attendance.date <= end_date)
    )

    if jenjang and jenjang.strip().lower() != "all":
        query = query.filter(Student.jenjang == jenjang.strip())

    if class_name and class_name.strip().lower() != "all":
        if class_name.strip() == "unassigned":
            query = query.filter(Student.class_name.is_(None))
        else:
            query = query.filter(Student.class_name == class_name.strip())

    rows = (
        query.group_by(Student.id, Student.name, Student.class_name, Student.jenjang)
        .order_by(Student.name)
        .all()
    )
    absence_reason_map = _get_absence_reason_map_for_range(db, start_date, end_date)

    results = []
    for row in rows:
        total = row.total_days
        attended = row.present_count + row.late_count + row.incomplete_count
        percentage = round((attended / total * 100), 1) if total > 0 else 0.0
        class_absence = absence_reason_map.get(
            row.class_name,
            {"sakit": 0, "izin": 0, "alfa": 0, "total_absence_reasons": 0},
        )


        results.append(
            {
                "student_id": row.student_id,
                "name": row.name,
                "class_name": row.class_name,
                "jenjang": row.jenjang,
                "present_count": row.present_count,
                "late_count": row.late_count,
                "absent_count": row.absent_count,
                "incomplete_count": row.incomplete_count,
                "sakit": class_absence["sakit"],
                "izin": class_absence["izin"],
                "alfa": class_absence["alfa"],
                "total_late_time_str": _format_duration(row.total_late_duration),
                "total_days": total,
                "attendance_percentage": percentage,
            }
        )

    # Calculate average late time for metadata (Average = Total Seconds / Total Late Instances)
    all_late_minutes = sum(
        row.total_late_duration if isinstance(row.total_late_duration, (int, float)) else int(row.total_late_duration.total_seconds() // 60)
        for row in rows
        if row.total_late_duration
    )
    all_late_count = sum(row.late_count for row in rows)

    avg_late_minutes = all_late_minutes / all_late_count if all_late_count > 0 else 0

    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/attendance-report: {elapsed:.3f}s")

    month_pairs = _month_pairs_in_range(start_date, end_date)
    total_heb_for_period = 0
    if jenjang and jenjang.strip().lower() != "all":
        norm_target = jenjang.strip().upper()
        orig_row = db.execute(
            select(func.trim(Student.jenjang).label("original"))
            .filter(_report_jenjang_expression(db) == norm_target)
            .limit(1)
        ).mappings().first()

        target_j = orig_row["original"] if orig_row else jenjang.strip()

        for pair_year, pair_month in month_pairs:
            total_heb_for_period += int(calculate_heb(db, target_j, pair_month, pair_year)["heb"] or 0)
    else:
        jenjang_list = [r[0] for r in db.query(Student.jenjang).filter(Student.jenjang.isnot(None)).distinct().all()]
        if jenjang_list:
            sum_heb = 0
            for j in jenjang_list:
                for pair_year, pair_month in month_pairs:
                    sum_heb += int(calculate_heb(db, j, pair_month, pair_year)["heb"] or 0)
            total_heb_for_period = round(sum_heb / len(jenjang_list))

    # Wrap in a dict to include summary metadata
    return {
        "results": results,
        "summary": {
            "avg_late_time_str": _format_duration(avg_late_minutes),
            "heb_days": total_heb_for_period
        }
    }



@router.get("/late-by-class")
def get_late_by_class(db: Session = Depends(get_db)):
    start = time.perf_counter()
    rows = (
        db.query(
            Student.class_name,
            func.count(Attendance.id).label("late_count"),
        )
        .join(Attendance, Student.id == Attendance.student_id)
        .filter(Attendance.status == "late")
        .group_by(Student.class_name)
        .order_by(desc("late_count"))
        .all()
    )
    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/late-by-class: {elapsed:.3f}s")
    return _serialize_rows(rows, ["class_name", "late_count"])


@router.get("/late-by-jenjang")
def get_late_by_jenjang(db: Session = Depends(get_db)):
    start = time.perf_counter()
    rows = (
        db.query(
            Student.jenjang,
            func.count(Attendance.id).label("late_count"),
        )
        .join(Attendance, Student.id == Attendance.student_id)
        .filter(Attendance.status == "late")
        .group_by(Student.jenjang)
        .order_by(desc("late_count"))
        .all()
    )
    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/late-by-jenjang: {elapsed:.3f}s")
    return _serialize_rows(rows, ["jenjang", "late_count"])


@router.get("/late-by-student")
def get_late_by_student(db: Session = Depends(get_db)):
    start = time.perf_counter()
    rows = (
        db.query(
            Student.id,
            Student.name,
            Student.class_name,
            Student.jenjang,
            func.count(Attendance.id).label("late_count"),
        )
        .join(Attendance, Student.id == Attendance.student_id)
        .filter(Attendance.status == "late")
        .group_by(Student.id, Student.name, Student.class_name, Student.jenjang)
        .order_by(desc("late_count"))
        .all()
    )

    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/late-by-student: {elapsed:.3f}s")

    return [
        {
            "no_id": str(row.id),
            "nama": row.name,
            "class_name": row.class_name,
            "jenjang": row.jenjang,
            "late_count": int(row.late_count),
        }
        for row in rows
    ]



@router.get("/attendance-rate/students")
def get_attendance_rate_per_student(db: Session = Depends(get_db)):
    start = time.perf_counter()
    month_expr = month_bucket_string_expression(db, Attendance.date).label("month")

    rows = (
        db.query(
            Student.id,
            Student.name,
            Student.class_name,
            Student.jenjang,
            month_expr,
            func.count(Attendance.id).label("present_days"),
        )
        .join(Attendance, Student.id == Attendance.student_id)
        .filter(Attendance.check_in.isnot(None))
        .group_by(Student.id, Student.name, Student.class_name, Student.jenjang, month_expr)
        .all()
    )

    all_month_rows = (
        db.query(month_expr)
        .filter(Attendance.date.isnot(None))
        .distinct()
        .order_by(month_expr)
        .all()
    )
    all_months = [row[0] for row in all_month_rows]

    heb_cache: dict[tuple[str, int, int], int] = {}
    student_data = {}

    for row in rows:
        month_key = row.month
        year, month = _parse_month_key(month_key)
        jenjang = row.jenjang
        cache_key = (jenjang, year, month)

        if cache_key not in heb_cache:
            heb_cache[cache_key] = calculate_heb(db, jenjang, month, year)["heb"]

        if row.id not in student_data:
            student_data[row.id] = {
                "no_id": str(row.id),
                "nama": row.name,
                "class_name": row.class_name,
                "jenjang": jenjang,
                "monthly_map": {},
            }

        present_days = int(row.present_days)
        heb = heb_cache[cache_key]
        student_data[row.id]["monthly_map"][month_key] = {
            "month": month_key,
            "present_days": present_days,
            "heb": heb,
            "rate": _safe_rate(present_days, heb),
        }

    response = []
    for item in student_data.values():
        monthly = [item["monthly_map"][month] for month in sorted(item["monthly_map"].keys())]
        total_present = sum(month_item["present_days"] for month_item in monthly)

        total_heb = 0
        for month_key in all_months:
            year, month = _parse_month_key(month_key)
            cache_key = (item["jenjang"], year, month)
            if cache_key not in heb_cache:
                heb_cache[cache_key] = calculate_heb(db, item["jenjang"], month, year)["heb"]
            total_heb += heb_cache[cache_key]

        response.append(
            {
                "no_id": item["no_id"],
                "nama": item["nama"],
                "class_name": item["class_name"],
                "jenjang": item["jenjang"],
                "monthly": monthly,
                "total": {
                    "present_days": total_present,
                    "heb": total_heb,
                    "rate": _safe_rate(total_present, total_heb),
                },
            }
        )

    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/attendance-rate/students: {elapsed:.3f}s")
    return sorted(response, key=lambda entry: entry["nama"])


@router.get("/attendance-rate/jenjang")
def get_attendance_rate_per_jenjang(db: Session = Depends(get_db)):
    start = time.perf_counter()
    month_expr = month_bucket_string_expression(db, Attendance.date).label("month")

    student_rows = (
        db.query(Student.id, Student.jenjang)
        .filter(Student.jenjang.isnot(None))
        .all()
    )
    student_ids_by_jenjang = defaultdict(list)
    for row in student_rows:
        jenjang = row.jenjang
        student_ids_by_jenjang[jenjang].append(row.id)


    present_rows = (
        db.query(
            Student.id,
            Student.class_name,
            month_expr,
            func.count(Attendance.id).label("present_days"),
        )
        .join(Attendance, Student.id == Attendance.student_id)
        .filter(Attendance.check_in.isnot(None))
        .group_by(Student.id, Student.class_name, month_expr)
        .all()
    )

    monthly_present_map: dict[tuple[int, str], int] = defaultdict(int)
    for row in present_rows:
        monthly_present_map[(row.id, row.month)] = int(row.present_days)

    all_month_rows = (
        db.query(month_expr)
        .filter(Attendance.date.isnot(None))
        .distinct()
        .order_by(month_expr)
        .all()
    )
    all_months = [row[0] for row in all_month_rows]

    heb_cache: dict[tuple[str, int, int], int] = {}
    response = []

    for jenjang, student_ids in student_ids_by_jenjang.items():
        if not student_ids:
            continue

        monthly = []
        total_heb = 0
        total_present_values = []

        for month_key in all_months:
            year, month = _parse_month_key(month_key)
            cache_key = (jenjang, year, month)
            if cache_key not in heb_cache:
                heb_cache[cache_key] = calculate_heb(db, jenjang, month, year)["heb"]

            heb = heb_cache[cache_key]
            total_heb += heb
            present_values = [monthly_present_map[(student_id, month_key)] for student_id in student_ids]
            avg_present = sum(present_values) / len(student_ids)
            total_present_values.append(sum(present_values))

            monthly.append(
                {
                    "month": month_key,
                    "avg_present_days": round(avg_present, 3),
                    "heb": heb,
                    "rate": _safe_rate(avg_present, heb),
                }
            )

        total_avg_present = (
            round(sum(total_present_values) / len(student_ids), 3)
            if student_ids
            else 0.0
        )

        response.append(
            {
                "jenjang": jenjang,
                "monthly": monthly,
                "total": {
                    "avg_present_days": total_avg_present,
                    "heb": total_heb,
                    "rate": _safe_rate(total_avg_present, total_heb),
                },
            }
        )

    elapsed = time.perf_counter() - start
    print(f"[PERF] /analytics/attendance-rate/jenjang: {elapsed:.3f}s")
    return sorted(response, key=lambda item: item["jenjang"])


def _format_late_duration_label(minutes: int) -> str:
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours}:{mins:02d}"


@router.get("/filters")
def get_analytics_filters(
    academic_year_id: int | None = Query(None),
    jenjang_id: int | None = Query(None),
    db: Session = Depends(get_db)
):
    from models.academic_year import AcademicYear
    from models.jenjang import Jenjang
    from models.subject import Subject
    from models.student_enrollment import StudentEnrollment

    academic_years = db.query(AcademicYear).order_by(AcademicYear.start_date.asc()).all()
    jenjangs = db.query(Jenjang).order_by(Jenjang.name.asc()).all()

    classes_q = db.query(StudentEnrollment.class_name).filter(StudentEnrollment.class_name.isnot(None)).distinct()
    if academic_year_id:
        classes_q = classes_q.filter(StudentEnrollment.academic_year_id == academic_year_id)
    if jenjang_id:
        classes_q = classes_q.filter(StudentEnrollment.jenjang_id == jenjang_id)

    class_names = [r[0] for r in classes_q.order_by(StudentEnrollment.class_name.asc()).all() if r[0] and r[0].strip()]

    subjects_q = db.query(Subject).order_by(Subject.name.asc())
    if jenjang_id:
        subjects_q = subjects_q.filter(Subject.jenjang_id == jenjang_id)
    subjects = subjects_q.all()

    return {
        "academic_years": [{"id": ay.id, "label": ay.label, "is_default": ay.is_default} for ay in academic_years],
        "jenjangs": [{"id": j.id, "name": j.name} for j in jenjangs],
        "class_names": class_names,
        "subjects": [{"id": s.id, "name": s.name, "jenjang_id": s.jenjang_id} for s in subjects]
    }


@router.get("/management-summary")
def get_management_summary(
    academic_year_id: int = Query(...),
    jenjang_id: int | None = Query(None),
    class_name: str | None = Query(None),
    term: str | None = Query(None),
    subject_id: int | None = Query(None),
    db: Session = Depends(get_db)
):
    return build_management_summary(
        db=db,
        academic_year_id=academic_year_id,
        jenjang_id=jenjang_id,
        class_name=class_name,
        term=term,
        subject_id=subject_id,
    )

    from models.academic_year import AcademicYear
    from models.jenjang import Jenjang
    from models.subject import Subject
    from models.student_enrollment import StudentEnrollment
    from models.student_subject_grade import StudentSubjectGrade
    from models.assessment_component import AssessmentComponent

    ay = db.query(AcademicYear).filter(AcademicYear.id == academic_year_id).first()
    if not ay:
        raise HTTPException(status_code=404, detail="Academic year not found")

    start_date = ay.start_date
    end_date = ay.end_date

    warnings = []

    jenjang_name = None
    if jenjang_id is not None:
        jenjang = db.query(Jenjang).filter(Jenjang.id == jenjang_id).first()
        if not jenjang:
            raise HTTPException(status_code=404, detail="Jenjang not found")
        jenjang_name = jenjang.name

    if subject_id is not None:
        sub = db.query(Subject).filter(Subject.id == subject_id).first()
        if not sub:
            raise HTTPException(status_code=404, detail="Subject not found")

    if term:
        try:
            term_num = int(term.split("_")[1])
            if term_num == 1:
                t_start = date(ay.start_date.year, 7, 1)
                t_end = date(ay.start_date.year, 9, 30)
            elif term_num == 2:
                t_start = date(ay.start_date.year, 10, 1)
                t_end = date(ay.start_date.year, 12, 31)
            elif term_num == 3:
                t_start = date(ay.end_date.year, 1, 1)
                t_end = date(ay.end_date.year, 3, 31)
            elif term_num == 4:
                t_start = date(ay.end_date.year, 4, 1)
                t_end = date(ay.end_date.year, 6, 30)
            else:
                raise ValueError("Invalid term number")

            start_date = max(start_date, t_start)
            end_date = min(end_date, t_end)
        except Exception:
            warnings.append(f"Term format '{term}' is invalid. Calculating for the whole academic year.")
            start_date = ay.start_date
            end_date = ay.end_date
    else:
        warnings.append("Term date mapping is not configured; analytics are calculated across the selected academic year.")

    month_pairs = _month_pairs_in_range(start_date, end_date)

    effective_status = func.coalesce(AttendanceOverride.override_status, Attendance.status)
    q_attendance = (
        db.query(
            effective_status.label("status"),
            func.count(Attendance.id).label("count")
        )
        .join(Student, Student.id == Attendance.student_id)
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .filter(Attendance.date >= start_date, Attendance.date <= end_date)
    )
    if jenjang_name:
        q_attendance = q_attendance.filter(Student.jenjang == jenjang_name)
    if class_name:
        q_attendance = q_attendance.filter(Student.class_name == class_name)

    attendance_counts = q_attendance.group_by(effective_status).all()

    hadir_count = 0
    for status, count in attendance_counts:
        if status in ("on-time", "late"):
            hadir_count += count

    q_absence = (
        db.query(
            func.sum(AbsenceReason.sakit).label("sakit"),
            func.sum(AbsenceReason.izin).label("izin"),
            func.sum(AbsenceReason.alfa).label("alfa")
        )
        .join(Student, Student.id == AbsenceReason.student_id)
    )
    if month_pairs:
        or_conds = [
            and_(AbsenceReason.year == y, AbsenceReason.month == m)
            for y, m in month_pairs
        ]
        q_absence = q_absence.filter(or_(*or_conds))
    else:
        q_absence = q_absence.filter(False)

    if jenjang_name:
        q_absence = q_absence.filter(Student.jenjang == jenjang_name)
    if class_name:
        q_absence = q_absence.filter(Student.class_name == class_name)

    absence_res = q_absence.first()
    sakit_count = int(absence_res.sakit or 0) if absence_res else 0
    izin_count = int(absence_res.izin or 0) if absence_res else 0
    alfa_count = int(absence_res.alfa or 0) if absence_res else 0

    total_records = hadir_count + sakit_count + izin_count + alfa_count
    if total_records > 0:
        status_percentages = {
            "hadir": round((hadir_count / total_records) * 100, 1),
            "sakit": round((sakit_count / total_records) * 100, 1),
            "izin": round((izin_count / total_records) * 100, 1),
            "alfa": round((alfa_count / total_records) * 100, 1),
        }
    else:
        status_percentages = {
            "hadir": 0.0,
            "sakit": 0.0,
            "izin": 0.0,
            "alfa": 0.0,
        }

    attendance_summary = {
        "total_records": total_records,
        "status_counts": {
            "hadir": hadir_count,
            "sakit": sakit_count,
            "izin": izin_count,
            "alfa": alfa_count
        },
        "status_percentages": status_percentages
    }

    q_lateness = (
        db.query(
            Student.class_name.label("class_name"),
            func.count(Attendance.id).label("late_days"),
            func.sum(Attendance.late_duration).label("late_minutes")
        )
        .join(Student, Student.id == Attendance.student_id)
        .outerjoin(AttendanceOverride, AttendanceOverride.attendance_id == Attendance.id)
        .filter(Attendance.date >= start_date, Attendance.date <= end_date)
        .filter(effective_status == "late")
    )
    if jenjang_name:
        q_lateness = q_lateness.filter(Student.jenjang == jenjang_name)
    if class_name:
        q_lateness = q_lateness.filter(Student.class_name == class_name)

    lateness_rows = q_lateness.group_by(Student.class_name).all()

    total_late_days = sum(int(r.late_days or 0) for r in lateness_rows)
    total_late_minutes = sum(int(r.late_minutes or 0) for r in lateness_rows)

    lateness_by_class = []
    for r in lateness_rows:
        c_name = r.class_name or "Unknown"
        ld = int(r.late_days or 0)
        lm = int(r.late_minutes or 0)

        lateness_by_class.append({
            "class_name": c_name,
            "late_days": ld,
            "late_minutes": lm,
            "late_duration_label": _format_late_duration_label(lm),
            "late_day_percentage": round((ld / total_late_days) * 100, 1) if total_late_days > 0 else 0.0,
            "late_duration_percentage": round((lm / total_late_minutes) * 100, 1) if total_late_minutes > 0 else 0.0
        })
    lateness_by_class.sort(key=lambda x: x["class_name"])

    q_students_by_class = (
        db.query(
            StudentEnrollment.class_name.label("class_name"),
            func.count(func.distinct(StudentEnrollment.student_id)).label("student_count")
        )
        .filter(StudentEnrollment.academic_year_id == academic_year_id)
    )
    if jenjang_id:
        q_students_by_class = q_students_by_class.filter(StudentEnrollment.jenjang_id == jenjang_id)
    if class_name:
        q_students_by_class = q_students_by_class.filter(StudentEnrollment.class_name == class_name)

    student_counts = {
        r.class_name: int(r.student_count or 0)
        for r in q_students_by_class.group_by(StudentEnrollment.class_name).all()
    }

    q_grades_by_class = (
        db.query(
            StudentEnrollment.class_name.label("class_name"),
            AssessmentComponent.assessment_type.label("assessment_type"),
            func.avg(StudentSubjectGrade.score).label("average_score")
        )
        .join(StudentSubjectGrade, StudentSubjectGrade.enrollment_id == StudentEnrollment.id)
        .join(AssessmentComponent, AssessmentComponent.id == StudentSubjectGrade.component_id)
        .filter(StudentEnrollment.academic_year_id == academic_year_id)
        .filter(StudentSubjectGrade.score.isnot(None))
    )
    if jenjang_id:
        q_grades_by_class = q_grades_by_class.filter(StudentEnrollment.jenjang_id == jenjang_id)
    if class_name:
        q_grades_by_class = q_grades_by_class.filter(StudentEnrollment.class_name == class_name)
    if subject_id:
        q_grades_by_class = q_grades_by_class.filter(StudentSubjectGrade.subject_id == subject_id)

    grade_class_rows = q_grades_by_class.group_by(StudentEnrollment.class_name, AssessmentComponent.assessment_type).all()

    grade_class_map = {}
    for r in grade_class_rows:
        c_name = r.class_name or "Unknown"
        if c_name not in grade_class_map:
            grade_class_map[c_name] = {"sumatif": None, "formatif": None}
        if r.assessment_type == "sumatif":
            grade_class_map[c_name]["sumatif"] = round(float(r.average_score), 1) if r.average_score is not None else None
        elif r.assessment_type == "formatif":
            grade_class_map[c_name]["formatif"] = round(float(r.average_score), 1) if r.average_score is not None else None

    grade_by_class = [
        {
            "class_name": c_name,
            "sumatif_average": vals["sumatif"],
            "formatif_average": vals["formatif"],
            "student_count": student_counts.get(c_name, 0)
        }
        for c_name, vals in grade_class_map.items()
    ]
    grade_by_class.sort(key=lambda x: x["class_name"])

    q_grades_by_subject = (
        db.query(
            Subject.id.label("subject_id"),
            Subject.name.label("subject_name"),
            AssessmentComponent.assessment_type.label("assessment_type"),
            func.avg(StudentSubjectGrade.score).label("average_score")
        )
        .join(StudentSubjectGrade, StudentSubjectGrade.subject_id == Subject.id)
        .join(StudentEnrollment, StudentEnrollment.id == StudentSubjectGrade.enrollment_id)
        .join(AssessmentComponent, AssessmentComponent.id == StudentSubjectGrade.component_id)
        .filter(StudentEnrollment.academic_year_id == academic_year_id)
        .filter(StudentSubjectGrade.score.isnot(None))
    )
    if jenjang_id:
        q_grades_by_subject = q_grades_by_subject.filter(StudentEnrollment.jenjang_id == jenjang_id)
    if class_name:
        q_grades_by_subject = q_grades_by_subject.filter(StudentEnrollment.class_name == class_name)
    if subject_id:
        q_grades_by_subject = q_grades_by_subject.filter(Subject.id == subject_id)

    grade_subject_rows = q_grades_by_subject.group_by(Subject.id, Subject.name, AssessmentComponent.assessment_type).all()

    grade_subject_map = {}
    for r in grade_subject_rows:
        s_id = r.subject_id
        s_name = r.subject_name
        if s_id not in grade_subject_map:
            grade_subject_map[s_id] = {"name": s_name, "sumatif": None, "formatif": None}
        if r.assessment_type == "sumatif":
            grade_subject_map[s_id]["sumatif"] = round(float(r.average_score), 1) if r.average_score is not None else None
        elif r.assessment_type == "formatif":
            grade_subject_map[s_id]["formatif"] = round(float(r.average_score), 1) if r.average_score is not None else None

    grade_by_subject = [
        {
            "subject_id": s_id,
            "subject_name": vals["name"],
            "sumatif_average": vals["sumatif"],
            "formatif_average": vals["formatif"]
        }
        for s_id, vals in grade_subject_map.items()
    ]
    grade_by_subject.sort(key=lambda x: x["subject_name"])

    q_grades_by_student = (
        db.query(
            Student.id.label("student_id"),
            Student.name.label("student_name"),
            StudentEnrollment.class_name.label("class_name"),
            AssessmentComponent.assessment_type.label("assessment_type"),
            func.avg(StudentSubjectGrade.score).label("average_score")
        )
        .join(StudentEnrollment, StudentEnrollment.student_id == Student.id)
        .join(StudentSubjectGrade, StudentSubjectGrade.enrollment_id == StudentEnrollment.id)
        .join(AssessmentComponent, AssessmentComponent.id == StudentSubjectGrade.component_id)
        .filter(StudentEnrollment.academic_year_id == academic_year_id)
        .filter(StudentSubjectGrade.score.isnot(None))
    )
    if jenjang_id:
        q_grades_by_student = q_grades_by_student.filter(StudentEnrollment.jenjang_id == jenjang_id)
    if class_name:
        q_grades_by_student = q_grades_by_student.filter(StudentEnrollment.class_name == class_name)
    if subject_id:
        q_grades_by_student = q_grades_by_student.filter(StudentSubjectGrade.subject_id == subject_id)

    grade_student_rows = q_grades_by_student.group_by(
        Student.id, Student.name, StudentEnrollment.class_name, AssessmentComponent.assessment_type
    ).all()

    grade_student_map = {}
    threshold_edelweiss = 85.0
    threshold_national = 75.0

    for r in grade_student_rows:
        s_id = r.student_id
        s_name = r.student_name
        c_name = r.class_name or "Unknown"
        key = (s_id, s_name, c_name)
        if key not in grade_student_map:
            grade_student_map[key] = {"sumatif": None, "formatif": None}
        if r.assessment_type == "sumatif":
            grade_student_map[key]["sumatif"] = round(float(r.average_score), 1) if r.average_score is not None else None
        elif r.assessment_type == "formatif":
            grade_student_map[key]["formatif"] = round(float(r.average_score), 1) if r.average_score is not None else None

    grade_by_student = []
    for key, vals in grade_student_map.items():
        s_id, s_name, c_name = key
        sum_avg = vals["sumatif"]
        for_avg = vals["formatif"]

        below = False
        if sum_avg is not None and sum_avg < threshold_edelweiss:
            below = True
        if for_avg is not None and for_avg < threshold_edelweiss:
            below = True

        grade_by_student.append({
            "student_id": s_id,
            "student_name": s_name,
            "class_name": c_name,
            "sumatif_average": sum_avg,
            "formatif_average": for_avg,
            "below_threshold": below
        })
    grade_by_student.sort(key=lambda x: x["student_name"])

    return {
        "filters": {
            "academic_year_id": academic_year_id,
            "jenjang_id": jenjang_id,
            "class_name": class_name,
            "term": term,
            "subject_id": subject_id
        },
        "attendance_summary": attendance_summary,
        "lateness_by_class": lateness_by_class,
        "grade_by_class": grade_by_class,
        "grade_by_subject": grade_by_subject,
        "grade_by_student": grade_by_student,
        "thresholds": {
            "kkm_edelweiss": threshold_edelweiss,
            "kkm_national": threshold_national
        },
        "warnings": warnings
    }


@router.get("/management-summary/export/pdf")
def export_management_summary_pdf(
    academic_year_id: int = Query(...),
    jenjang_id: int | None = Query(None),
    class_name: str | None = Query(None),
    term: str | None = Query(None),
    subject_id: int | None = Query(None),
    db: Session = Depends(get_db),
):
    summary = build_management_summary(
        db=db,
        academic_year_id=academic_year_id,
        jenjang_id=jenjang_id,
        class_name=class_name,
        term=term,
        subject_id=subject_id,
    )
    filename = build_management_report_filename(summary, "pdf")

    return StreamingResponse(
        BytesIO(build_management_summary_pdf(summary)),
        media_type=PDF_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/management-summary/export/excel")
def export_management_summary_excel(
    academic_year_id: int = Query(...),
    jenjang_id: int | None = Query(None),
    class_name: str | None = Query(None),
    term: str | None = Query(None),
    subject_id: int | None = Query(None),
    mode: str | None = Query("summary"),
    db: Session = Depends(get_db),
):
    summary = build_management_summary(
        db=db,
        academic_year_id=academic_year_id,
        jenjang_id=jenjang_id,
        class_name=class_name,
        term=term,
        subject_id=subject_id,
    )
    filename = build_management_report_filename(summary, "xlsx")

    return StreamingResponse(
        BytesIO(build_management_summary_excel(summary, {"mode": mode})),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
