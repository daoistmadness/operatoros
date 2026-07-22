from datetime import date
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from core.database import Base
from models.academic_master import AcademicClass, AcademicGrade, AcademicProgram
from models.academic_year import AcademicYear
from models.attendance import Attendance
from models.jenjang import Jenjang
from models.student import Student
from models.student_enrollment import StudentEnrollment
from models.student_enrollment import StudentEnrollmentLifecycleAudit
from models.student_master import (
    StudentDeviceIdentity, StudentEnrollmentClassHistory, StudentImportBatch,
    StudentMaster, StudentMasterChangeHistory,
)
from models.student_import_session import StudentImportAppliedAction, StudentImportSession
from schemas.student_management import (
    DeviceReassignRequest, DeviceReplaceRequest, DeviceRetireRequest,
    EnrollmentLifecycleRequest, EnrollmentTransferRequest,
    StudentCreateRequest, StudentProfilePatch,
)
from services.student_management import (
    DEVICE_REASSIGN_CONFIRMATION, DEVICE_REPLACE_CONFIRMATION, DEVICE_RETIRE_CONFIRMATION,
    ENROLLMENT_DELETE_CONFIRMATION, ENROLLMENT_TRANSFER_CONFIRMATION,
    add_or_replace_device, create_student, enrollment_deletion_status, hard_delete_enrollment,
    reassign_device, record_version, retire_device, serialize_student_detail,
    transfer_enrollment, transition_enrollment, update_student,
)
from services.spreadsheet_security import validate_xlsx_upload
from services.student_workbook import (
    UPDATE_CONFIRMATION, commit_update_preview, create_update_preview,
    export_student_workbook, result_workbook, serialize_update_batch,
)


@pytest.fixture
def student_db():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    @event.listens_for(engine, "connect")
    def enable_fks(connection, _record):
        connection.execute("PRAGMA foreign_keys=ON")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    db = Session()
    year = AcademicYear(label="2026/2027", start_date=date(2026, 7, 1), end_date=date(2027, 6, 30), status="active", is_default=True)
    jenjang = Jenjang(name="Synthetic Primary")
    db.add_all([year, jenjang]); db.flush()
    program = AcademicProgram(jenjang_id=jenjang.id, name="Synthetic Primary", active=True)
    db.add(program); db.flush()
    grade = AcademicGrade(jenjang_id=jenjang.id, program_id=program.id, name="Grade 1", sequence_number=1, active=True)
    db.add(grade); db.flush()
    first_class = AcademicClass(academic_year_id=year.id, grade_id=grade.id, class_name="Synthetic 1A", section_code="A", active=True)
    second_class = AcademicClass(academic_year_id=year.id, grade_id=grade.id, class_name="Synthetic 1B", section_code="B", active=True)
    db.add_all([first_class, second_class]); db.commit()
    yield db, year, first_class, second_class
    db.close(); Base.metadata.drop_all(engine)


def create_payload(name="Synthetic Student", device="000901", nisn="0000000901", nik="3201000000000901", class_id=None, year_id=None):
    data = {
        "identity": {"full_name": name, "nisn": nisn, "nik": nik, "student_status": "active"},
        "contact": {"address": "Synthetic Street", "student_email": "synthetic@example.test"},
        "guardians": [{"guardian_type": "guardian", "name": "Synthetic Guardian", "phone": "0800000000"}],
        "device_identity": {"device_identifier": device, "device_source": "attendance_machine", "effective_from": "2026-07-01", "reason": "Synthetic test assignment"},
    }
    if class_id and year_id:
        data["enrollment"] = {"academic_year_id": year_id, "academic_class_id": class_id, "effective_from": "2026-07-01"}
    return StudentCreateRequest.model_validate(data)


def test_manual_create_minimal_complete_duplicate_and_same_name(student_db):
    db, year, first_class, _second = student_db
    minimal = create_student(db, StudentCreateRequest.model_validate({"identity": {"full_name": "Minimal Synthetic"}}), "admin")
    assert minimal.id and minimal.student_status == "active"
    complete = create_student(db, create_payload(class_id=first_class.id, year_id=year.id), "admin")
    assert db.query(StudentDeviceIdentity).filter_by(student_master_id=complete.id, is_active=True).one().device_identifier == "000901"
    assert db.query(StudentEnrollment).filter_by(student_master_id=complete.id).one().academic_class_id == first_class.id
    assert db.query(StudentMasterChangeHistory).filter_by(student_master_id=complete.id, action="student_created").count() == 1
    with pytest.raises(HTTPException) as duplicate:
        create_student(db, create_payload(name="Other", device="000902"), "admin")
    assert duplicate.value.status_code == 409
    same_name = StudentCreateRequest.model_validate({"identity": {"full_name": "Minimal Synthetic"}, "duplicate_override_reason": "Confirmed separate student"})
    assert create_student(db, same_name, "admin").id != minimal.id


def test_create_rolls_back_on_device_conflict(student_db):
    db, _year, _first, _second = student_db
    create_student(db, create_payload(), "admin")
    before = db.query(StudentMaster).count()
    with pytest.raises(HTTPException):
        create_student(db, create_payload(name="Rollback Synthetic", device="000901", nisn="0000000902", nik="3201000000000902"), "admin")
    assert db.query(StudentMaster).count() == before


def test_profile_optimistic_update_uniqueness_and_audit(student_db):
    db, _year, _first, _second = student_db
    student = create_student(db, create_payload(), "admin")
    original_version = record_version(student)
    body = StudentProfilePatch.model_validate({
        "record_version": original_version,
        "identity": {**{field: getattr(student, field) for field in (
            "full_name", "preferred_name", "nipd", "nisn", "nik", "birth_place", "birth_date", "gender", "religion", "citizenship", "blood_type", "student_status", "admission_date", "admission_type", "previous_school"
        )}, "preferred_name": "Synth"},
        "contact": {"address": "Updated Synthetic Street"},
        "reason": "Correct synthetic profile",
    })
    updated = update_student(db, student, body, "admin")
    assert updated.preferred_name == "Synth"
    assert db.query(StudentMasterChangeHistory).filter_by(student_master_id=student.id, field_name="preferred_name").count() == 1
    with pytest.raises(HTTPException) as stale:
        update_student(db, student, body, "admin")
    assert stale.value.detail["code"] == "STALE_RECORD"


def test_device_history_transfer_and_attendance_preservation(student_db):
    db, year, first_class, second_class = student_db
    student = create_student(db, create_payload(class_id=first_class.id, year_id=year.id), "admin")
    old = db.query(StudentDeviceIdentity).filter_by(student_master_id=student.id, is_active=True).one()
    db.add(Attendance(student_id=old.legacy_student_id, date=date(2026, 7, 2), status="Hadir")); db.commit()
    attendance_before = db.query(Attendance).count()
    replacement = add_or_replace_device(db, student, DeviceReplaceRequest.model_validate({
        "device_identifier": "000902", "device_source": "attendance_machine",
        "effective_from": "2026-08-01", "reason": "Synthetic device replacement",
        "confirmation": DEVICE_REPLACE_CONFIRMATION,
    }), "admin")
    db.refresh(old)
    assert replacement.is_active and not old.is_active and db.query(Attendance).count() == attendance_before
    enrollment = db.query(StudentEnrollment).filter_by(student_master_id=student.id).one()
    transfer_enrollment(db, enrollment, EnrollmentTransferRequest.model_validate({
        "target_class_id": second_class.id, "effective_date": "2026-09-01",
        "reason": "Synthetic class transfer", "confirmation": ENROLLMENT_TRANSFER_CONFIRMATION,
    }), "admin")
    assert enrollment.academic_class_id == second_class.id
    assert db.query(StudentEnrollmentClassHistory).filter_by(enrollment_id=enrollment.id).count() == 2
    histories = db.query(StudentEnrollmentClassHistory).filter_by(enrollment_id=enrollment.id).order_by(StudentEnrollmentClassHistory.id).all()
    assert histories[0].effective_to == date(2026, 9, 1)
    retire_device(db, student, replacement, DeviceRetireRequest.model_validate({
        "effective_to": "2026-10-01", "reason": "Synthetic retirement",
        "confirmation": DEVICE_RETIRE_CONFIRMATION,
    }), "admin")
    assert not replacement.is_active


def test_enrollment_without_device_identity_and_device_can_be_added_later(student_db):
    db, year, first_class, _second = student_db
    body = StudentCreateRequest.model_validate({
        "identity": {"full_name": "Academic Only Student", "student_status": "active"},
        "enrollment": {"academic_year_id": year.id, "academic_class_id": first_class.id, "effective_from": "2026-07-01"},
    })
    student = create_student(db, body, "admin")
    enrollment = db.query(StudentEnrollment).filter_by(student_master_id=student.id).one()
    assert enrollment.student_id is None
    assert enrollment.lifecycle_state == "ACTIVE"
    mapping = add_or_replace_device(db, student, DeviceReplaceRequest.model_validate({
        "device_identifier": "000955", "device_source": "attendance_machine",
        "effective_from": "2026-08-01", "reason": "Machine link added after admission",
        "confirmation": DEVICE_REPLACE_CONFIRMATION,
    }), "admin")
    assert mapping.is_active
    db.refresh(enrollment)
    assert enrollment.student_id == mapping.legacy_student_id


def test_lifecycle_withdraw_reactivate_graduate_and_invalid_backward_transition(student_db):
    db, year, first_class, _second = student_db
    student = create_student(db, create_payload(class_id=first_class.id, year_id=year.id), "admin")
    enrollment = db.query(StudentEnrollment).filter_by(student_master_id=student.id).one()
    withdrawn = EnrollmentLifecycleRequest.model_validate({
        "effective_date": "2026-09-10", "reason": "Family relocation",
        "reason_code": "FAMILY_RELOCATION", "confirmation": "WITHDRAW_STUDENT_ENROLLMENT",
    })
    transition_enrollment(db, enrollment, "WITHDRAWN", withdrawn, "admin", "manual_withdraw")
    assert enrollment.lifecycle_state == "WITHDRAWN" and not enrollment.class_assigned
    reactivate = EnrollmentLifecycleRequest.model_validate({
        "effective_date": "2026-10-01", "reason": "Student returned",
        "reason_code": "RETURNED", "confirmation": "REACTIVATE_STUDENT_ENROLLMENT",
    })
    transition_enrollment(db, enrollment, "ACTIVE", reactivate, "admin", "manual_reactivate")
    graduate = EnrollmentLifecycleRequest.model_validate({
        "effective_date": "2027-06-30", "reason": "Completed terminal grade",
        "reason_code": "PROGRAM_COMPLETED", "confirmation": "GRADUATE_STUDENT_ENROLLMENT",
    })
    transition_enrollment(db, enrollment, "GRADUATED", graduate, "admin", "manual_graduate")
    with pytest.raises(HTTPException) as invalid:
        transition_enrollment(db, enrollment, "ACTIVE", reactivate, "admin", "manual_reactivate")
    assert invalid.value.detail["code"] == "INVALID_LIFECYCLE_TRANSITION"
    audits = db.query(StudentEnrollmentLifecycleAudit).filter_by(enrollment_id=enrollment.id).all()
    assert [(row.prior_state, row.new_state) for row in audits] == [
        ("ACTIVE", "WITHDRAWN"), ("WITHDRAWN", "ACTIVE"), ("ACTIVE", "GRADUATED")
    ]


def test_hard_delete_only_unused_draft(student_db):
    db, year, first_class, _second = student_db
    student = create_student(db, StudentCreateRequest.model_validate({"identity": {"full_name": "Draft Student"}}), "admin")
    draft = StudentEnrollment(
        student_master_id=student.id, student_id=None, academic_year_id=year.id,
        jenjang_id=db.get(AcademicGrade, first_class.grade_id).jenjang_id,
        academic_class_id=first_class.id, class_name=first_class.class_name,
        class_assigned=False, effective_from=year.start_date, lifecycle_state="DRAFT",
    )
    db.add(draft); db.commit()
    assert enrollment_deletion_status(db, draft)["can_hard_delete"] is True
    draft_id = draft.id
    assert hard_delete_enrollment(db, draft, ENROLLMENT_DELETE_CONFIRMATION)["deleted"] == 1
    assert db.get(StudentEnrollment, draft_id) is None

    operational_student = create_student(db, create_payload(name="Operational Student", device="000956", nisn="0000000956", nik="3201000000000956", class_id=first_class.id, year_id=year.id), "admin")
    operational = db.query(StudentEnrollment).filter_by(student_master_id=operational_student.id).one()
    with pytest.raises(HTTPException) as protected:
        hard_delete_enrollment(db, operational, ENROLLMENT_DELETE_CONFIRMATION)
    assert protected.value.status_code == 409
    assert "CLASS_HISTORY" in protected.value.detail["dependencies"]


def test_xlsx_export_preview_commit_round_trip_and_stale_protection(student_db):
    db, year, first_class, _second = student_db
    student = create_student(db, create_payload(class_id=first_class.id, year_id=year.id), "admin")
    exported = export_student_workbook(db)
    workbook = load_workbook(BytesIO(exported))
    assert workbook.sheetnames == ["Student Data", "Reference Values", "Instructions"]
    sheet = workbook["Student Data"]
    headers = [cell.value for cell in sheet[1]]
    assert sheet.freeze_panes == "A2" and sheet.auto_filter.ref
    for header in ("NIPD", "NISN", "NIK", "Attendance Device No. ID"):
        assert sheet.cell(2, headers.index(header) + 1).number_format == "@"
    sheet.cell(2, headers.index("Preferred Name") + 1).value = "Round Trip"
    sheet.cell(2, headers.index("Attendance Device No. ID") + 1).value = "000903"
    changed = BytesIO(); workbook.save(changed)
    batch = create_update_preview(db, changed.getvalue(), "synthetic-update.xlsx", "admin")
    preview = serialize_update_batch(db, batch)
    assert preview["summary"]["updates"] == 1
    assert db.get(StudentImportSession, batch.session_id).preview_checksum == preview["preview_checksum"]
    row = preview["rows"][0]
    assert row["differences"]["preferred_name"]["uploaded"] == "Round Trip"
    result = commit_update_preview(db, batch.id, [row["id"]], UPDATE_CONFIRMATION, preview["preview_checksum"], "admin")
    assert result["updated"] == 1
    session = db.get(StudentImportSession, batch.session_id)
    assert session.status == "COMMITTED" and session.provenance_status == "COMPLETE_ACTION_PROVENANCE"
    actions = db.query(StudentImportAppliedAction).filter_by(session_id=session.id).all()
    assert actions and len({action.operation_id for action in actions}) == len(actions)
    assert all("000903" not in str(action.after_state) for action in actions)
    replay = commit_update_preview(db, batch.id, [row["id"]], UPDATE_CONFIRMATION, preview["preview_checksum"], "admin")
    assert replay["idempotent_replay"] is True
    db.refresh(student)
    assert student.preferred_name == "Round Trip"
    assert db.query(StudentDeviceIdentity).filter_by(student_master_id=student.id, is_active=True).one().device_identifier == "000903"
    assert load_workbook(BytesIO(result_workbook(db, batch))).active.title == "Validation Summary"

    stale_book = load_workbook(BytesIO(exported)); stale_sheet = stale_book["Student Data"]
    stale_sheet.cell(2, headers.index("Preferred Name") + 1).value = "Stale Value"
    stale_bytes = BytesIO(); stale_book.save(stale_bytes)
    stale_batch = create_update_preview(db, stale_bytes.getvalue(), "stale.xlsx", "admin")
    assert serialize_update_batch(db, stale_batch)["rows"][0]["classification"] == "CONFLICT"


def test_guarded_device_reassignment_preserves_both_audit_sides(student_db):
    db, _year, _first, _second = student_db
    source = create_student(db, create_payload(), "admin")
    target = create_student(db, StudentCreateRequest.model_validate({"identity": {"full_name": "Reassignment Target"}}), "admin")
    moved = reassign_device(db, target, DeviceReassignRequest.model_validate({
        "device_identifier": "000901", "device_source": "attendance_machine",
        "effective_from": "2026-08-01", "reason": "Confirmed synthetic reassignment",
        "previous_student_master_id": source.id, "confirmation": DEVICE_REASSIGN_CONFIRMATION,
    }), "admin")
    assert moved.student_master_id == target.id and moved.is_active
    assert db.query(StudentDeviceIdentity).filter_by(student_master_id=source.id, is_active=True).count() == 0
    assert db.query(StudentMasterChangeHistory).filter_by(student_master_id=source.id, action="device_identity_reassigned_out").count() == 1
    assert db.query(StudentMasterChangeHistory).filter_by(student_master_id=target.id, action="device_identity_reassigned_in").count() == 1


def test_sensitive_profile_view_is_masked_and_identifier_change_requires_confirmation(student_db):
    db, _year, _first, _second = student_db
    student = create_student(db, create_payload(), "admin")
    masked = serialize_student_detail(db, student, include_sensitive=False)
    assert masked["identity"]["nik"] == "************0901"
    assert masked["identity"]["nisn"] == "******0901"
    assert masked["contact"]["address"] != "Synthetic Street"
    assert masked["guardians"][0]["phone"] != "0800000000"
    assert masked["device_identities"][0]["device_identifier"] != "000901"

    identity = {field: getattr(student, field) for field in (
        "full_name", "preferred_name", "nipd", "nisn", "nik", "birth_place",
        "birth_date", "gender", "religion", "citizenship", "blood_type",
        "student_status", "admission_date", "admission_type", "previous_school",
    )}
    identity["nik"] = "3201000000000999"
    body = StudentProfilePatch.model_validate({
        "record_version": record_version(student),
        "identity": identity,
        "reason": "Correct synthetic identifier",
    })
    with pytest.raises(HTTPException) as missing_confirmation:
        update_student(db, student, body, "admin")
    assert missing_confirmation.value.status_code == 400

    confirmed = body.model_copy(update={"sensitive_confirmation": "CHANGE_SENSITIVE_STUDENT_IDENTIFIERS"})
    update_student(db, student, confirmed, "admin")
    audit = db.query(StudentMasterChangeHistory).filter_by(student_master_id=student.id, field_name="nik").one()
    assert audit.old_value == "************0901"
    assert audit.new_value == "************0999"


def test_xlsx_security_rejects_malformed_and_formula_workbooks():
    with pytest.raises(HTTPException):
        validate_xlsx_upload(b"not-a-zip", "../../unsafe.xlsx")
    workbook = Workbook()
    workbook.active["A1"] = "=1+1"
    output = BytesIO(); workbook.save(output)
    with pytest.raises(HTTPException) as formula:
        validate_xlsx_upload(output.getvalue(), "formula.xlsx")
    assert formula.value.detail == "Formula cells are not accepted"
