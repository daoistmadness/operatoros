from __future__ import annotations

import importlib
import io
import sys
from datetime import date, datetime, time
from pathlib import Path

import pytest
from openpyxl import load_workbook

MODULE_PREFIXES = ("src", "api", "core", "models", "services")
SOURCE_ROOT = Path(__file__).resolve().parents[1] / "src"


def unload_app_modules() -> None:
    for name in list(sys.modules):
        if name == "src" or name.startswith(MODULE_PREFIXES):
            sys.modules.pop(name, None)


@pytest.fixture
def trends_db(monkeypatch, tmp_path):
    db_path = tmp_path / "phase18.db"
    monkeypatch.syspath_prepend(str(SOURCE_ROOT))
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    unload_app_modules()
    db_module = importlib.import_module("core.database")
    db_module.init_db()

    AcademicYear = importlib.import_module("models.academic_year").AcademicYear
    Jenjang = importlib.import_module("models.jenjang").Jenjang
    Student = importlib.import_module("models.student").Student
    StudentEnrollment = importlib.import_module("models.student_enrollment").StudentEnrollment
    Attendance = importlib.import_module("models.attendance").Attendance
    AbsenceReason = importlib.import_module("models.absence_reason").AbsenceReason
    Subject = importlib.import_module("models.subject").Subject
    AssessmentComponent = importlib.import_module("models.assessment_component").AssessmentComponent
    StudentSubjectGrade = importlib.import_module("models.student_subject_grade").StudentSubjectGrade
    KkmThreshold = importlib.import_module("models.academic_config").KkmThreshold
    AcademicTermConfig = importlib.import_module("models.academic_config").AcademicTermConfig
    AcademicIntervention = importlib.import_module("models.academic_intervention").AcademicIntervention

    db = db_module.SessionLocal()
    db.query(AcademicYear).update({AcademicYear.is_default: False})
    ay_prev = db.query(AcademicYear).filter(AcademicYear.label == "2024/2025").first()
    if ay_prev is None:
        ay_prev = AcademicYear(label="2024/2025", start_date=date(2024, 7, 1), end_date=date(2025, 6, 30), status="closed", is_default=False)
        db.add(ay_prev)
    else:
        ay_prev.start_date = date(2024, 7, 1)
        ay_prev.end_date = date(2025, 6, 30)
        ay_prev.status = "closed"
        ay_prev.is_default = False
    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    if ay is None:
        ay = AcademicYear(label="2025/2026", start_date=date(2025, 7, 1), end_date=date(2026, 6, 30), status="active", is_default=True)
        db.add(ay)
    else:
        ay.start_date = date(2025, 7, 1)
        ay.end_date = date(2026, 6, 30)
        ay.status = "active"
        ay.is_default = True
    jen = db.query(Jenjang).filter(Jenjang.name == "Primary").first()
    if jen is None:
        jen = Jenjang(name="Primary")
        db.add(jen)
    db.flush()
    for year in (ay_prev, ay):
        db.add_all([
            AcademicTermConfig(academic_year_id=year.id, term_number=1, label="Term 1", start_date=date(year.start_date.year, 7, 1), end_date=date(year.start_date.year, 9, 30)),
            AcademicTermConfig(academic_year_id=year.id, term_number=2, label="Term 2", start_date=date(year.start_date.year, 10, 1), end_date=date(year.start_date.year, 12, 31)),
            AcademicTermConfig(academic_year_id=year.id, term_number=3, label="Term 3", start_date=date(year.end_date.year, 1, 1), end_date=date(year.end_date.year, 3, 31)),
            AcademicTermConfig(academic_year_id=year.id, term_number=4, label="Term 4", start_date=date(year.end_date.year, 4, 1), end_date=date(year.end_date.year, 6, 30)),
        ])
    math = Subject(name="Math", jenjang_id=jen.id, supports_sumatif=True, supports_formatif=True)
    db.add(math)
    db.flush()
    db.add_all([
        KkmThreshold(academic_year_id=ay.id, jenjang_id=jen.id, subject_id=math.id, assessment_type="sumatif", threshold=80.0),
        KkmThreshold(academic_year_id=ay.id, jenjang_id=jen.id, subject_id=math.id, assessment_type="formatif", threshold=78.0),
    ])
    sumatif = AssessmentComponent(name="Math Sumatif", assessment_type="sumatif", subject_id=math.id)
    formatif = AssessmentComponent(name="Math Formatif", assessment_type="formatif", subject_id=math.id)
    db.add_all([sumatif, formatif])
    db.flush()
    students = [Student(name=f"Student {idx}", jenjang="Primary", class_name="P3") for idx in range(1, 4)]
    db.add_all(students)
    db.flush()
    enrollments = []
    for year in (ay_prev, ay):
        for student in students:
            enrollment = StudentEnrollment(student_id=student.id, academic_year_id=year.id, jenjang_id=jen.id, class_name="P3", class_assigned=True)
            db.add(enrollment)
            enrollments.append((year.id, student.id, enrollment))
    db.flush()

    attendance_specs = [
        (ay_prev, date(2024, 7, 10), 3, 0, 0, 0, 1, 15),
        (ay, date(2025, 7, 10), 3, 0, 0, 0, 1, 20),
        (ay, date(2025, 10, 10), 2, 1, 0, 0, 2, 45),
        (ay, date(2026, 1, 10), 2, 0, 1, 0, 2, 60),
    ]
    for year, day, hadir, sakit, izin, alfa, late_days, late_minutes in attendance_specs:
        for idx, student in enumerate(students):
            is_late = idx < late_days
            db.add(Attendance(
                student_id=student.id,
                date=day,
                check_in=time(7, 20) if is_late else time(7, 0),
                check_out=time(14, 0),
                status="late" if is_late else "on-time",
                late_duration=late_minutes // late_days if is_late and late_days else 0,
            ))
        db.add(AbsenceReason(student_id=students[0].id, class_name="P3", month=day.month, year=day.year, sakit=sakit, izin=izin, alfa=alfa, entered_by="test"))

    score_by_year = {
        ay_prev.id: [(75.0, 74.0), (82.0, 80.0), (86.0, 84.0)],
        ay.id: [(76.0, 75.0), (84.0, 82.0), (88.0, 85.0)],
    }
    for year_id, student_id, enrollment in enrollments:
        for sum_score, for_score in [score_by_year[year_id][student_id - students[0].id]]:
            db.add(StudentSubjectGrade(enrollment_id=enrollment.id, subject_id=math.id, component_id=sumatif.id, score=sum_score))
            db.add(StudentSubjectGrade(enrollment_id=enrollment.id, subject_id=math.id, component_id=formatif.id, score=for_score))
    db.add_all([
        AcademicIntervention(
            student_id=students[0].id, academic_year_id=ay.id, jenjang_id=jen.id, subject_id=math.id,
            class_name="P3", student_name=students[0].name, subject_name="Math", assessment_type="sumatif",
            term="term_1", effective_threshold=80.0, threshold_source="subject-specific",
            status="resolved", priority="high", created_at=datetime(2025, 7, 12), resolved_at=datetime(2025, 7, 20),
        ),
        AcademicIntervention(
            student_id=students[0].id, academic_year_id=ay.id, jenjang_id=jen.id, subject_id=math.id,
            class_name="P3", student_name=students[0].name, subject_name="Math", assessment_type="formatif",
            term="term_2", effective_threshold=78.0, threshold_source="subject-specific",
            status="open", priority="urgent", follow_up_date=date(2025, 11, 1),
        ),
    ])
    db.commit()
    yield {"db": db, "ay_id": ay.id, "prev_ay_id": ay_prev.id, "jen_id": jen.id, "subject_id": math.id}
    db.close()


def test_forecast_methods_and_sufficiency(trends_db):
    forecast = importlib.import_module("services.analytics_forecast")
    insufficient = forecast.forecast_metric("attendance_percentage", [95.0], method="linear_trend")
    assert insufficient["forecast_value"] is None
    assert insufficient["data_sufficiency"] == "insufficient"
    assert forecast.forecast_metric("late_days", [1, 2, 3], method="moving_average")["forecast_value"] == 2.0
    assert forecast.forecast_metric("late_days", [1, 2, 3], method="weighted_moving_average")["forecast_value"] == pytest.approx(2.3, abs=0.1)
    assert forecast.forecast_metric("late_days", [1, 2, 3], method="linear_trend")["forecast_value"] == 4.0


def test_historical_trend_service_series_and_insights(trends_db):
    service = importlib.import_module("services.analytics_trends")
    data = service.build_historical_trends(
        trends_db["db"],
        academic_year_id=trends_db["ay_id"],
        from_academic_year_id=trends_db["prev_ay_id"],
        jenjang_id=trends_db["jen_id"],
        subject_id=trends_db["subject_id"],
        include_forecast=True,
        forecast_method="linear_trend",
    )
    assert data["trend_series"]["attendance"]["by_term"]
    assert data["trend_series"]["lateness"]["by_term"]
    assert data["trend_series"]["grades"]["by_term"]
    assert data["trend_series"]["interventions"]["by_term"]
    assert data["forecast_series"][0]["method"] in {"linear_trend", "none"}
    assert all("kkm_threshold_source" not in str(row) for row in data["effective_kkm_metadata"])
    assert data["executive_insights"]


def test_historical_trends_api_and_export_parity(trends_db):
    main = importlib.import_module("main")
    route_paths = {getattr(route, "path", "") for route in main.app.routes}
    assert "/api/analytics/historical-trends" in route_paths
    service = importlib.import_module("services.analytics_trends")
    payload = service.build_historical_trends(
        trends_db["db"],
        academic_year_id=trends_db["ay_id"],
        jenjang_id=trends_db["jen_id"],
        subject_id=trends_db["subject_id"],
        include_forecast=True,
    )
    assert payload["forecast_series"]
    assert payload["trend_series"]["grades"]["effective_kkm_by_term"]

    summary_builder = importlib.import_module("services.management_analytics").build_management_summary
    exports = importlib.import_module("services.management_report_export")
    summary = summary_builder(
        trends_db["db"],
        academic_year_id=trends_db["ay_id"],
        jenjang_id=trends_db["jen_id"],
        subject_id=trends_db["subject_id"],
    )
    summary["historical_trends"] = payload
    pdf_bytes = exports.build_management_summary_pdf(summary)
    assert b"Historical Trends summary page" in pdf_bytes
    assert b"Forecast methodology notes" in pdf_bytes

    excel_bytes = exports.build_management_summary_excel(summary, {"mode": "editable"})
    workbook = load_workbook(io.BytesIO(excel_bytes), read_only=False)
    for sheet_name in [
        "Trend_Attendance_Data",
        "Trend_Lateness_Data",
        "Trend_Grades_Data",
        "Trend_Interventions_Data",
        "Forecast_Data",
        "Trend_Insights",
    ]:
        assert sheet_name in workbook.sheetnames
    assert workbook["Trend_Grades_Data"]["F1"].value == "Threshold Source"
