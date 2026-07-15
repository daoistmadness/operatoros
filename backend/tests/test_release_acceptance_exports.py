"""Release-acceptance contracts for executive report artifacts.

Artifacts stay in memory so the disposable database and generated exports are
removed automatically when pytest tears down the fixture.
"""

from datetime import date, time
from io import BytesIO
import sys
from pathlib import Path

from openpyxl import load_workbook

SOURCE_ROOT = Path(__file__).resolve().parents[1] / "src"
if str(SOURCE_ROOT) not in sys.path:
    sys.path.insert(0, str(SOURCE_ROOT))

from models.attendance import Attendance
from test_report_exports import PDF_MIME, XLSX_MIME, export
from test_reports import report_context


def _add_second_month(context):
    primary = context["students"]["Primary A"][0]
    secondary = context["students"]["Secondary"][0]
    context["db"].add_all(
        [
            Attendance(
                student_id=primary.id,
                date=date(2024, 3, 1),
                check_in=time(7, 0),
                check_out=time(14, 0),
                late_duration=0,
                status="on-time",
            ),
            Attendance(
                student_id=secondary.id,
                date=date(2024, 3, 1),
                check_in=time(7, 20),
                check_out=time(14, 0),
                late_duration=20,
                status="late",
            ),
        ]
    )
    context["db"].commit()


def test_monthly_and_annual_pdf_release_acceptance(report_context):
    _add_second_month(report_context)

    monthly = export(report_context, "monthly", "pdf")
    annual = export(report_context, "annual", "pdf")

    assert monthly.status_code == annual.status_code == 200
    assert monthly.headers["content-type"] == annual.headers["content-type"] == PDF_MIME
    assert 'filename="executive-report_monthly_combined_2024-02.pdf"' in monthly.headers["content-disposition"]
    assert 'filename="executive-report_annual_combined_2023-2024.pdf"' in annual.headers["content-disposition"]
    for response in (monthly, annual):
        assert response.content.startswith(b"%PDF-")
        assert response.content.rstrip().endswith(b"%%EOF")
        assert b"Executive Report" in response.content

    annual_json = report_context["client"].get(
        "/api/reports/annual",
        params={"academic_year_id": report_context["year"].id, "scope": "combined"},
    ).json()
    trend_months = {row["month"] for row in annual_json["trends"]}
    assert {"2024-02", "2024-03"}.issubset(trend_months)


def test_monthly_and_annual_xlsx_release_acceptance(report_context):
    _add_second_month(report_context)

    monthly = export(report_context, "monthly", "xlsx")
    annual = export(report_context, "annual", "xlsx")

    assert monthly.status_code == annual.status_code == 200
    assert monthly.headers["content-type"] == annual.headers["content-type"] == XLSX_MIME
    assert 'filename="executive-report_monthly_combined_2024-02.xlsx"' in monthly.headers["content-disposition"]
    assert 'filename="executive-report_annual_combined_2023-2024.xlsx"' in annual.headers["content-disposition"]

    monthly_book = load_workbook(BytesIO(monthly.content), data_only=False)
    annual_book = load_workbook(BytesIO(annual.content), data_only=False)
    assert monthly_book.sheetnames == [
        "Executive Summary",
        "Attendance",
        "Student Distribution",
        "Academic Summary",
        "Data Quality",
    ]
    assert annual_book.sheetnames == [
        "Executive Summary",
        "Attendance",
        "Student Distribution",
        "Academic Summary",
        "Annual Trends",
        "Data Quality",
    ]
    trend_months = {
        row[0]
        for row in annual_book["Annual Trends"].iter_rows(min_row=3, values_only=True)
    }
    assert {"2024-02", "2024-03"}.issubset(trend_months)
    assert monthly_book["Attendance"].freeze_panes == "A3"
    assert annual_book["Annual Trends"].freeze_panes == "A3"
