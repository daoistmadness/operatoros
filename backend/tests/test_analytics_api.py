import importlib
import sqlite3
import sys
from datetime import date, time
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

MODULE_PREFIXES = ("src", "api", "core", "models", "services")
SOURCE_ROOT = Path(__file__).resolve().parents[1] / "src"


def unload_app_modules() -> None:
    for name in list(sys.modules):
        if name == "src" or name.startswith(MODULE_PREFIXES):
            sys.modules.pop(name, None)


def prepare_source_imports(monkeypatch) -> None:
    monkeypatch.syspath_prepend(str(SOURCE_ROOT))


def create_legacy_grade_table(db_path: Path) -> None:
    connection = sqlite3.connect(db_path)
    try:
        connection.execute(
            """
            CREATE TABLE student_term_grades (
                id INTEGER PRIMARY KEY,
                student_id INTEGER NOT NULL,
                academic_year VARCHAR NOT NULL,
                term_1 FLOAT,
                term_2 FLOAT,
                term_3 FLOAT,
                term_4 FLOAT
            )
            """
        )
        connection.commit()
    finally:
        connection.close()


@pytest.fixture
def app_context(monkeypatch, tmp_path):
    db_path = tmp_path / "attendance-test.db"
    create_legacy_grade_table(db_path)
    prepare_source_imports(monkeypatch)
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    unload_app_modules()

    main_module = importlib.import_module("src.main")
    db_module = importlib.import_module("core.database")
    student_module = importlib.import_module("models.student")
    attendance_module = importlib.import_module("models.attendance")
    absence_module = importlib.import_module("models.absence_reason")
    academic_year_module = importlib.import_module("models.academic_year")
    jenjang_module = importlib.import_module("models.jenjang")
    subject_module = importlib.import_module("models.subject")
    component_module = importlib.import_module("models.assessment_component")
    enrollment_module = importlib.import_module("models.student_enrollment")
    grade_module = importlib.import_module("models.student_subject_grade")
    upload_log_module = importlib.import_module("models.upload_log")
    academic_config_module = importlib.import_module("models.academic_config")
    intervention_module = importlib.import_module("models.academic_intervention")

    return {
        "app": main_module.app,
        "db_module": db_module,
        "Student": student_module.Student,
        "Attendance": attendance_module.Attendance,
        "AbsenceReason": absence_module.AbsenceReason,
        "AcademicYear": academic_year_module.AcademicYear,
        "Jenjang": jenjang_module.Jenjang,
        "Subject": subject_module.Subject,
        "AssessmentComponent": component_module.AssessmentComponent,
        "StudentEnrollment": enrollment_module.StudentEnrollment,
        "StudentSubjectGrade": grade_module.StudentSubjectGrade,
        "UploadLog": upload_log_module.UploadLog,
        "KkmThreshold": academic_config_module.KkmThreshold,
        "AcademicTermConfig": academic_config_module.AcademicTermConfig,
        "AcademicIntervention": intervention_module.AcademicIntervention,
    }


def test_management_summary_empty_state(app_context):
    client = TestClient(app_context["app"])
    db_module = app_context["db_module"]
    AcademicYear = app_context["AcademicYear"]

    db = db_module.SessionLocal()
    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    assert ay is not None
    ay_id = ay.id
    db.close()

    # Query endpoint
    response = client.get(f"/api/analytics/management-summary?academic_year_id={ay_id}")
    assert response.status_code == 200
    data = response.json()

    assert data["filters"]["academic_year_id"] == ay_id
    assert data["attendance_summary"]["total_records"] == 0
    assert data["attendance_summary"]["status_counts"]["hadir"] == 0
    assert data["lateness_by_class"] == []
    assert data["grade_by_class"] == []
    assert data["grade_by_subject"] == []
    assert data["grade_by_student"] == []
    assert len(data["warnings"]) > 0


def test_management_summary_invalid_id(app_context):
    client = TestClient(app_context["app"])
    response = client.get("/api/analytics/management-summary?academic_year_id=99999")
    assert response.status_code == 404
    assert "Academic year not found" in response.json()["detail"]


def seed_management_export_dataset(app_context):
    db_module = app_context["db_module"]
    Student = app_context["Student"]
    Attendance = app_context["Attendance"]
    AbsenceReason = app_context["AbsenceReason"]
    AcademicYear = app_context["AcademicYear"]
    Jenjang = app_context["Jenjang"]
    Subject = app_context["Subject"]
    AssessmentComponent = app_context["AssessmentComponent"]
    StudentEnrollment = app_context["StudentEnrollment"]
    StudentSubjectGrade = app_context["StudentSubjectGrade"]

    db = db_module.SessionLocal()
    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    jen = db.query(Jenjang).filter(Jenjang.name == "Primary").first()
    ay_id = ay.id
    jen_id = jen.id

    student = Student(name="Export Student", jenjang="Primary", class_name="P3A")
    db.add(student)
    db.flush()

    enrollment = StudentEnrollment(
        student_id=student.id,
        academic_year_id=ay_id,
        jenjang_id=jen_id,
        class_name="P3A",
        class_assigned=True,
    )
    db.add(enrollment)
    db.flush()

    db.add_all(
        [
            Attendance(
                student_id=student.id,
                date=date(2025, 10, 7),
                check_in=time(7, 45),
                check_out=time(14, 0),
                status="late",
                late_duration=30,
            ),
            Attendance(
                student_id=student.id,
                date=date(2026, 2, 7),
                check_in=time(7, 30),
                check_out=time(14, 0),
                status="late",
                late_duration=20,
            ),
            AbsenceReason(
                student_id=student.id,
                class_name="P3A",
                month=10,
                year=2025,
                sakit=1,
                izin=0,
                alfa=0,
                entered_by="operator",
            ),
        ]
    )

    subject = Subject(name="Export Math", jenjang_id=jen_id, supports_sumatif=True, supports_formatif=True)
    db.add(subject)
    db.flush()

    comp_sum = AssessmentComponent(name="Export Sumatif", assessment_type="sumatif", subject_id=subject.id)
    comp_for = AssessmentComponent(name="Export Formatif", assessment_type="formatif", subject_id=subject.id)
    db.add_all([comp_sum, comp_for])
    db.flush()

    db.add_all(
        [
            StudentSubjectGrade(
                enrollment_id=enrollment.id,
                subject_id=subject.id,
                component_id=comp_sum.id,
                score=74.0,
            ),
            StudentSubjectGrade(
                enrollment_id=enrollment.id,
                subject_id=subject.id,
                component_id=comp_for.id,
                score=None,
            ),
        ]
    )
    db.commit()
    subject_id = subject.id
    db.close()
    return ay_id, jen_id, subject_id


def test_management_summary_calculations(app_context):
    client = TestClient(app_context["app"])
    db_module = app_context["db_module"]
    Student = app_context["Student"]
    Attendance = app_context["Attendance"]
    AbsenceReason = app_context["AbsenceReason"]
    AcademicYear = app_context["AcademicYear"]
    Jenjang = app_context["Jenjang"]
    Subject = app_context["Subject"]
    AssessmentComponent = app_context["AssessmentComponent"]
    StudentEnrollment = app_context["StudentEnrollment"]
    StudentSubjectGrade = app_context["StudentSubjectGrade"]

    db = db_module.SessionLocal()

    # 1. Fetch automatically seeded academic year and jenjang
    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    assert ay is not None
    jen = db.query(Jenjang).filter(Jenjang.name == "Primary").first()
    assert jen is not None
    ay_id = ay.id
    jen_id = jen.id

    # 2. Seed Students
    s1 = Student(name="Student One", jenjang="Primary", class_name="P1A")
    s2 = Student(name="Student Two", jenjang="Primary", class_name="P1A")
    db.add_all([s1, s2])
    db.flush()

    # 3. Seed Enrollments
    e1 = StudentEnrollment(student_id=s1.id, academic_year_id=ay_id, jenjang_id=jen_id, class_name="P1A", class_assigned=True)
    e2 = StudentEnrollment(student_id=s2.id, academic_year_id=ay_id, jenjang_id=jen_id, class_name="P1A", class_assigned=True)
    db.add_all([e1, e2])
    db.flush()

    # 4. Seed Attendance (Daily)
    att1 = Attendance(student_id=s1.id, date=date(2025, 10, 5), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0)
    att2 = Attendance(student_id=s2.id, date=date(2025, 10, 5), check_in=time(7, 45), check_out=time(14, 0), status="late", late_duration=45)
    db.add_all([att1, att2])

    # 5. Seed Absence Reasons (Sakit/Izin/Alfa monthly)
    abs1 = AbsenceReason(student_id=s1.id, class_name="P1A", month=10, year=2025, sakit=2, izin=1, alfa=0, entered_by="operator")
    abs2 = AbsenceReason(student_id=s2.id, class_name="P1A", month=10, year=2025, sakit=0, izin=0, alfa=1, entered_by="operator")
    db.add_all([abs1, abs2])

    # 6. Seed Subject & Assessment Components
    sub = Subject(name="Math", jenjang_id=jen_id, supports_sumatif=True, supports_formatif=True)
    db.add(sub)
    db.flush()

    comp_sum = AssessmentComponent(name="Sumatif 1", assessment_type="sumatif", subject_id=sub.id)
    comp_for = AssessmentComponent(name="Formatif 1", assessment_type="formatif", subject_id=sub.id)
    db.add_all([comp_sum, comp_for])
    db.flush()

    # 7. Seed Grades
    g1 = StudentSubjectGrade(enrollment_id=e1.id, subject_id=sub.id, component_id=comp_sum.id, score=90.0)
    g2 = StudentSubjectGrade(enrollment_id=e1.id, subject_id=sub.id, component_id=comp_for.id, score=80.0)
    g3 = StudentSubjectGrade(enrollment_id=e2.id, subject_id=sub.id, component_id=comp_sum.id, score=70.0)
    db.add_all([g1, g2, g3])

    db.commit()
    db.close()

    # Query management-summary
    response = client.get(f"/api/analytics/management-summary?academic_year_id={ay_id}&jenjang_id={jen_id}")
    assert response.status_code == 200
    data = response.json()

    # 8. Assert Attendance Summary
    att_sum = data["attendance_summary"]
    assert att_sum["total_records"] == 6
    assert att_sum["status_counts"]["hadir"] == 2
    assert att_sum["status_counts"]["sakit"] == 2
    assert att_sum["status_counts"]["izin"] == 1
    assert att_sum["status_counts"]["alfa"] == 1

    assert att_sum["status_percentages"]["hadir"] == 33.3
    assert att_sum["status_percentages"]["sakit"] == 33.3
    assert att_sum["status_percentages"]["izin"] == 16.7
    assert att_sum["status_percentages"]["alfa"] == 16.7

    # 9. Assert Lateness by Class
    lat_class = data["lateness_by_class"]
    assert len(lat_class) == 1
    assert lat_class[0]["class_name"] == "P1A"
    assert lat_class[0]["late_days"] == 1
    assert lat_class[0]["late_minutes"] == 45
    assert lat_class[0]["late_duration_label"] == "0:45"
    assert lat_class[0]["late_day_percentage"] == 100.0
    assert lat_class[0]["late_duration_percentage"] == 100.0

    # 10. Assert Grade by Class
    gr_class = data["grade_by_class"]
    assert len(gr_class) == 1
    assert gr_class[0]["class_name"] == "P1A"
    assert gr_class[0]["sumatif_average"] == 80.0
    assert gr_class[0]["formatif_average"] == 80.0
    assert gr_class[0]["student_count"] == 2

    # 11. Assert Grade by Subject
    gr_sub = data["grade_by_subject"]
    assert len(gr_sub) == 1
    assert gr_sub[0]["subject_name"] == "Math"
    assert gr_sub[0]["sumatif_average"] == 80.0
    assert gr_sub[0]["formatif_average"] == 80.0

    # 12. Assert Grade by Student
    gr_stu = data["grade_by_student"]
    assert len(gr_stu) == 2
    s1_entry = next(x for x in gr_stu if x["student_name"] == "Student One")
    s2_entry = next(x for x in gr_stu if x["student_name"] == "Student Two")

    assert s1_entry["sumatif_average"] == 90.0
    assert s1_entry["formatif_average"] == 80.0
    assert s1_entry["below_threshold"] is True

    assert s2_entry["sumatif_average"] == 70.0
    assert s2_entry["formatif_average"] is None
    assert s2_entry["below_threshold"] is True


def test_management_summary_pdf_export_returns_file_and_warning(app_context):
    client = TestClient(app_context["app"])
    ay_id, jen_id, subject_id = seed_management_export_dataset(app_context)

    response = client.get(
        f"/api/analytics/management-summary/export/pdf?academic_year_id={ay_id}&jenjang_id={jen_id}&subject_id={subject_id}"
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert "management-analytics-report-2025-2026-all-terms" in response.headers["content-disposition"]
    assert response.content.startswith(b"%PDF-1.4")
    assert b"No Term filter selected. This report aggregates the full academic year." in response.content
    assert b"Export Student" in response.content
    assert b"Below-KKM Alerts" in response.content


def test_management_summary_pdf_export_respects_term_filter(app_context):
    client = TestClient(app_context["app"])
    ay_id, jen_id, _subject_id = seed_management_export_dataset(app_context)

    response = client.get(
        f"/api/analytics/management-summary/export/pdf?academic_year_id={ay_id}&jenjang_id={jen_id}&term=term_2"
    )

    assert response.status_code == 200
    assert b"Late days: 1" in response.content
    assert b"No Term filter selected" not in response.content
    assert "management-analytics-report-2025-2026-term-2" in response.headers["content-disposition"]


def test_management_summary_excel_export_returns_workbook(app_context):
    client = TestClient(app_context["app"])
    ay_id, jen_id, subject_id = seed_management_export_dataset(app_context)

    response = client.get(
        f"/api/analytics/management-summary/export/excel?academic_year_id={ay_id}&jenjang_id={jen_id}&subject_id={subject_id}&term=term_2"
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(filename=BytesIO(response.content), read_only=True)
    assert "Summary" in workbook.sheetnames
    assert "Below-KKM Alerts" in workbook.sheetnames
    alert_sheet = workbook["Below-KKM Alerts"]
    values = [cell for row in alert_sheet.iter_rows(values_only=True) for cell in row if cell is not None]
    assert "Export Student" in values
    assert "Export Math" in values


def test_management_summary_export_is_read_only(app_context):
    client = TestClient(app_context["app"])
    db_module = app_context["db_module"]
    StudentEnrollment = app_context["StudentEnrollment"]
    UploadLog = app_context["UploadLog"]
    ay_id, jen_id, _subject_id = seed_management_export_dataset(app_context)

    db = db_module.SessionLocal()
    before_upload_logs = db.query(UploadLog).count()
    before_enrollments = db.query(StudentEnrollment).count()
    db.close()

    response = client.get(f"/api/analytics/management-summary/export/pdf?academic_year_id={ay_id}&jenjang_id={jen_id}")
    assert response.status_code == 200

    db = db_module.SessionLocal()
    assert db.query(UploadLog).count() == before_upload_logs
    assert db.query(StudentEnrollment).count() == before_enrollments
    db.close()


def test_kkm_threshold_crud_validation_and_effective_resolution(app_context):
    client = TestClient(app_context["app"])
    ay_id, jen_id, subject_id = seed_management_export_dataset(app_context)

    create_response = client.post(
        "/api/academic-config/kkm-thresholds",
        json={
            "academic_year_id": ay_id,
            "jenjang_id": jen_id,
            "subject_id": subject_id,
            "assessment_type": "sumatif",
            "threshold": 80.0,
        },
    )
    assert create_response.status_code == 200
    created = create_response.json()
    assert created["threshold"] == 80.0

    low_response = client.post(
        "/api/academic-config/kkm-thresholds",
        json={
            "academic_year_id": ay_id,
            "assessment_type": "overall",
            "threshold": -1.0,
        },
    )
    assert low_response.status_code == 422

    high_response = client.post(
        "/api/academic-config/kkm-thresholds",
        json={
            "academic_year_id": ay_id,
            "assessment_type": "overall",
            "threshold": 101.0,
        },
    )
    assert high_response.status_code == 422

    duplicate_response = client.post(
        "/api/academic-config/kkm-thresholds",
        json={
            "academic_year_id": ay_id,
            "jenjang_id": jen_id,
            "subject_id": subject_id,
            "assessment_type": "sumatif",
            "threshold": 82.0,
        },
    )
    assert duplicate_response.status_code == 409

    effective_response = client.get(
        "/api/academic-config/kkm-effective",
        params={
            "academic_year_id": ay_id,
            "jenjang_id": jen_id,
            "subject_id": subject_id,
            "assessment_type": "sumatif",
        },
    )
    assert effective_response.status_code == 200
    effective = effective_response.json()
    assert effective["threshold"] == 80.0
    assert effective["threshold_source"] == "subject-specific"

    fallback_response = client.get(
        "/api/academic-config/kkm-effective",
        params={"academic_year_id": ay_id, "assessment_type": "formatif"},
    )
    assert fallback_response.status_code == 200
    assert fallback_response.json()["threshold"] == 85.0
    assert fallback_response.json()["threshold_source"] == "legacy-fallback"

    update_response = client.put(
        f"/api/academic-config/kkm-thresholds/{created['id']}",
        json={"threshold": 78.0},
    )
    assert update_response.status_code == 200
    assert update_response.json()["threshold"] == 78.0

    delete_response = client.delete(f"/api/academic-config/kkm-thresholds/{created['id']}")
    assert delete_response.status_code == 200
    assert delete_response.json()["deleted"] == 1


def test_term_config_crud_validation_and_effective_terms(app_context):
    client = TestClient(app_context["app"])
    ay_id, _jen_id, _subject_id = seed_management_export_dataset(app_context)

    default_response = client.get(f"/api/academic-config/terms/effective?academic_year_id={ay_id}")
    assert default_response.status_code == 200
    default_terms = default_response.json()
    assert default_terms[1]["source"] == "default"
    assert default_terms[1]["start_date"] == "2025-10-01"

    create_response = client.post(
        "/api/academic-config/terms",
        json={
            "academic_year_id": ay_id,
            "term_number": 2,
            "label": "Custom Intervention Term",
            "start_date": "2025-12-01",
            "end_date": "2025-12-31",
        },
    )
    assert create_response.status_code == 200
    created = create_response.json()
    assert created["source"] == "custom"

    invalid_number = client.post(
        "/api/academic-config/terms",
        json={
            "academic_year_id": ay_id,
            "term_number": 5,
            "label": "Bad Term",
            "start_date": "2025-10-01",
            "end_date": "2025-10-31",
        },
    )
    assert invalid_number.status_code == 422

    invalid_order = client.post(
        "/api/academic-config/terms",
        json={
            "academic_year_id": ay_id,
            "term_number": 1,
            "label": "Bad Dates",
            "start_date": "2025-09-30",
            "end_date": "2025-07-01",
        },
    )
    assert invalid_order.status_code == 400

    overlap_response = client.post(
        "/api/academic-config/terms",
        json={
            "academic_year_id": ay_id,
            "term_number": 1,
            "label": "Overlapping Term",
            "start_date": "2025-09-15",
            "end_date": "2025-12-15",
        },
    )
    assert overlap_response.status_code == 400

    effective_response = client.get(f"/api/academic-config/terms/effective?academic_year_id={ay_id}")
    assert effective_response.status_code == 200
    effective_term_2 = next(row for row in effective_response.json() if row["term_number"] == 2)
    assert effective_term_2["source"] == "custom"
    assert effective_term_2["start_date"] == "2025-12-01"

    update_response = client.put(
        f"/api/academic-config/terms/{created['id']}",
        json={"label": "Updated Intervention Term", "start_date": "2025-11-01", "end_date": "2025-11-30"},
    )
    assert update_response.status_code == 200
    assert update_response.json()["label"] == "Updated Intervention Term"

    delete_response = client.delete(f"/api/academic-config/terms/{created['id']}")
    assert delete_response.status_code == 200
    assert delete_response.json()["deleted"] == 1


def test_management_summary_uses_configured_kkm_and_term_config(app_context):
    client = TestClient(app_context["app"])
    ay_id, jen_id, subject_id = seed_management_export_dataset(app_context)

    client.post(
        "/api/academic-config/kkm-thresholds",
        json={
            "academic_year_id": ay_id,
            "jenjang_id": jen_id,
            "subject_id": subject_id,
            "assessment_type": "sumatif",
            "threshold": 70.0,
        },
    )
    client.post(
        "/api/academic-config/terms",
        json={
            "academic_year_id": ay_id,
            "term_number": 2,
            "label": "No October Term",
            "start_date": "2025-12-01",
            "end_date": "2025-12-31",
        },
    )

    summary_response = client.get(
        f"/api/analytics/management-summary?academic_year_id={ay_id}&jenjang_id={jen_id}&subject_id={subject_id}&term=term_2"
    )
    assert summary_response.status_code == 200
    summary = summary_response.json()
    assert summary["filters"]["date_start"] == "2025-12-01"
    assert summary["filters"]["term_source"] == "custom"
    assert summary["lateness_by_class"] == []
    student = summary["grade_by_student"][0]
    assert student["sumatif_kkm_threshold"] == 70.0
    assert student["sumatif_threshold_source"] == "subject-specific"
    assert student["below_threshold"] is False


def test_management_exports_include_configured_kkm_and_term_metadata(app_context):
    client = TestClient(app_context["app"])
    ay_id, jen_id, subject_id = seed_management_export_dataset(app_context)

    client.post(
        "/api/academic-config/kkm-thresholds",
        json={
            "academic_year_id": ay_id,
            "jenjang_id": jen_id,
            "subject_id": subject_id,
            "assessment_type": "sumatif",
            "threshold": 80.0,
        },
    )
    client.post(
        "/api/academic-config/terms",
        json={
            "academic_year_id": ay_id,
            "term_number": 2,
            "label": "Configured Term 2",
            "start_date": "2025-12-01",
            "end_date": "2025-12-31",
        },
    )

    pdf_response = client.get(
        f"/api/analytics/management-summary/export/pdf?academic_year_id={ay_id}&jenjang_id={jen_id}&subject_id={subject_id}&term=term_2"
    )
    assert pdf_response.status_code == 200
    assert b"Configured Term 2" in pdf_response.content
    assert b"source subject-specific" in pdf_response.content

    excel_response = client.get(
        f"/api/analytics/management-summary/export/excel?academic_year_id={ay_id}&jenjang_id={jen_id}&subject_id={subject_id}&term=term_2"
    )
    assert excel_response.status_code == 200
    workbook = load_workbook(filename=BytesIO(excel_response.content), read_only=True)
    summary_values = [cell for row in workbook["Summary"].iter_rows(values_only=True) for cell in row if cell is not None]
    alert_values = [cell for row in workbook["Below-KKM Alerts"].iter_rows(values_only=True) for cell in row if cell is not None]
    assert "Configured Term 2" in summary_values
    assert 80.0 in alert_values
    assert "subject-specific" in alert_values


def test_academic_config_operations_are_read_only_for_operational_tables(app_context):
    client = TestClient(app_context["app"])
    db_module = app_context["db_module"]
    StudentEnrollment = app_context["StudentEnrollment"]
    UploadLog = app_context["UploadLog"]
    ay_id, jen_id, subject_id = seed_management_export_dataset(app_context)

    db = db_module.SessionLocal()
    before_upload_logs = db.query(UploadLog).count()
    before_enrollments = db.query(StudentEnrollment).count()
    db.close()

    response = client.post(
        "/api/academic-config/kkm-thresholds",
        json={
            "academic_year_id": ay_id,
            "jenjang_id": jen_id,
            "subject_id": subject_id,
            "assessment_type": "sumatif",
            "threshold": 80.0,
        },
    )
    assert response.status_code == 200

    db = db_module.SessionLocal()
    assert db.query(UploadLog).count() == before_upload_logs
    assert db.query(StudentEnrollment).count() == before_enrollments
    db.close()


def _intervention_payload(app_context, ay_id, jen_id, subject_id, status="open"):
    db_module = app_context["db_module"]
    Student = app_context["Student"]
    StudentEnrollment = app_context["StudentEnrollment"]

    db = db_module.SessionLocal()
    student = db.query(Student).filter(Student.name == "Export Student").first()
    enrollment = db.query(StudentEnrollment).filter(StudentEnrollment.student_id == student.id).first()
    payload = {
        "student_id": student.id,
        "enrollment_id": enrollment.id,
        "academic_year_id": ay_id,
        "jenjang_id": jen_id,
        "subject_id": subject_id,
        "assessment_type": "sumatif",
        "term": "term_2",
        "class_name": "P3A",
        "student_name": "Export Student",
        "subject_name": "Export Math",
        "effective_threshold": 85.0,
        "threshold_source": "legacy-fallback",
        "current_average": 74.0,
        "status": status,
        "priority": "high",
        "owner_name": "Academic Lead",
        "planned_action": "Schedule remedial review",
        "notes": "Initial intervention plan",
        "follow_up_date": "2025-11-15",
    }
    db.close()
    return payload


def test_academic_intervention_crud_duplicate_and_filtering(app_context):
    client = TestClient(app_context["app"])
    ay_id, jen_id, subject_id = seed_management_export_dataset(app_context)
    payload = _intervention_payload(app_context, ay_id, jen_id, subject_id)

    create_response = client.post("/api/academic-interventions/from-alert", json=payload)
    assert create_response.status_code == 200
    created = create_response.json()
    assert created["status"] == "open"
    assert created["priority"] == "high"

    duplicate_response = client.post("/api/academic-interventions", json=payload)
    assert duplicate_response.status_code == 409

    invalid_status = client.post("/api/academic-interventions", json={**payload, "status": "waiting"})
    assert invalid_status.status_code == 422

    invalid_threshold = client.post("/api/academic-interventions", json={**payload, "effective_threshold": 101.0})
    assert invalid_threshold.status_code == 422

    update_response = client.patch(
        f"/api/academic-interventions/{created['id']}",
        json={
            "status": "resolved",
            "priority": "medium",
            "owner_name": "Counselor",
            "outcome": "Student completed remediation",
        },
    )
    assert update_response.status_code == 200
    updated = update_response.json()
    assert updated["status"] == "resolved"
    assert updated["resolved_at"] is not None
    assert updated["owner_name"] == "Counselor"

    second_response = client.post("/api/academic-interventions", json={**payload, "priority": "urgent"})
    assert second_response.status_code == 200
    assert second_response.json()["priority"] == "urgent"

    list_response = client.get(
        "/api/academic-interventions",
        params={"academic_year_id": ay_id, "subject_id": subject_id, "status": "open", "priority": "urgent"},
    )
    assert list_response.status_code == 200
    rows = list_response.json()
    assert len(rows) == 1
    assert rows[0]["student_name"] == "Export Student"


def test_management_summary_and_exports_include_intervention_metadata(app_context):
    client = TestClient(app_context["app"])
    ay_id, jen_id, subject_id = seed_management_export_dataset(app_context)
    payload = _intervention_payload(app_context, ay_id, jen_id, subject_id)
    create_response = client.post("/api/academic-interventions/from-alert", json=payload)
    assert create_response.status_code == 200
    intervention_id = create_response.json()["id"]

    summary_response = client.get(
        f"/api/analytics/management-summary?academic_year_id={ay_id}&jenjang_id={jen_id}&subject_id={subject_id}&term=term_2"
    )
    assert summary_response.status_code == 200
    alerts = summary_response.json()["below_kkm_alerts"]
    assert alerts[0]["intervention_id"] == intervention_id
    assert alerts[0]["intervention_status"] == "open"
    assert alerts[0]["intervention_priority"] == "high"
    assert alerts[0]["intervention_owner"] == "Academic Lead"
    assert alerts[0]["follow_up_date"] == "2025-11-15"

    pdf_response = client.get(
        f"/api/analytics/management-summary/export/pdf?academic_year_id={ay_id}&jenjang_id={jen_id}&subject_id={subject_id}&term=term_2"
    )
    assert pdf_response.status_code == 200
    assert b"intervention open" in pdf_response.content
    assert b"Academic Lead" in pdf_response.content

    excel_response = client.get(
        f"/api/analytics/management-summary/export/excel?academic_year_id={ay_id}&jenjang_id={jen_id}&subject_id={subject_id}&term=term_2"
    )
    assert excel_response.status_code == 200
    workbook = load_workbook(filename=BytesIO(excel_response.content), read_only=True)
    values = [cell for row in workbook["Below-KKM Alerts"].iter_rows(values_only=True) for cell in row if cell is not None]
    assert "Intervention status" in values
    assert "Academic Lead" in values
    assert intervention_id in values


def test_analytics_filters(app_context):
    client = TestClient(app_context["app"])
    db_module = app_context["db_module"]
    AcademicYear = app_context["AcademicYear"]
    Jenjang = app_context["Jenjang"]
    Subject = app_context["Subject"]
    StudentEnrollment = app_context["StudentEnrollment"]
    Student = app_context["Student"]

    db = db_module.SessionLocal()
    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    jen = db.query(Jenjang).filter(Jenjang.name == "Primary").first()
    ay_id = ay.id
    jen_id = jen.id

    # Seed student first to satisfy foreign key constraint
    s = Student(id=99, name="Test Student")
    db.add(s)
    db.flush()

    # Seed an enrollment to generate class names
    e = StudentEnrollment(student_id=99, academic_year_id=ay_id, jenjang_id=jen_id, class_name="P2B", class_assigned=True)
    db.add(e)

    db.commit()
    db.close()

    response = client.get(f"/api/analytics/filters?academic_year_id={ay_id}&jenjang_id={jen_id}")
    assert response.status_code == 200
    data = response.json()

    assert len(data["academic_years"]) == 1
    assert data["academic_years"][0]["label"] == "2025/2026"
    assert len(data["jenjangs"]) == 1
    assert data["jenjangs"][0]["name"] == "Primary"
    assert "P2B" in data["class_names"]
    assert len(data["subjects"]) == 1
    assert data["subjects"][0]["name"] == "Language"


def test_legacy_analytics_filters_route_is_preserved_temporarily(app_context):
    client = TestClient(app_context["app"])
    db_module = app_context["db_module"]
    AcademicYear = app_context["AcademicYear"]
    db = db_module.SessionLocal()
    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    ay_id = ay.id
    db.close()

    response = client.get(f"/analytics/filters?academic_year_id={ay_id}")
    assert response.status_code == 200


def test_management_summary_export_excel_editable(app_context):
    client = TestClient(app_context["app"])
    db_module = app_context["db_module"]
    AcademicYear = app_context["AcademicYear"]
    db = db_module.SessionLocal()
    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    ay_id = ay.id
    db.close()

    response = client.get(
        f"/api/analytics/management-summary/export/excel?academic_year_id={ay_id}&mode=editable"
    )
    assert response.status_code == 200
    workbook = load_workbook(filename=BytesIO(response.content), read_only=True)
    sheet_names = workbook.sheetnames
    assert "README" in sheet_names
    assert "Config" in sheet_names
    assert "Attendance_Data" in sheet_names
    assert "Charts" in sheet_names
