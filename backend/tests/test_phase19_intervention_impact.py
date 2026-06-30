from __future__ import annotations

import importlib
import io
import sys
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

MODULE_PREFIXES = ("src", "api", "core", "models", "services")
SOURCE_ROOT = Path(__file__).resolve().parents[1] / "src"


def unload_app_modules() -> None:
    for name in list(sys.modules):
        if name == "src" or name.startswith(MODULE_PREFIXES):
            sys.modules.pop(name, None)


@pytest.fixture
def impact_db(monkeypatch, tmp_path):
    db_path = tmp_path / "phase19.db"
    monkeypatch.syspath_prepend(str(SOURCE_ROOT))
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    unload_app_modules()
    db_module = importlib.import_module("core.database")
    db_module.init_db()

    AcademicYear = importlib.import_module("models.academic_year").AcademicYear
    Jenjang = importlib.import_module("models.jenjang").Jenjang
    Student = importlib.import_module("models.student").Student
    StudentEnrollment = importlib.import_module("models.student_enrollment").StudentEnrollment
    Subject = importlib.import_module("models.subject").Subject
    AssessmentComponent = importlib.import_module("models.assessment_component").AssessmentComponent
    StudentSubjectGrade = importlib.import_module("models.student_subject_grade").StudentSubjectGrade
    AcademicIntervention = importlib.import_module("models.academic_intervention").AcademicIntervention

    db = db_module.SessionLocal()
    db.query(AcademicYear).update({AcademicYear.is_default: False})
    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    if ay is None:
        ay = AcademicYear(label="2025/2026", start_date=date(2025, 7, 1), end_date=date(2026, 6, 30), status="active", is_default=True)
        db.add(ay)
    else:
        ay.start_date = date(2025, 7, 1)
        ay.end_date = date(2026, 6, 30)
        ay.is_default = True
    jen = db.query(Jenjang).filter(Jenjang.name == "Primary").first()
    if jen is None:
        jen = Jenjang(name="Primary")
        db.add(jen)
    db.flush()

    math = Subject(name="Math", jenjang_id=jen.id, supports_sumatif=True, supports_formatif=True)
    science = Subject(name="Science", jenjang_id=jen.id, supports_sumatif=True, supports_formatif=True)
    db.add_all([math, science])
    db.flush()
    math_sum = AssessmentComponent(name="Math Sumatif", assessment_type="sumatif", subject_id=math.id)
    sci_sum = AssessmentComponent(name="Science Sumatif", assessment_type="sumatif", subject_id=science.id)
    db.add_all([math_sum, sci_sum])
    db.flush()

    alice = Student(name="Alice", jenjang="Primary", class_name="P3")
    bob = Student(name="Bob", jenjang="Primary", class_name="P4")
    db.add_all([alice, bob])
    db.flush()
    en_alice = StudentEnrollment(student_id=alice.id, academic_year_id=ay.id, jenjang_id=jen.id, class_name="P3", class_assigned=True)
    en_bob = StudentEnrollment(student_id=bob.id, academic_year_id=ay.id, jenjang_id=jen.id, class_name="P4", class_assigned=True)
    db.add_all([en_alice, en_bob])
    db.flush()
    db.add_all([
        StudentSubjectGrade(enrollment_id=en_alice.id, subject_id=math.id, component_id=math_sum.id, score=86.0),
        StudentSubjectGrade(enrollment_id=en_bob.id, subject_id=science.id, component_id=sci_sum.id, score=70.0),
    ])
    db.flush()
    db.add_all([
        AcademicIntervention(
            student_id=alice.id, enrollment_id=en_alice.id, academic_year_id=ay.id, jenjang_id=jen.id,
            subject_id=math.id, assessment_type="sumatif", term="term_3", class_name="P3",
            student_name="Alice", subject_name="Math", effective_threshold=80.0,
            threshold_source="subject-specific", current_average=72.0, status="resolved",
            priority="high", owner_name="Teacher A", follow_up_date=date(2026, 3, 15),
            created_at=datetime(2026, 3, 1), resolved_at=datetime(2026, 3, 20),
        ),
        AcademicIntervention(
            student_id=bob.id, enrollment_id=en_bob.id, academic_year_id=ay.id, jenjang_id=jen.id,
            subject_id=science.id, assessment_type="sumatif", term="term_3", class_name="P4",
            student_name="Bob", subject_name="Science", effective_threshold=85.0,
            threshold_source="legacy-fallback", current_average=72.0, status="open",
            priority="urgent", owner_name="Teacher B", follow_up_date=date(2026, 1, 10),
            created_at=datetime(2026, 1, 1),
        ),
    ])
    db.commit()
    yield {"db": db, "ay_id": ay.id, "jen_id": jen.id, "math_id": math.id}
    db.close()


def _impact(impact_db):
    service = importlib.import_module("services.intervention_impact")
    return service.build_intervention_impact(
        impact_db["db"],
        academic_year_id=impact_db["ay_id"],
        jenjang_id=impact_db["jen_id"],
    )


def test_intervention_impact_endpoint_route_exists(impact_db):
    main = importlib.import_module("main")
    route_paths = {getattr(route, "path", "") for route in main.app.routes}
    assert "/api/analytics/intervention-impact" in route_paths


def test_score_delta_moved_above_kkm_and_days_open(impact_db):
    data = _impact(impact_db)
    alice = next(row for row in data["impact_rows"] if row["student_name"] == "Alice")
    assert alice["baseline_average"] == 72.0
    assert alice["latest_average"] == 86.0
    assert alice["score_delta"] == 14.0
    assert alice["moved_above_kkm"] is True
    assert alice["days_open"] == 19
    assert alice["threshold_source"] == "subject-specific"


def test_overdue_risk_reasons_summary_and_breakdowns(impact_db):
    data = _impact(impact_db)
    bob = next(row for row in data["impact_rows"] if row["student_name"] == "Bob")
    assert bob["is_overdue"] is True
    assert bob["risk_level"] == "critical"
    assert "Still below effective KKM" in bob["risk_reasons"]
    assert "Follow-up overdue" in bob["risk_reasons"]
    assert "No score improvement after intervention" in bob["risk_reasons"]
    assert data["summary"]["total_interventions"] == 2
    assert data["summary"]["overdue_interventions"] == 1
    assert data["summary"]["percent_improved"] == 50.0
    assert data["summary"]["percent_moved_above_kkm"] == 50.0
    assert any(row["class_name"] == "P4" and row["high_risk_count"] == 1 for row in data["class_breakdown"])
    assert any(row["subject_name"] == "Science" and row["overdue_interventions"] == 1 for row in data["subject_breakdown"])
    assert any(row["owner_name"] == "Teacher B" and row["overdue_interventions"] == 1 for row in data["owner_workload_summary"])
    assert any(insight["category"] == "intervention_impact" for insight in data["executive_insights"])


def test_pdf_excel_intervention_impact_parity(impact_db):
    summary_builder = importlib.import_module("services.management_analytics").build_management_summary
    exports = importlib.import_module("services.management_report_export")
    summary = summary_builder(impact_db["db"], academic_year_id=impact_db["ay_id"], jenjang_id=impact_db["jen_id"])
    summary["intervention_impact"] = _impact(impact_db)
    pdf_bytes = exports.build_management_summary_pdf(summary)
    assert b"Intervention Impact Analysis" in pdf_bytes
    assert b"Impact rows: 2" in pdf_bytes

    workbook = load_workbook(io.BytesIO(exports.build_management_summary_excel(summary, {"mode": "editable"})), read_only=False)
    for sheet_name in [
        "Intervention_Impact_Data",
        "Intervention_Impact_Summary",
        "Risk_Students_Data",
        "Owner_Workload_Data",
    ]:
        assert sheet_name in workbook.sheetnames
    rows = {
        workbook["Intervention_Impact_Data"].cell(row=row_idx, column=2).value: row_idx
        for row_idx in range(2, workbook["Intervention_Impact_Data"].max_row + 1)
    }
    assert workbook["Intervention_Impact_Data"].cell(row=rows["Alice"], column=9).value == 14
    assert workbook["Intervention_Impact_Data"].cell(row=rows["Bob"], column=13).value == "critical"
