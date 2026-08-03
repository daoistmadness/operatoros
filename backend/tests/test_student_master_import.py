from datetime import date

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from core.database import Base
from core.student_master_import import apply_import, validate_import
from core.student_import_schema_migration import ensure_student_import_schema
from models.academic_master import AcademicClass, AcademicGrade, AcademicProgram
from models.academic_year import AcademicYear
from models.attendance import Attendance
from models.jenjang import Jenjang
from models.student import Student
from models.student_enrollment import StudentEnrollment
from models.student_master import StudentDeviceIdentity, StudentMaster


def _workbook(path, *, student=None, class_values=None):
    student = student or {
        "peserta_didik_id": "PD-001", "name": "Synthetic Student", "school_id": "SCH-1",
        "rombongan_belajar_id": "ROM-1", "nis": "000123", "nisn": "0000000001",
        "gender": "female", "birth_place": "Bandung", "birth_date": date(2012, 1, 2),
        "status": "active", "legacy_student_id": None,
    }
    class_values = class_values or {
        "rombongan_belajar_id": "ROM-1", "semester_id": "SEM-1", "school_id": "SCH-1",
        "jenjang": "Primary", "program": "Primary", "grade": "Primary 1", "class_name": "P1A",
    }
    workbook = Workbook()
    school = workbook.active
    school.title = "School"
    school.append(["school_id", "name", "npsn"])
    school.append(["SCH-1", "Synthetic School", "00000001"])
    classes = workbook.create_sheet("Classes")
    classes.append(list(class_values))
    classes.append([class_values[key] for key in class_values])
    students = workbook.create_sheet("Students")
    headers = [
        "peserta_didik_id", "name", "school_id", "rombongan_belajar_id", "nis", "nisn",
        "gender", "birth_place", "birth_date", "status", "legacy_student_id",
    ]
    students.append(headers)
    students.append([student.get(key) for key in headers])
    workbook.save(path)


@pytest.fixture
def import_db(tmp_path):
    path = tmp_path / "student-import.sqlite"
    engine = create_engine(f"sqlite:///{path}")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    db = Session()
    year = AcademicYear(label="2026/2027", start_date=date(2026, 7, 1), end_date=date(2027, 6, 30), status="active", is_default=True)
    jenjang = Jenjang(name="Primary", active=True)
    db.add_all([year, jenjang]); db.flush()
    program = AcademicProgram(jenjang_id=jenjang.id, name="Primary", active=True)
    db.add(program); db.flush()
    grade = AcademicGrade(jenjang_id=jenjang.id, program_id=program.id, name="Primary 1", sequence_number=1, active=True)
    db.add(grade); db.flush()
    db.add(AcademicClass(academic_year_id=year.id, grade_id=grade.id, class_name="P1A", section_code="A", active=True))
    db.commit(); db.close(); engine.dispose()
    ensure_student_import_schema(path)
    yield path


def test_validate_is_read_only_and_classifies_new_student(import_db, tmp_path):
    source = tmp_path / "students.xlsx"
    _workbook(source)
    engine = create_engine(f"sqlite:///{import_db}")
    Session = sessionmaker(bind=engine)
    before = {table: Session().execute(text(f"SELECT COUNT(*) FROM {table}")).scalar() for table in ("student_masters", "student_enrollments", "student_import_sessions")}
    Session().close(); engine.dispose()

    report = validate_import(source, import_db, "2026/2027")
    assert report["status"] == "READY"
    assert report["counts"]["NEW_STUDENT"] == 1
    assert report["rows"][0]["identifiers"]["nisn"] == "******0001"
    assert "Synthetic Student" not in str(report)

    engine = create_engine(f"sqlite:///{import_db}")
    with engine.connect() as connection:
        after = {table: connection.exec_driver_sql(f"SELECT COUNT(*) FROM {table}").scalar() for table in before}
    engine.dispose()
    assert after == before


def test_apply_creates_canonical_student_enrollment_and_is_idempotent(import_db, tmp_path):
    source = tmp_path / "students.xlsx"
    _workbook(source)
    first = apply_import(source, import_db, "2026/2027", [2], "admin", "Synthetic Registrar", date(2026, 8, 1), "APPLY_CANONICAL_STUDENT_IMPORT")
    assert first["status"] == "committed"
    engine = create_engine(f"sqlite:///{import_db}")
    Session = sessionmaker(bind=engine)
    db = Session()
    assert db.query(StudentMaster).count() == 1
    assert db.query(StudentEnrollment).count() == 1
    master = db.query(StudentMaster).one()
    assert master.nipd == "000123"
    assert master.nisn == "0000000001"
    assert db.query(StudentDeviceIdentity).count() == 0
    db.close(); engine.dispose()

    replay = apply_import(source, import_db, "2026/2027", [2], "admin", "Synthetic Registrar", date(2026, 8, 1), "APPLY_CANONICAL_STUDENT_IMPORT")
    assert replay["idempotent_replay"] is True


def test_same_name_and_birth_date_is_review_only(import_db, tmp_path):
    engine = create_engine(f"sqlite:///{import_db}")
    Session = sessionmaker(bind=engine)
    db = Session()
    db.add(StudentMaster(full_name="Synthetic Student", normalized_name="synthetic student", birth_date=date(2012, 1, 2), student_status="active"))
    db.commit(); db.close(); engine.dispose()
    source = tmp_path / "duplicate.xlsx"
    _workbook(source)
    report = validate_import(source, import_db, "2026/2027")
    assert report["rows"][0]["classification"] == "REVIEW_REQUIRED"
    assert "POSSIBLE_DUPLICATE" in report["rows"][0]["errors"]


def test_explicit_legacy_link_preserves_attendance(import_db, tmp_path):
    engine = create_engine(f"sqlite:///{import_db}")
    Session = sessionmaker(bind=engine)
    db = Session()
    db.add(Student(id=901, name="Legacy Student")); db.flush()
    db.add(Attendance(student_id=901, date=date(2026, 8, 1), status="on-time", late_duration=0, is_absent=False))
    db.commit(); db.close(); engine.dispose()
    source = tmp_path / "linked.xlsx"
    _workbook(source, student={
        "peserta_didik_id": "PD-002", "name": "Legacy Student", "school_id": "SCH-1",
        "rombongan_belajar_id": "ROM-1", "nis": "000124", "nisn": "0000000002",
        "gender": "male", "birth_place": "Bandung", "birth_date": date(2012, 2, 2),
        "status": "active", "legacy_student_id": "901",
    })
    result = apply_import(source, import_db, "2026/2027", [2], "admin", "Synthetic Registrar", date(2026, 8, 1), "APPLY_CANONICAL_STUDENT_IMPORT")
    assert result["status"] == "committed"
    engine = create_engine(f"sqlite:///{import_db}")
    Session = sessionmaker(bind=engine)
    db = Session()
    enrollment = db.query(StudentEnrollment).one()
    assert enrollment.student_id == 901
    assert db.query(Attendance).one().student_id == 901
    assert db.query(StudentDeviceIdentity).one().device_source == "legacy_students"
    db.close(); engine.dispose()


def test_protected_attendance_filename_is_rejected_without_opening(tmp_path):
    with pytest.raises(ValueError, match="PROTECTED_OPERATIONAL_DATABASE_REJECTED"):
        validate_import(tmp_path / "missing.xlsx", "/tmp/attendance.db", "2026/2027")


def test_schema_upgrade_is_idempotent_and_integrity_safe(import_db):
    assert ensure_student_import_schema(import_db) == ensure_student_import_schema(import_db)
    engine = create_engine(f"sqlite:///{import_db}")
    with engine.connect() as connection:
        assert connection.execute(text("PRAGMA integrity_check")).scalar() == "ok"
        assert connection.execute(text("PRAGMA foreign_key_check")).all() == []
        indexes = {row[1] for row in connection.execute(text("PRAGMA index_list(student_masters)"))}
    engine.dispose()
    assert "uq_student_masters_dapodik_peserta_didik_id" in indexes


def test_duplicate_and_suspicious_source_values_are_review_only(import_db, tmp_path):
    source = tmp_path / "unsafe.xlsx"
    _workbook(source)
    workbook = load_workbook(source)
    students = workbook["Students"]
    students["F2"] = 123
    students["I2"] = "not-a-date"
    students.append([cell.value for cell in students[2]])
    workbook.save(source)
    report = validate_import(source, import_db, "2026/2027")
    assert all(row["classification"] == "REVIEW_REQUIRED" for row in report["rows"])
    errors = {error for row in report["rows"] for error in row["errors"]}
    assert "SUSPICIOUS_NUMERIC_IDENTIFIER" in errors
    assert "INVALID_DATE" in errors
    assert "DUPLICATE_SOURCE_IDENTIFIER" in errors
    assert "Synthetic Student" not in str(report)
