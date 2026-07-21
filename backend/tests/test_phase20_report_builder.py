import importlib
import io
import sqlite3
import sys
from datetime import date, datetime, time
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

MODULE_PREFIXES = ("src", "api", "core", "models", "services")
SOURCE_ROOT = Path(__file__).resolve().parents[1] / "src"


def unload_app_modules() -> None:
    for name in list(sys.modules):
        if name == "src" or name.startswith(MODULE_PREFIXES):
            sys.modules.pop(name, None)


def create_legacy_grade_table(db_path: Path) -> None:
    connection = sqlite3.connect(db_path)
    try:
        connection.execute(
            """
            CREATE TABLE student_term_grades (
                id INTEGER PRIMARY KEY,
                student_id INTEGER NOT NULL,
                academic_year VARCHAR NOT NULL,
                term_1 FLOAT,
                term_2 FLOAT,
                term_3 FLOAT,
                term_4 FLOAT
            )
            """
        )
        connection.commit()
    finally:
        connection.close()


@pytest.fixture
def builder_app(monkeypatch, tmp_path):
    import src.main as main_module
    
    db_path = tmp_path / "report-builder.db"
    create_legacy_grade_table(db_path)
    monkeypatch.syspath_prepend(str(SOURCE_ROOT))
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    
    # Snapshot overrides
    original_overrides = main_module.app.dependency_overrides.copy()
    main_module.app.dependency_overrides.clear()
    
    unload_app_modules()
    
    importlib.invalidate_caches()
    main_module = importlib.import_module("src.main")
    db_module = importlib.import_module("core.database")
    student_module = importlib.import_module("models.student")
    attendance_module = importlib.import_module("models.attendance")
    academic_year_module = importlib.import_module("models.academic_year")
    jenjang_module = importlib.import_module("models.jenjang")
    subject_module = importlib.import_module("models.subject")
    enrollment_module = importlib.import_module("models.student_enrollment")
    component_module = importlib.import_module("models.assessment_component")
    grade_module = importlib.import_module("models.student_subject_grade")
    config_module = importlib.import_module("models.academic_config")
    intervention_module = importlib.import_module("models.academic_intervention")
    report_module = importlib.import_module("models.report_builder")

    db_module.init_db()
    db = db_module.SessionLocal()
    AcademicYear = academic_year_module.AcademicYear
    Jenjang = jenjang_module.Jenjang
    Subject = subject_module.Subject
    Student = student_module.Student
    Attendance = attendance_module.Attendance
    StudentEnrollment = enrollment_module.StudentEnrollment
    AssessmentComponent = component_module.AssessmentComponent
    StudentSubjectGrade = grade_module.StudentSubjectGrade
    KkmThreshold = config_module.KkmThreshold
    AcademicTermConfig = config_module.AcademicTermConfig
    AcademicIntervention = intervention_module.AcademicIntervention
    ReportTemplate = report_module.ReportTemplate

    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    if ay is None:
        ay = AcademicYear(label="2025/2026", start_date=date(2025, 7, 1), end_date=date(2026, 6, 30), status="active", is_default=True)
        db.add(ay)
        db.flush()
    jen = db.query(Jenjang).filter(Jenjang.name == "Primary").first()
    if jen is None:
        jen = Jenjang(name="Primary")
        db.add(jen)
        db.flush()
    subject = db.query(Subject).filter(Subject.name == "Math", Subject.jenjang_id == jen.id).first()
    if subject is None:
        subject = Subject(name="Math", jenjang_id=jen.id, supports_sumatif=True, supports_formatif=True)
        db.add(subject)
        db.flush()
    student = db.query(Student).filter(Student.name == "Template Student").first()
    if student is None:
        student = Student(name="Template Student", jenjang="Primary", class_name="P3A")
        db.add(student)
        db.flush()
    enrollment = db.query(StudentEnrollment).filter(StudentEnrollment.student_id == student.id, StudentEnrollment.academic_year_id == ay.id).first()
    if enrollment is None:
        enrollment = StudentEnrollment(student_id=student.id, academic_year_id=ay.id, jenjang_id=jen.id, class_name="P3A", class_assigned=True)
        db.add(enrollment)
        db.flush()
    sum_component = db.query(AssessmentComponent).filter(AssessmentComponent.name == "Math Sumatif", AssessmentComponent.subject_id == subject.id).first()
    if sum_component is None:
        sum_component = AssessmentComponent(name="Math Sumatif", assessment_type="sumatif", subject_id=subject.id)
        db.add(sum_component)
        db.flush()
    for_component = db.query(AssessmentComponent).filter(AssessmentComponent.name == "Math Formatif", AssessmentComponent.subject_id == subject.id).first()
    if for_component is None:
        for_component = AssessmentComponent(name="Math Formatif", assessment_type="formatif", subject_id=subject.id)
        db.add(for_component)
        db.flush()

    if db.query(Attendance).count() == 0:
      db.add_all(
        [
          Attendance(student_id=student.id, date=date(2025, 7, 10), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0),
          Attendance(student_id=student.id, date=date(2025, 10, 10), check_in=time(7, 35), check_out=time(14, 0), status="late", late_duration=35),
        ]
      )

    if db.query(StudentSubjectGrade).count() == 0:
      db.add_all(
        [
          StudentSubjectGrade(enrollment_id=enrollment.id, subject_id=subject.id, component_id=sum_component.id, score=74.0),
          StudentSubjectGrade(enrollment_id=enrollment.id, subject_id=subject.id, component_id=for_component.id, score=82.0),
        ]
      )

    if db.query(KkmThreshold).count() == 0:
      db.add(KkmThreshold(academic_year_id=ay.id, jenjang_id=jen.id, subject_id=subject.id, assessment_type="sumatif", threshold=80.0))

    if db.query(AcademicTermConfig).count() == 0:
      db.add_all(
        [
          AcademicTermConfig(academic_year_id=ay.id, term_number=1, label="Term 1", start_date=date(2025, 7, 1), end_date=date(2025, 9, 30)),
          AcademicTermConfig(academic_year_id=ay.id, term_number=2, label="Term 2", start_date=date(2025, 10, 1), end_date=date(2025, 12, 31)),
          AcademicTermConfig(academic_year_id=ay.id, term_number=3, label="Term 3", start_date=date(2026, 1, 1), end_date=date(2026, 3, 31)),
          AcademicTermConfig(academic_year_id=ay.id, term_number=4, label="Term 4", start_date=date(2026, 4, 1), end_date=date(2026, 6, 30)),
        ]
      )

    if db.query(AcademicIntervention).count() == 0:
      db.add(
        AcademicIntervention(
          student_id=student.id,
          academic_year_id=ay.id,
          jenjang_id=jen.id,
          subject_id=subject.id,
          class_name="P3A",
          student_name=student.name,
          subject_name=subject.name,
          assessment_type="sumatif",
          term="term_1",
          effective_threshold=80.0,
          threshold_source="configured",
          current_average=74.0,
          status="resolved",
          priority="high",
          owner_name="Teacher A",
          follow_up_date=date(2025, 7, 20),
          created_at=datetime(2025, 7, 12),
          resolved_at=datetime(2025, 7, 20),
        )
      )

    db.commit()
    db.close()

    sec_module = importlib.import_module("security.dependencies")
    main_module.app.dependency_overrides[sec_module.get_current_user] = lambda: type("MockUser", (), {"role": "admin"})()

    try:
        yield {
            "app": main_module.app,
            "db_module": db_module,
            "ReportTemplate": ReportTemplate,
            "AcademicYear": AcademicYear,
            "Jenjang": Jenjang,
            "Subject": Subject,
        }
    finally:
        main_module.app.dependency_overrides.clear()
        main_module.app.dependency_overrides.update(original_overrides)
        if hasattr(db_module, "engine"):
            db_module.engine.dispose()


def test_report_builder_routes_and_default_seeding(builder_app):
    client = TestClient(builder_app["app"])
    route_paths = {getattr(route, "path", "") for route in builder_app["app"].routes}
    assert "/api/report-builder/templates" in route_paths
    assert "/api/report-builder/branding" in route_paths
    assert "/api/report-builder/preview" in route_paths
    assert "/api/report-builder/export/pdf" in route_paths
    assert "/api/report-builder/export/excel" in route_paths

    response = client.get("/api/report-builder/templates")
    assert response.status_code == 200
    templates = response.json()
    assert len(templates) >= 3

    db = builder_app["db_module"].SessionLocal()
    try:
        count_before = db.query(builder_app["ReportTemplate"]).count()
    finally:
        db.close()

    builder_app["db_module"].init_db()
    db = builder_app["db_module"].SessionLocal()
    try:
        count_after = db.query(builder_app["ReportTemplate"]).count()
    finally:
        db.close()

    assert count_before == count_after


def test_report_builder_validation_and_crud(builder_app):
    client = TestClient(builder_app["app"])
    db = builder_app["db_module"].SessionLocal()
    try:
        default_template = db.query(builder_app["ReportTemplate"]).filter(builder_app["ReportTemplate"].name == "Full Management Review").first()
        assert default_template is not None
        ay = db.query(builder_app["AcademicYear"]).filter(builder_app["AcademicYear"].label == "2025/2026").first()
        assert ay is not None
        jen = db.query(builder_app["Jenjang"]).filter(builder_app["Jenjang"].name == "Primary").first()
        subject = db.query(builder_app["Subject"]).filter(builder_app["Subject"].name == "Math").first()
    finally:
        db.close()

    invalid = client.post(
        "/api/report-builder/templates",
        json={
            "name": "Broken Template",
            "description": "invalid",
            "template_type": "management_summary",
            "output_format": "both",
            "page_order_json": ["executive_summary", "unknown_section"],
            "section_visibility_json": {"executive_summary": True, "unknown_section": True},
            "chart_visibility_json": {"executive_summary": True},
            "excel_sheet_visibility_json": {"README": True},
            "default_filters_json": {},
            "export_options_json": {},
            "is_default": False,
            "is_active": True,
        },
    )
    assert invalid.status_code == 400

    color_invalid = client.post(
        "/api/report-builder/branding",
        json={
            "school_name": "EDELWEISS SCHOOL",
            "foundation_name": "Foundation",
            "report_header_title": "Title",
            "report_subtitle": "Subtitle",
            "primary_color": "blue",
            "secondary_color": "#0F172A",
            "accent_color": "#F97316",
            "logo_path": None,
            "logo_label": "Logo",
            "footer_text": "Footer",
            "prepared_by": "Tester",
            "is_default": False,
        },
    )
    assert color_invalid.status_code == 400

    created = client.post(
        "/api/report-builder/templates",
        json={
            "name": "Phase 20 Custom",
            "description": "Custom template",
            "template_type": "management_summary",
            "output_format": "both",
            "page_order_json": ["executive_summary", "attendance", "metadata"],
            "section_visibility_json": {"executive_summary": True, "attendance": True, "metadata": True},
            "chart_visibility_json": {"attendance": False},
            "excel_sheet_visibility_json": {"README": True, "Config": True, "Attendance_Data": True},
            "default_filters_json": {"academic_year_id": ay.id, "jenjang_id": jen.id, "subject_id": subject.id},
            "export_options_json": {"include_charts": False},
            "is_default": True,
            "is_active": True,
        },
    )
    assert created.status_code == 200
    created_template = created.json()
    assert created_template["is_default"] is True

    list_response = client.get("/api/report-builder/templates")
    default_count = sum(1 for row in list_response.json() if row["template_type"] == "management_summary" and row["output_format"] == "both" and row["is_default"])
    assert default_count == 1

    patch_response = client.patch(
        f"/api/report-builder/templates/{created_template['id']}",
        json={"description": "Updated custom template", "is_active": True},
    )
    assert patch_response.status_code == 200
    assert patch_response.json()["description"] == "Updated custom template"

    delete_response = client.delete(f"/api/report-builder/templates/{created_template['id']}")
    assert delete_response.status_code == 200
    assert delete_response.json()["deleted"] == 1


def test_report_builder_preview_pdf_excel_and_parity(builder_app):
    client = TestClient(builder_app["app"])
    db = builder_app["db_module"].SessionLocal()
    try:
        template = db.query(builder_app["ReportTemplate"]).filter(builder_app["ReportTemplate"].name == "Full Management Review").first()
        ay = db.query(builder_app["AcademicYear"]).filter(builder_app["AcademicYear"].label == "2025/2026").first()
        assert template is not None and ay is not None
        jen = db.query(builder_app["Jenjang"]).filter(builder_app["Jenjang"].name == "Primary").first()
        subject = db.query(builder_app["Subject"]).filter(builder_app["Subject"].name == "Math").first()
    finally:
        db.close()

    preview = client.post(
        "/api/report-builder/preview",
        json={
            "template_id": template.id,
            "filters": {
                "academic_year_id": ay.id,
                "jenjang_id": jen.id,
                "class_name": "P3A",
                "subject_id": subject.id,
                "term": "term_1",
            },
            "include_trends": False,
            "include_forecast": False,
            "forecast_method": "linear_trend",
            "granularity": "term",
        },
    )
    assert preview.status_code == 200
    preview_data = preview.json()
    assert "executive_summary" in preview_data["resolved_sections"]
    assert preview_data["estimated_pdf_pages"] >= 1
    assert preview_data["selected_template"]["name"] == "Full Management Review"

    pdf_response = client.post(
        "/api/report-builder/export/pdf",
        json={
            "template_id": template.id,
            "filters": {
                "academic_year_id": ay.id,
                "jenjang_id": jen.id,
                "class_name": "P3A",
                "subject_id": subject.id,
                "term": "term_1",
            },
            "include_trends": False,
            "include_forecast": False,
            "forecast_method": "linear_trend",
            "granularity": "term",
            "mode": "editable",
        },
    )
    assert pdf_response.status_code == 200
    assert b"Executive Summary" in pdf_response.content

    excel_response = client.post(
        "/api/report-builder/export/excel",
        json={
            "template_id": template.id,
            "filters": {
                "academic_year_id": ay.id,
                "jenjang_id": jen.id,
                "class_name": "P3A",
                "subject_id": subject.id,
                "term": "term_1",
            },
            "include_trends": False,
            "include_forecast": False,
            "forecast_method": "linear_trend",
            "granularity": "term",
            "mode": "editable",
        },
    )
    assert excel_response.status_code == 200
    workbook = load_workbook(io.BytesIO(excel_response.content), read_only=False)
    assert "README" in workbook.sheetnames
    assert "Config" in workbook.sheetnames
    assert "Attendance_Data" in workbook.sheetnames

    custom_template = client.post(
        "/api/report-builder/templates",
        json={
            "name": "Metadata Only",
            "description": "Minimal template",
            "template_type": "management_summary",
            "output_format": "both",
            "page_order_json": ["executive_summary", "metadata"],
            "section_visibility_json": {"executive_summary": True, "metadata": True},
            "chart_visibility_json": {},
            "excel_sheet_visibility_json": {"README": True, "Config": True},
            "default_filters_json": {"academic_year_id": ay.id},
            "export_options_json": {},
            "is_default": False,
            "is_active": True,
        },
    ).json()

    minimal_pdf = client.post(
        "/api/report-builder/export/pdf",
        json={
            "template_id": custom_template["id"],
            "filters": {
                "academic_year_id": ay.id,
                "jenjang_id": jen.id,
                "class_name": "P3A",
                "subject_id": subject.id,
                "term": "term_1",
            },
            "include_trends": False,
            "include_forecast": False,
            "forecast_method": "linear_trend",
            "granularity": "term",
            "mode": "editable",
        },
    )
    assert minimal_pdf.status_code == 200
    assert b"Attendance Summary" not in minimal_pdf.content
