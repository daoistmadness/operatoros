from datetime import date, time
from io import BytesIO
import re
import sys
from pathlib import Path

SOURCE_ROOT = Path(__file__).resolve().parents[1] / "src"
if str(SOURCE_ROOT) not in sys.path:
    sys.path.insert(0, str(SOURCE_ROOT))

import pytest
from openpyxl import load_workbook
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from core.database import Base, get_db
from models.absence_reason import AbsenceReason
from models.academic_config import KkmThreshold
from models.academic_year import AcademicYear
from models.assessment_component import AssessmentComponent
from models.attendance import Attendance
from models.attendance_review import AttendanceOverride, AttendanceOverrideHistory
from models.jenjang import Jenjang
from models.student import Student
from models.student_enrollment import StudentEnrollment
from models.student_master import StudentAddress, StudentMaster
from models.student_subject_grade import StudentSubjectGrade
from models.subject import Subject
from api.reports import router
from security.dependencies import get_current_user
from services.report_grouping import canonical_scope_for_level, level_matches_scope


@pytest.fixture
def report_context():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    db = Session()

    year = AcademicYear(
        label="2023/2024", start_date=date(2023, 7, 1), end_date=date(2024, 6, 30), status="active", is_default=True
    )
    levels = {name: Jenjang(name=name) for name in ("Early Year Program", "Primary", "Secondary", "PKBM", "Mystery")}
    db.add(year)
    db.add_all(levels.values())
    db.flush()
    students = {}
    for index, (name, level, class_name) in enumerate((
        ("Early", "Early Year Program", "EY A"),
        ("Primary A", "Primary", "P1 A"),
        ("Primary B", "Primary", "P1 B"),
        ("Secondary", "Secondary", "S1 A"),
        ("PKBM Student", "PKBM", "PKBM A"),
        ("Unknown Student", "Mystery", "M A"),
    ), start=1):
        student = Student(id=1000 + index, name=name, jenjang="legacy-wrong", class_name="legacy-wrong")
        db.add(student)
        db.flush()
        enrollment = StudentEnrollment(
            student_id=student.id,
            academic_year_id=year.id,
            jenjang_id=levels[level].id,
            class_name=class_name,
            class_assigned=True,
        )
        db.add(enrollment)
        db.flush()
        students[name] = (student, enrollment)

    math = Subject(name="Math", jenjang_id=levels["Primary"].id)
    science = Subject(name="Science", jenjang_id=levels["Primary"].id)
    db.add_all([math, science])
    db.flush()
    sum_component = AssessmentComponent(name="Exam", assessment_type="sumatif", subject_id=math.id)
    format_component = AssessmentComponent(name="Quiz", assessment_type="formatif", subject_id=math.id)
    science_component = AssessmentComponent(name="Project", assessment_type="sumatif", subject_id=science.id)
    db.add_all([sum_component, format_component, science_component])
    db.flush()
    db.add_all([
        StudentSubjectGrade(enrollment_id=students["Primary A"][1].id, subject_id=math.id, component_id=sum_component.id, score=60),
        StudentSubjectGrade(enrollment_id=students["Primary A"][1].id, subject_id=math.id, component_id=format_component.id, score=None),
        StudentSubjectGrade(enrollment_id=students["Primary B"][1].id, subject_id=math.id, component_id=sum_component.id, score=100),
        StudentSubjectGrade(enrollment_id=students["Primary B"][1].id, subject_id=science.id, component_id=science_component.id, score=90),
    ])
    db.add(KkmThreshold(
        academic_year_id=year.id, jenjang_id=levels["Primary"].id, subject_id=math.id, assessment_type="sumatif", threshold=75
    ))

    def attendance(student_name, day, status, late_duration=0):
        row = Attendance(
            student_id=students[student_name][0].id,
            date=date(2024, 2, day),
            check_in=time(7, 0),
            check_out=time(14, 0),
            late_duration=late_duration,
            status=status,
        )
        db.add(row)
        db.flush()
        return row

    attendance("Early", 1, "on-time")
    attendance("Primary A", 1, "on-time")
    overridden = attendance("Primary A", 2, "incomplete")
    db.add(AttendanceOverride(
        attendance_id=overridden.id,
        original_status="incomplete",
        override_status="late",
        note="Approved late arrival",
        reviewed_by="Admin",
    ))
    attendance("Primary B", 1, "incomplete")
    attendance("Primary B", 2, "absent")
    attendance("Secondary", 1, "late", 30)
    attendance("Secondary", 2, "on-time")
    attendance("PKBM Student", 1, "late", 999)
    db.add(AbsenceReason(
        student_id=students["Primary B"][0].id,
        class_name="P1 B", month=2, year=2024, sakit=1, izin=1, alfa=1, entered_by="Admin"
    ))
    db.commit()

    app = FastAPI()
    app.include_router(router, prefix="/api/reports")

    def override_db():
        yield db

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = lambda: type("MockUser", (), {"role": "admin"})()
    yield {"client": TestClient(app), "db": db, "year": year, "levels": levels, "students": students, "math": math}
    db.close()
    Base.metadata.drop_all(engine)


def monthly(client, year_id, **params):
    query = {"academic_year_id": year_id, "month": "2024-02", "scope": "combined", **params}
    return client.get("/api/reports/monthly", params=query)


def test_filters_academic_year_month_default_and_order(report_context):
    response = report_context["client"].get("/api/reports/filters")
    assert response.status_code == 200
    data = response.json()
    assert data["academic_years"][0]["name"] == "2023/2024"
    assert data["default_academic_year_id"] == report_context["year"].id
    assert data["months"][0]["value"] == "2023-07"
    assert data["months"][-1]["value"] == "2024-06"
    assert [row["value"] for row in data["scopes"]] == ["combined", "early_year", "primary", "secondary"]
    assert "PKBM A" not in data["classes"]


def test_filters_missing_academic_year_returns_404(report_context):
    assert report_context["client"].get("/api/reports/filters", params={"academic_year_id": 999}).status_code == 404


@pytest.mark.parametrize("month", ["2024-2", "24-02", "2024/02", "hello", "2024-13"])
def test_malformed_month_returns_422(report_context, month):
    response = monthly(report_context["client"], report_context["year"].id, month=month)
    assert response.status_code == 422


@pytest.mark.parametrize("month", ["2023-06", "2024-07"])
def test_out_of_range_month_returns_422(report_context, month):
    assert monthly(report_context["client"], report_context["year"].id, month=month).status_code == 422


def test_missing_year_returns_404(report_context):
    assert monthly(report_context["client"], 999).status_code == 404


def test_leap_year_period_and_utc_timestamp(report_context):
    data = monthly(report_context["client"], report_context["year"].id).json()
    assert data["meta"]["period"] == {"start": "2024-02-01", "end": "2024-02-29"}
    assert data["meta"]["generated_at"].endswith("Z")
    assert data["trends"] == []


@pytest.mark.parametrize(
    ("scope", "expected"),
    [
        ("combined", {"Early Year Program", "Primary", "Secondary"}),
        ("early_year", {"Early Year Program"}),
        ("primary", {"Primary"}),
        ("secondary", {"Secondary"}),
    ],
)
def test_scope_membership_and_pkbm_exclusion(report_context, scope, expected):
    data = monthly(report_context["client"], report_context["year"].id, scope=scope).json()
    assert {row["name"] for row in data["student_distribution"]["by_level"]} == expected
    assert "Mystery" in data["data_quality"]["unmapped_levels"]
    assert "PKBM" in data["data_quality"]["unmapped_levels"]


@pytest.mark.parametrize(
    ("value", "scope"),
    [(" kb ", "early_year"), ("Kindergarten", "early_year"), ("SD", "primary"), (" smp ", "secondary")],
)
def test_grouping_aliases_are_normalized(value, scope):
    assert canonical_scope_for_level(value) == scope
    assert level_matches_scope(value, scope)


def test_weighted_attendance_override_incomplete_and_lateness(report_context):
    data = monthly(report_context["client"], report_context["year"].id).json()
    summary = data["attendance_summary"]
    assert summary == {
        "present": 5, "sakit": 1, "izin": 1, "alfa": 1, "incomplete": 1,
        "late_days": 2, "late_minutes": 30, "attendance_rate": 62.5, "late_rate": 40.0,
    }
    assert data["executive_summary"]["data_completeness_rate"] == 88.9
    assert data["data_quality"]["incomplete_attendance"] == 1
    assert any("effective absent" in warning for warning in data["data_quality"]["warnings"])


def test_zero_denominators_return_null(report_context):
    response = monthly(report_context["client"], report_context["year"].id, class_name="EY A", month="2024-03")
    data = response.json()
    assert data["attendance_summary"]["attendance_rate"] is None
    assert data["attendance_summary"]["late_rate"] is None
    assert data["executive_summary"]["data_completeness_rate"] is None


def test_population_is_distinct_enrollment_scoped_and_demographics_stable(report_context):
    data = monthly(report_context["client"], report_context["year"].id).json()
    assert data["executive_summary"]["total_students"] == 4
    assert data["student_distribution"]["by_gender"] == []
    assert data["student_distribution"]["by_religion"] == []
    assert data["student_distribution"]["by_domicile"] == []
    assert data["data_quality"]["missing_gender"] == 4
    assert any("within-year" in warning for warning in data["data_quality"]["warnings"])


def test_academic_raw_average_null_cells_and_below_kkm_unit(report_context):
    data = monthly(report_context["client"], report_context["year"].id, scope="primary").json()
    academic = data["academic_summary"]
    assert academic["sumatif_average"] == pytest.approx(83.3)
    assert academic["formatif_average"] is None
    assert academic["below_kkm_count"] == 1
    assert data["data_quality"]["empty_grade_cells"] == 1
    # Counting unit matches Management Analytics: student + subject + assessment type.
    assert academic["by_subject"][0]["below_kkm_count"] == 1


def test_subject_filter_only_changes_academics(report_context):
    unfiltered = monthly(report_context["client"], report_context["year"].id, scope="primary").json()
    filtered = monthly(
        report_context["client"], report_context["year"].id, scope="primary", subject_id=report_context["math"].id
    ).json()
    assert filtered["executive_summary"]["total_students"] == unfiltered["executive_summary"]["total_students"]
    assert filtered["attendance_summary"] == unfiltered["attendance_summary"]
    assert filtered["academic_summary"]["sumatif_average"] == 80.0


def test_class_filter_affects_all_sections(report_context):
    data = monthly(report_context["client"], report_context["year"].id, scope="primary", class_name=" P1 A ").json()
    assert data["executive_summary"]["total_students"] == 1
    assert data["attendance_summary"]["present"] == 2
    assert data["academic_summary"]["sumatif_average"] == 60.0
    assert data["student_distribution"]["by_class"][0]["name"] == "P1 A"


def test_empty_academic_contract(report_context):
    data = monthly(report_context["client"], report_context["year"].id, scope="secondary").json()
    assert data["academic_summary"] == {
        "availability": False,
        "reason": "Academic data is not available for the selected report context.",
        "sumatif_average": None,
        "formatif_average": None,
        "below_kkm_count": 0,
        "by_subject": [],
    }


def test_invalid_scope_and_subject_are_rejected(report_context):
    assert monthly(report_context["client"], report_context["year"].id, scope="all").status_code == 422
    assert monthly(report_context["client"], report_context["year"].id, subject_id=999).status_code == 404


def test_database_constraints_protect_duplicate_rows(report_context):
    enrollment = report_context["students"]["Primary A"][1]
    duplicate = StudentEnrollment(
        student_id=enrollment.student_id,
        academic_year_id=enrollment.academic_year_id,
        jenjang_id=enrollment.jenjang_id,
    )
    report_context["db"].add(duplicate)
    with pytest.raises(Exception):
        report_context["db"].flush()
    report_context["db"].rollback()


def test_monthly_time_basis_is_backend_owned(report_context):
    data = monthly(report_context["client"], report_context["year"].id).json()
    assert data["report_period"]["sections"] == {
        "attendance": {"basis": "calendar_month", "month_bound": True, "label": "February 2024"},
        "population": {"basis": "academic_year_enrollment_snapshot", "month_bound": False, "label": "Academic Year 2023/2024"},
        "academics": {"basis": "available_academic_year_records", "month_bound": False, "label": "Available Academic Records - AY 2023/2024"},
    }


def test_management_demographics_reconcile_without_forcing_totals(report_context):
    db = report_context["db"]
    primary_a = report_context["students"]["Primary A"][1]
    primary_b = report_context["students"]["Primary B"][1]
    first = StudentMaster(full_name="Synthetic One", normalized_name="synthetic one", student_status="active", gender="male", religion="Islam")
    second = StudentMaster(full_name="Synthetic Two", normalized_name="synthetic two", student_status="active", gender="female", religion=None)
    db.add_all([first, second]); db.flush()
    primary_a.student_master_id = first.id; primary_b.student_master_id = second.id
    db.add(StudentAddress(student_master_id=first.id, kelurahan="  jatibening BARU "))
    db.commit()
    response = report_context["client"].get("/api/reports/management/monthly", params={
        "academic_year_id": report_context["year"].id, "month": "2024-02", "scope": "primary",
    })
    assert response.status_code == 200
    data = response.json()
    assert data["metadata"]["report_type"] == "monthly_management"
    assert data["data_quality"]["reconciliation"] == {
        "population_total": 2, "student_master_linked": 2, "student_master_unlinked": 0,
        "religion_known": 1, "religion_unknown": 1, "gender_known": 2, "gender_unknown": 0,
        "location_known": 1, "location_unknown": 1,
    }
    assert data["demographics"]["religion"]["rows"][0]["percentage_of_known"] == 100.0
    assert data["demographics"]["religion"]["rows"][0]["percentage_of_eligible"] == 50.0
    assert data["demographics"]["residential_area"]["rows"][0]["name"] == "Jatibening Baru"
    assert all(row["reconciles"] for row in data["data_quality"]["sections"].values())


@pytest.mark.parametrize(("format", "mime"), [
    ("pdf", "application/pdf"),
    ("xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
])
def test_management_export_contract(report_context, format, mime):
    response = report_context["client"].get("/api/reports/management/monthly/export", params={
        "academic_year_id": report_context["year"].id, "month": "2024-02", "scope": "primary", "format": format,
    })
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(mime)
    assert f"management-report_monthly_primary_2024-02.{format}" in response.headers["content-disposition"]
    assert len(response.content) > 1000
    if format == "pdf":
        assert len(re.findall(rb"/Type\s*/Page\b", response.content)) >= 2
        assert b"Data Quality & Coverage" in response.content
        assert b"not restricted to the selected calendar month" in response.content
    else:
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        assert workbook.sheetnames == [
            "Executive Summary", "Population by Jenjang", "Population by Class", "Religion", "Gender",
            "Residential Area", "Monthly Attendance", "Academic Summary", "Data Quality",
        ]
