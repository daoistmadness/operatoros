from io import BytesIO
import sys
from pathlib import Path

SOURCE_ROOT = Path(__file__).resolve().parents[1] / "src"
if str(SOURCE_ROOT) not in sys.path:
    sys.path.insert(0, str(SOURCE_ROOT))

import pytest
from openpyxl import load_workbook

from test_reports import report_context


PDF_MIME = "application/pdf"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def export(context, report_type, export_format, **params):
    query = {
        "academic_year_id": context["year"].id,
        "scope": "combined",
        "format": export_format,
        **params,
    }
    if report_type == "monthly":
        query.setdefault("month", "2024-02")
    return context["client"].get(f"/api/reports/{report_type}/export", params=query)


@pytest.mark.parametrize(
    ("report_type", "export_format", "mime", "filename"),
    [
        ("monthly", "pdf", PDF_MIME, "executive-report_monthly_combined_2024-02.pdf"),
        ("annual", "pdf", PDF_MIME, "executive-report_annual_combined_2023-2024.pdf"),
        ("monthly", "xlsx", XLSX_MIME, "executive-report_monthly_combined_2024-02.xlsx"),
        ("annual", "xlsx", XLSX_MIME, "executive-report_annual_combined_2023-2024.xlsx"),
    ],
)
def test_export_endpoint_mime_filename_and_nonempty_content(report_context, report_type, export_format, mime, filename):
    response = export(report_context, report_type, export_format)
    assert response.status_code == 200
    assert response.headers["content-type"] == mime
    assert response.headers["content-disposition"] == f'attachment; filename="{filename}"'
    assert len(response.content) > 1000


@pytest.mark.parametrize("report_type", ["monthly", "annual"])
def test_pdf_integrity_metadata_and_json_parity(report_context, report_type):
    json_params = {"academic_year_id": report_context["year"].id, "scope": "combined"}
    if report_type == "monthly":
        json_params["month"] = "2024-02"
    report = report_context["client"].get(f"/api/reports/{report_type}", params=json_params).json()
    response = export(report_context, report_type, "pdf")
    content = response.content
    assert content.startswith(b"%PDF-")
    assert content.rstrip().endswith(b"%%EOF")
    assert b"Executive Report" in content
    assert report["meta"]["academic_year"]["name"].encode() in content
    assert str(report["executive_summary"]["total_students"]).encode() in content
    attendance_text = f"{report['executive_summary']['attendance_rate']:.1f}%".encode()
    assert attendance_text in content
    if report_type == "annual":
        assert b"Monthly Trends" in content
        assert b"Annual Comparisons" in content


@pytest.mark.parametrize(
    ("report_type", "expected_sheets"),
    [
        ("monthly", ["Executive Summary", "Attendance", "Student Distribution", "Academic Summary", "Data Quality"]),
        ("annual", ["Executive Summary", "Attendance", "Student Distribution", "Academic Summary", "Annual Trends", "Data Quality"]),
    ],
)
def test_xlsx_integrity_sheets_and_json_parity(report_context, report_type, expected_sheets):
    json_params = {"academic_year_id": report_context["year"].id, "scope": "combined"}
    if report_type == "monthly":
        json_params["month"] = "2024-02"
    report = report_context["client"].get(f"/api/reports/{report_type}", params=json_params).json()
    response = export(report_context, report_type, "xlsx")
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == expected_sheets
    summary_values = {row[0]: row[1] for row in workbook["Executive Summary"].iter_rows(min_row=3, values_only=True)}
    assert summary_values["Total Students"] == report["executive_summary"]["total_students"]
    assert summary_values["Attendance Rate"] == report["executive_summary"]["attendance_rate"]
    attendance = workbook["Attendance"]
    assert attendance["B3"].value == report["attendance_summary"]["present"]
    assert attendance["I3"].value == report["attendance_summary"]["attendance_rate"]
    assert attendance["I3"].number_format == '0.0"%"'
    assert attendance.freeze_panes == "A3"
    if report_type == "annual":
        trends = workbook["Annual Trends"]
        exported = list(trends.iter_rows(min_row=3, values_only=True))
        assert len(exported) == len(report["trends"])
        assert [row[0] for row in exported] == [row["month"] for row in report["trends"]]
        assert [row[8] for row in exported] == [row["attendance_rate"] for row in report["trends"]]


def test_combined_weighted_values_match_in_both_formats(report_context):
    report = report_context["client"].get("/api/reports/monthly", params={
        "academic_year_id": report_context["year"].id, "month": "2024-02", "scope": "combined"
    }).json()
    assert report["attendance_summary"]["attendance_rate"] == 62.5
    pdf = export(report_context, "monthly", "pdf").content
    assert b"62.5%" in pdf
    workbook = load_workbook(BytesIO(export(report_context, "monthly", "xlsx").content), data_only=True)
    assert workbook["Attendance"]["I3"].value == 62.5


@pytest.mark.parametrize(
    ("report_type", "params", "status"),
    [
        ("monthly", {"format": "csv", "month": "2024-02"}, 422),
        ("annual", {"format": "csv"}, 422),
        ("annual", {"academic_year_id": 999999}, 404),
        ("monthly", {"month": "bad"}, 422),
        ("monthly", {"month": "2024-07"}, 422),
    ],
)
def test_export_errors(report_context, report_type, params, status):
    export_format = params.pop("format", "pdf")
    response = export(report_context, report_type, export_format, **params)
    assert response.status_code == status
