# test_report_parity.py
# Phase 17: Report Accuracy QA - Management Review Parity.
# Tech Stack: FastAPI / SQLAlchemy / pytest / openpyxl / pandas
from __future__ import annotations
import importlib, io, sys
from datetime import date, time
from pathlib import Path
import pytest

MODULE_PREFIXES = ("src", "api", "core", "models", "services")
SOURCE_ROOT = Path(__file__).resolve().parents[1] / "src"


def unload_app_modules() -> None:
    for name in list(sys.modules):
        if name == "src" or name.startswith(MODULE_PREFIXES):
            sys.modules.pop(name, None)


@pytest.fixture
def golden_db(monkeypatch, tmp_path):
    """Golden fixture: 4 students, 2 classes, 2 subjects, custom KKM+term config."""
    db_path = tmp_path / "golden.db"
    monkeypatch.syspath_prepend(str(SOURCE_ROOT))
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path}")
    unload_app_modules()
    db_module = importlib.import_module("core.database")
    db_module.init_db()
    AcademicYear = importlib.import_module("models.academic_year").AcademicYear
    Jenjang = importlib.import_module("models.jenjang").Jenjang
    Student = importlib.import_module("models.student").Student
    StudentEnrollment = importlib.import_module("models.student_enrollment").StudentEnrollment
    Attendance = importlib.import_module("models.attendance").Attendance
    AbsenceReason = importlib.import_module("models.absence_reason").AbsenceReason
    Subject = importlib.import_module("models.subject").Subject
    AssessmentComponent = importlib.import_module("models.assessment_component").AssessmentComponent
    StudentSubjectGrade = importlib.import_module("models.student_subject_grade").StudentSubjectGrade
    KkmThreshold = importlib.import_module("models.academic_config").KkmThreshold
    AcademicTermConfig = importlib.import_module("models.academic_config").AcademicTermConfig
    AcademicIntervention = importlib.import_module("models.academic_intervention").AcademicIntervention
    db = db_module.SessionLocal()
    ay = db.query(AcademicYear).filter(AcademicYear.label == "2025/2026").first()
    if ay is None:
        ay = AcademicYear(label="2025/2026", start_date=date(2025, 7, 1), end_date=date(2026, 6, 30))
        db.add(ay)
        db.flush()
    else:
        ay.start_date = date(2025, 7, 1)
        ay.end_date = date(2026, 6, 30)
        db.flush()
    jen = db.query(Jenjang).filter(Jenjang.name == "Primary").first()
    if jen is None:
        jen = Jenjang(name="Primary")
        db.add(jen)
        db.flush()

    db.add_all([
        AcademicTermConfig(academic_year_id=ay.id, term_number=1, label="Term 1", start_date=date(2025, 7, 1), end_date=date(2025, 9, 30)),
        AcademicTermConfig(academic_year_id=ay.id, term_number=2, label="Term 2", start_date=date(2025, 10, 1), end_date=date(2025, 12, 31)),
        AcademicTermConfig(academic_year_id=ay.id, term_number=3, label="Term 3", start_date=date(2026, 1, 1), end_date=date(2026, 3, 31)),
        AcademicTermConfig(academic_year_id=ay.id, term_number=4, label="Term 4", start_date=date(2026, 4, 1), end_date=date(2026, 6, 30)),
    ])
    db.flush()

    math = Subject(name="Math", jenjang_id=jen.id, supports_sumatif=True, supports_formatif=True)
    science = Subject(name="Science", jenjang_id=jen.id, supports_sumatif=True, supports_formatif=True)
    db.add_all([math, science]); db.flush()
    db.add(KkmThreshold(academic_year_id=ay.id, jenjang_id=jen.id, subject_id=math.id, assessment_type="sumatif", threshold=80.0)); db.flush()
    math_sum = AssessmentComponent(name="Math Sumatif", assessment_type="sumatif", subject_id=math.id)
    math_for = AssessmentComponent(name="Math Formatif", assessment_type="formatif", subject_id=math.id)
    sci_sum = AssessmentComponent(name="Science Sumatif", assessment_type="sumatif", subject_id=science.id)
    sci_for = AssessmentComponent(name="Science Formatif", assessment_type="formatif", subject_id=science.id)
    db.add_all([math_sum, math_for, sci_sum, sci_for]); db.flush()
    s_alice = Student(name="Alice", jenjang="Primary", class_name="P1A")
    s_bob = Student(name="Bob", jenjang="Primary", class_name="P1A")
    s_carol = Student(name="Carol", jenjang="Primary", class_name="P1B")
    s_dave = Student(name="Dave", jenjang="Primary", class_name="P1B")
    db.add_all([s_alice, s_bob, s_carol, s_dave]); db.flush()
    en_alice = StudentEnrollment(student_id=s_alice.id, academic_year_id=ay.id, jenjang_id=jen.id, class_name="P1A", class_assigned=True)
    en_bob = StudentEnrollment(student_id=s_bob.id, academic_year_id=ay.id, jenjang_id=jen.id, class_name="P1A", class_assigned=True)
    en_carol = StudentEnrollment(student_id=s_carol.id, academic_year_id=ay.id, jenjang_id=jen.id, class_name="P1B", class_assigned=True)
    en_dave = StudentEnrollment(student_id=s_dave.id, academic_year_id=ay.id, jenjang_id=jen.id, class_name="P1B", class_assigned=True)
    db.add_all([en_alice, en_bob, en_carol, en_dave]); db.flush()
    db.add_all([
        Attendance(student_id=s_alice.id, date=date(2025, 10, 1), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0),
        Attendance(student_id=s_alice.id, date=date(2025, 10, 2), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0),
        Attendance(student_id=s_alice.id, date=date(2025, 10, 3), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0),
        Attendance(student_id=s_bob.id, date=date(2025, 10, 1), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0),
        Attendance(student_id=s_bob.id, date=date(2025, 10, 2), check_in=time(8, 32), check_out=time(14, 0), status="late", late_duration=92),
        Attendance(student_id=s_bob.id, date=date(2025, 10, 3), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0),
        Attendance(student_id=s_carol.id, date=date(2025, 10, 1), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0),
        Attendance(student_id=s_carol.id, date=date(2025, 10, 2), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0),
        Attendance(student_id=s_carol.id, date=date(2025, 10, 3), check_in=time(7, 28), check_out=time(14, 0), status="late", late_duration=28),
        Attendance(student_id=s_dave.id, date=date(2025, 10, 1), check_in=time(7, 0), check_out=time(14, 0), status="on-time", late_duration=0),
        Attendance(student_id=s_dave.id, date=date(2025, 10, 2), check_in=time(8, 0), check_out=time(14, 0), status="late", late_duration=60),
    ])
    db.add_all([
        AbsenceReason(student_id=s_alice.id, class_name="P1A", month=10, year=2025, sakit=1, izin=0, alfa=0, entered_by="operator"),
        AbsenceReason(student_id=s_bob.id, class_name="P1A", month=10, year=2025, sakit=0, izin=1, alfa=0, entered_by="operator"),
        AbsenceReason(student_id=s_dave.id, class_name="P1B", month=10, year=2025, sakit=0, izin=0, alfa=1, entered_by="operator"),
    ])
    db.add_all([
        StudentSubjectGrade(enrollment_id=en_alice.id, subject_id=math.id, component_id=math_sum.id, score=85.0),
        StudentSubjectGrade(enrollment_id=en_alice.id, subject_id=math.id, component_id=math_for.id, score=None),
        StudentSubjectGrade(enrollment_id=en_alice.id, subject_id=science.id, component_id=sci_sum.id, score=90.0),
        StudentSubjectGrade(enrollment_id=en_alice.id, subject_id=science.id, component_id=sci_for.id, score=88.0),
        StudentSubjectGrade(enrollment_id=en_bob.id, subject_id=math.id, component_id=math_sum.id, score=75.0),
        StudentSubjectGrade(enrollment_id=en_bob.id, subject_id=math.id, component_id=math_for.id, score=70.0),
        StudentSubjectGrade(enrollment_id=en_bob.id, subject_id=science.id, component_id=sci_sum.id, score=None),
        StudentSubjectGrade(enrollment_id=en_bob.id, subject_id=science.id, component_id=sci_for.id, score=65.0),
        StudentSubjectGrade(enrollment_id=en_carol.id, subject_id=math.id, component_id=math_sum.id, score=92.0),
        StudentSubjectGrade(enrollment_id=en_carol.id, subject_id=math.id, component_id=math_for.id, score=95.0),
        StudentSubjectGrade(enrollment_id=en_carol.id, subject_id=science.id, component_id=sci_sum.id, score=78.0),
        StudentSubjectGrade(enrollment_id=en_carol.id, subject_id=science.id, component_id=sci_for.id, score=80.0),
        StudentSubjectGrade(enrollment_id=en_dave.id, subject_id=math.id, component_id=math_sum.id, score=60.0),
        StudentSubjectGrade(enrollment_id=en_dave.id, subject_id=math.id, component_id=math_for.id, score=55.0),
        StudentSubjectGrade(enrollment_id=en_dave.id, subject_id=science.id, component_id=sci_sum.id, score=72.0),
        StudentSubjectGrade(enrollment_id=en_dave.id, subject_id=science.id, component_id=sci_for.id, score=68.0),
    ])
    db.add(AcademicIntervention(
        student_id=s_bob.id, academic_year_id=ay.id, jenjang_id=jen.id, subject_id=math.id,
        class_name="P1A", student_name="Bob", subject_name="Math", assessment_type="sumatif",
        effective_threshold=80.0, threshold_source="kkm_configured",
        status="open", priority="high", owner_name="Wali Kelas P1A", follow_up_date=date(2026, 3, 15),
    ))
    db.commit()
    yield {"db": db, "ay_id": ay.id, "jen_id": jen.id}
    db.close()


def _build():
    return importlib.import_module("services.management_analytics").build_management_summary


def _exports():
    mod = importlib.import_module("services.management_report_export")
    return mod.build_management_summary_pdf, mod.build_management_summary_excel


def _data(golden_db):
    return _build()(golden_db["db"], academic_year_id=golden_db["ay_id"], jenjang_id=golden_db["jen_id"])


def _lates(golden_db):
    return {r["class_name"]: r for r in _data(golden_db)["lateness_by_class"]}


def _excel(golden_db):
    _, build_excel = _exports()
    s = _data(golden_db)
    return s, build_excel(s, {"mode": "editable"})


def _pdf(golden_db):
    build_pdf, _ = _exports()
    s = _data(golden_db)
    return build_pdf(s), s


# 1. ATTENDANCE

def test_attendance_hadir_count(golden_db):
    assert _data(golden_db)["attendance_summary"]["status_counts"]["hadir"] == 11

def test_attendance_absence_counts(golden_db):
    att = _data(golden_db)["attendance_summary"]
    assert att["status_counts"]["sakit"] == 1
    assert att["status_counts"]["izin"] == 1
    assert att["status_counts"]["alfa"] == 1

def test_attendance_total_records(golden_db):
    assert _data(golden_db)["attendance_summary"]["total_records"] == 14

def test_attendance_hadir_percentage(golden_db):
    pct = _data(golden_db)["attendance_summary"]["status_percentages"]["hadir"]
    assert pct == pytest.approx(round(11/14*100, 1), abs=0.2)

def test_attendance_percentage_sums_100(golden_db):
    total = sum(_data(golden_db)["attendance_summary"]["status_percentages"].values())
    assert total == pytest.approx(100.0, abs=0.5)


# 2. LATENESS

def test_lateness_p1a(golden_db):
    l = _lates(golden_db)
    assert l["P1A"]["late_days"] == 1 and l["P1A"]["late_minutes"] == 92

def test_lateness_p1b(golden_db):
    l = _lates(golden_db)
    assert l["P1B"]["late_days"] == 2 and l["P1B"]["late_minutes"] == 88

def test_lateness_label(golden_db):
    l = _lates(golden_db)
    assert l["P1A"]["late_duration_label"] == "1:32"
    assert l["P1B"]["late_duration_label"] == "1:28"

def test_lateness_pct_sums_100(golden_db):
    total = sum(r["late_day_percentage"] for r in _data(golden_db)["lateness_by_class"])
    assert total == pytest.approx(100.0, abs=0.5)

def test_no_unknown_class(golden_db):
    assert "Unknown" not in _lates(golden_db)

def test_total_late_days_3(golden_db):
    assert sum(r["late_days"] for r in _data(golden_db)["lateness_by_class"]) == 3


# 3. GRADES

def test_grade_class_p1a_sumatif(golden_db):
    p1a = next(g for g in _data(golden_db)["grade_by_class"] if g["class_name"] == "P1A")
    assert p1a["sumatif_average"] == pytest.approx(83.3, abs=0.2)

def test_grade_class_p1a_formatif_null_excluded(golden_db):
    p1a = next(g for g in _data(golden_db)["grade_by_class"] if g["class_name"] == "P1A")
    assert p1a["formatif_average"] == pytest.approx(74.3, abs=0.2)

def test_grade_class_p1b_sumatif(golden_db):
    p1b = next(g for g in _data(golden_db)["grade_by_class"] if g["class_name"] == "P1B")
    assert p1b["sumatif_average"] == pytest.approx(75.5, abs=0.2)


def test_grade_subject_math(golden_db):
    math = next(g for g in _data(golden_db)["grade_by_subject"] if g["subject_name"] == "Math")
    assert math["sumatif_average"] == pytest.approx(78.0, abs=0.2)


def test_grade_subject_science_null_excluded(golden_db):
    sci = next(g for g in _data(golden_db)["grade_by_subject"] if g["subject_name"] == "Science")
    assert sci["sumatif_average"] == pytest.approx(80.0, abs=0.2)

def test_student_alice_null_formatif_is_none(golden_db):
    alice = next((g for g in _data(golden_db)["grade_by_student"] if g["student_name"] == "Alice" and g["subject_name"] == "Math"), None)
    assert alice is not None and alice["formatif_average"] is None

def test_student_bob_below_kkm(golden_db):
    bob = next((g for g in _data(golden_db)["grade_by_student"] if g["student_name"] == "Bob" and g["subject_name"] == "Math"), None)
    assert bob is not None and bob["below_threshold"] is True

def test_student_alice_not_below_kkm(golden_db):
    alice = next((g for g in _data(golden_db)["grade_by_student"] if g["student_name"] == "Alice" and g["subject_name"] == "Math"), None)
    assert alice is not None and alice["below_threshold"] is False

def test_two_classes(golden_db):
    assert {"P1A", "P1B"}.issubset({g["class_name"] for g in _data(golden_db)["grade_by_class"]})

def test_two_subjects(golden_db):
    assert {"Math", "Science"}.issubset({g["subject_name"] for g in _data(golden_db)["grade_by_subject"]})


# 4. KKM RESOLUTION

def test_custom_kkm_math_sumatif(golden_db):
    alerts = [a for a in _data(golden_db)["below_kkm_alerts"] if a["student_name"] == "Bob" and a["subject_name"] == "Math" and a["assessment_type"] == "sumatif"]
    assert len(alerts) >= 1 and alerts[0]["kkm_threshold"] == pytest.approx(80.0, abs=0.01)

def test_legacy_kkm_science(golden_db):
    alerts = [a for a in _data(golden_db)["below_kkm_alerts"] if a["subject_name"] == "Science"]
    assert any(a["kkm_threshold"] == pytest.approx(85.0, abs=0.01) for a in alerts)

def test_below_kkm_count_gte_3(golden_db):
    assert len(_data(golden_db)["below_kkm_alerts"]) >= 3

def test_kkm_source_present(golden_db):
    for a in _data(golden_db)["below_kkm_alerts"]:
        assert a.get("threshold_source") is not None


# 5. INTERVENTIONS

def test_bob_alert_has_intervention(golden_db):
    alerts = [a for a in _data(golden_db)["below_kkm_alerts"] if a["student_name"] == "Bob" and a["assessment_type"] == "sumatif"]
    assert len(alerts) >= 1
    assert alerts[0].get("intervention_status") == "open"
    assert alerts[0].get("intervention_priority") == "high"

def test_interventions_open(golden_db):
    assert _data(golden_db)["interventions_summary"]["status_counts"]["open"] >= 1

def test_interventions_by_class_p1a(golden_db):
    assert "P1A" in _data(golden_db)["interventions_summary"]["by_class"]


# 6. TERM BREAKDOWN

def test_four_terms(golden_db):
    assert len(_data(golden_db)["terms_breakdown"]) == 4

def test_term2_oct_hadir(golden_db):
    t2 = next(t for t in _data(golden_db)["terms_breakdown"] if t["term_number"] == 2)
    assert t2["hadir"] == 11

def test_term1_no_hadir(golden_db):
    t1 = next(t for t in _data(golden_db)["terms_breakdown"] if t["term_number"] == 1)
    assert t1["hadir"] == 0

def test_term_source_custom(golden_db):
    assert "custom" in {t["source"] for t in _data(golden_db)["terms_breakdown"]}


# 7. INSIGHTS

def test_below_kkm_insight(golden_db):
    assert any("Keterlambatan Akademik" in i["title"] or i["category"] == "below_kkm" for i in _data(golden_db)["executive_insights"])

def test_insights_critical_first(golden_db):
    order = {"critical": 0, "warning": 1, "info": 2}
    vals = [order.get(i["severity"], 3) for i in _data(golden_db)["executive_insights"]]
    assert vals == sorted(vals)

def test_insights_required_fields(golden_db):
    for i in _data(golden_db)["executive_insights"]:
        for k in ("severity", "category", "title", "message", "recommended_action"):
            assert k in i

def test_null_warning(golden_db):
    assert any("null" in w.lower() or "Null" in w for w in _data(golden_db)["warnings"])

def test_full_year_insight(golden_db):
    assert any("Laporan Tahunan Penuh" in i["title"] for i in _data(golden_db)["executive_insights"])


# 8. EXCEL PARITY

def test_excel_sheets(golden_db):
    import pandas as pd
    _, xb = _excel(golden_db)
    xl = pd.ExcelFile(io.BytesIO(xb), engine="openpyxl")
    for s in ["README","Config","Attendance_Data","Lateness_Data","Grade_Class_Data","Grade_Subject_Data","Grade_Student_Data","Below_KKM_Data","Interventions_Data","Insights","Charts"]:
        assert s in xl.sheet_names

def test_excel_attendance_cols(golden_db):
    import pandas as pd
    _, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Attendance_Data", header=0, engine="openpyxl")
    for c in ("Term","Hadir (Hari)","Sakit (Hari)","Izin (Hari)","Alfa (Hari)","Total Kehadiran","Persentase Kehadiran"):
        assert c in df.columns, f"Missing: {c}"

def test_excel_lateness_cols(golden_db):
    import pandas as pd
    _, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Lateness_Data", header=0, engine="openpyxl")
    for c in ("Kelas","Hari Terlambat","Total Menit Terlambat","Persentase Hari"):
        assert c in df.columns

def test_excel_grade_class_kkm_col(golden_db):
    import pandas as pd
    _, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Grade_Class_Data", header=0, engine="openpyxl")
    assert "KKM Edelweiss" in df.columns

def test_excel_grade_class_rows(golden_db):
    import pandas as pd
    s, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Grade_Class_Data", header=0, engine="openpyxl")
    assert len(df) == len(s["grade_by_class"])

def test_excel_below_kkm_cols(golden_db):
    import pandas as pd
    _, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Below_KKM_Data", header=0, engine="openpyxl")
    assert "Status Intervensi" in df.columns and "Prioritas" in df.columns

def test_excel_insights_rows(golden_db):
    import pandas as pd
    _, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Insights", header=0, engine="openpyxl")
    assert len(df) > 0

def test_excel_insights_cols(golden_db):
    import pandas as pd
    _, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Insights", header=0, engine="openpyxl")
    for c in ("Tingkat Keparahan","Kategori","Judul Temuan","Deskripsi Analisis"):
        assert c in df.columns

def test_excel_grade_student_names(golden_db):
    import pandas as pd
    _, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Grade_Student_Data", header=0, engine="openpyxl")
    names = set(df["Nama Siswa"].dropna().tolist())
    for n in ("Alice","Bob","Carol","Dave"):
        assert n in names

def test_excel_lateness_rows(golden_db):
    import pandas as pd
    s, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Lateness_Data", header=0, engine="openpyxl")
    assert len(df) == len(s["lateness_by_class"])

def test_excel_readme_school(golden_db):
    from openpyxl import load_workbook
    _, xb = _excel(golden_db)
    wb = load_workbook(io.BytesIO(xb), read_only=True)
    txt = " ".join(str(c) for row in wb["README"].iter_rows(values_only=True) for c in row if c)
    assert "EDELWEISS" in txt


# 9. PDF PARITY

def test_pdf_valid(golden_db):
    p, _ = _pdf(golden_db); assert p.startswith(b"%PDF")

def test_pdf_school(golden_db):
    p, _ = _pdf(golden_db); assert b"EDELWEISS" in p

def test_pdf_insights(golden_db):
    p, _ = _pdf(golden_db); assert b"EXECUTIVE ANALYTICS INSIGHTS" in p

def test_pdf_context(golden_db):
    p, _ = _pdf(golden_db); assert b"REPORT FILTER" in p

def test_pdf_kkm(golden_db):
    p, _ = _pdf(golden_db); assert b"KKM" in p

def test_pdf_bob(golden_db):
    p, _ = _pdf(golden_db); assert b"Bob" in p

def test_pdf_attendance(golden_db):
    p, _ = _pdf(golden_db); assert b"Attendance" in p or b"ATTENDANCE" in p

def test_pdf_size(golden_db):
    p, _ = _pdf(golden_db); assert len(p) > 5000


# 10. DATA QUALITY

def test_dq_null_warning(golden_db):
    assert any("null" in w.lower() or "Null" in w for w in _data(golden_db)["warnings"])

def test_dq_no_unknown_class_insight(golden_db):
    unknowns = [i for i in _data(golden_db)["executive_insights"] if "Unknown" in i.get("message","") and "class" in i.get("message","").lower()]
    assert len(unknowns) == 0

def test_dq_math_not_legacy(golden_db):
    bob = next((a for a in _data(golden_db)["below_kkm_alerts"] if a["student_name"]=="Bob" and a["subject_name"]=="Math" and a["assessment_type"]=="sumatif"), None)
    assert bob is not None and bob.get("threshold_source") != "legacy-fallback"

def test_dq_data_quality_category(golden_db):
    import pandas as pd
    s, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Insights", header=0, engine="openpyxl")
    cats = df["Kategori"].dropna().tolist() if "Kategori" in df.columns else []
    assert "data_quality" in cats


# 11. CROSS-FORMAT PARITY

def test_parity_hadir(golden_db):
    import pandas as pd
    s, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Attendance_Data", header=0, engine="openpyxl")
    assert int(df["Hadir (Hari)"].sum()) == s["attendance_summary"]["status_counts"]["hadir"]

def test_parity_grade_sumatif(golden_db):
    import pandas as pd
    s, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Grade_Class_Data", header=0, engine="openpyxl")
    m = {r["Kelas"]: r["Rata-rata Sumatif"] for _, r in df.iterrows()}
    for gc in s["grade_by_class"]:
        if gc["sumatif_average"] is not None and gc["class_name"] in m:
            assert abs(float(m[gc["class_name"]]) - gc["sumatif_average"]) < 0.2

def test_parity_below_kkm_rows(golden_db):
    import pandas as pd
    s, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Below_KKM_Data", header=0, engine="openpyxl")
    assert len(df) == len(s["below_kkm_alerts"])

def test_parity_insights_rows(golden_db):
    import pandas as pd
    s, xb = _excel(golden_db)
    df = pd.read_excel(io.BytesIO(xb), sheet_name="Insights", header=0, engine="openpyxl")
    assert len(df) == len(s["executive_insights"])
