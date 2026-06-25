from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COLUMN_ALIASES = {
    "date": ["Date", "Tanggal", "date", "tanggal"],
    "student_id": ["Student_ID", "No. ID", "No ID", "student_id"],
    "name": ["Name", "Nama", "name", "nama"],
    "class": ["Class", "Kelas", "class", "kelas", "class_name"],
    "level": ["Level", "Jenjang", "level", "jenjang"],
    "late_indicator": ["Terlambat", "Late", "late_duration", "late_indicator"],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an executive dashboard for lateness metrics by jenjang or across all jenjang."
    )
    parser.add_argument("input_file", help="Path to the source Excel or CSV file")
    parser.add_argument(
        "--output",
        default="lateness_executive_dashboard.xlsx",
        help="Output workbook path",
    )
    parser.add_argument("--sheet", default=None, help="Excel sheet name to read")
    parser.add_argument(
        "--term-days",
        type=int,
        default=90,
        help="Number of school days in the academic term",
    )
    parser.add_argument(
        "--level-value",
        default="PRIMARY",
        help="Level value to filter on, or use ALL to combine all jenjang",
    )
    parser.add_argument(
        "--default-level",
        default=None,
        help="Fallback level value to assign when the source file has no Level/Jenjang column",
    )
    return parser.parse_args()


def resolve_column(df: pd.DataFrame, logical_name: str) -> str | None:
    for candidate in COLUMN_ALIASES[logical_name]:
        if candidate in df.columns:
            return candidate
    return None


def load_source(path: Path, sheet_name: str | None) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path)
    return pd.read_excel(path, sheet_name=sheet_name)


def is_late_value(value: object) -> bool:
    if pd.isna(value):
        return False
    text = str(value).strip().lower()
    return text not in {"", "0", "0:00", "0:00:00", "false", "none", "nan"}


def build_late_incident_frame(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str | None]]:
    date_col = resolve_column(df, "date")
    student_id_col = resolve_column(df, "student_id")
    name_col = resolve_column(df, "name")
    class_col = resolve_column(df, "class")
    level_col = resolve_column(df, "level")
    late_indicator_col = resolve_column(df, "late_indicator")

    if date_col is None:
        raise ValueError("Missing date column. Expected one of: Date, Tanggal")

    if late_indicator_col:
        incidents = df[df[late_indicator_col].map(is_late_value)].copy()
    else:
        incidents = df.copy()

    incidents[date_col] = pd.to_datetime(incidents[date_col], errors="coerce")
    incidents = incidents.dropna(subset=[date_col])
    incidents[date_col] = incidents[date_col].dt.date

    columns = {
        "date": date_col,
        "student_id": student_id_col,
        "name": name_col,
        "class": class_col,
        "level": level_col,
        "late_indicator": late_indicator_col,
    }
    return incidents, columns


def ensure_level(incidents: pd.DataFrame, level_col: str | None, default_level: str | None) -> tuple[pd.DataFrame, str]:
    if level_col:
        incidents[level_col] = incidents[level_col].astype(str).str.strip()
        return incidents, level_col
    if default_level:
        incidents = incidents.copy()
        incidents["Level"] = default_level
        return incidents, "Level"
    raise ValueError(
        "Missing Level/Jenjang column. Provide a file with level data or use --default-level PRIMARY."
    )


def build_level_summary_df(incidents: pd.DataFrame, date_col: str, level_col: str, term_days: int) -> pd.DataFrame:
    grouped = (
        incidents.groupby(level_col)
        .agg(
            total_late_incidents=(date_col, "size"),
            unique_late_days=(date_col, "nunique"),
        )
        .reset_index()
    )
    grouped["school_impact_rate"] = grouped["unique_late_days"].map(
        lambda value: (value / term_days * 100) if term_days else 0
    )
    grouped["average_lateness_density"] = grouped.apply(
        lambda row: row["total_late_incidents"] / row["unique_late_days"] if row["unique_late_days"] else 0,
        axis=1,
    )
    return grouped.rename(
        columns={
            level_col: "Level",
            "total_late_incidents": "Total Late Incidents",
            "unique_late_days": "Unique Late Days",
            "school_impact_rate": "School Impact Rate (%)",
            "average_lateness_density": "Average Lateness Density",
        }
    ).sort_values(["Total Late Incidents", "Unique Late Days"], ascending=[False, False])


def autosize_columns(ws, widths: dict[int, int] | None = None) -> None:
    widths = widths or {}
    for idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(idx, 20)


def apply_table_style(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border


def build_dashboard(
    output_file: Path,
    incidents: pd.DataFrame,
    columns: dict[str, str | None],
    level_value: str,
    term_days: int,
) -> dict[str, float | int]:
    date_col = columns["date"]
    student_id_col = columns["student_id"] or "Student_ID"
    name_col = columns["name"] or "Name"
    class_col = columns["class"] or "Class"
    level_col = columns["level"] or "Level"

    total_late_incidents = int(len(incidents))
    unique_late_days = int(incidents[date_col].nunique())
    school_impact_rate = (unique_late_days / term_days * 100) if term_days else 0
    average_lateness_density = (
        total_late_incidents / unique_late_days if unique_late_days > 0 else 0
    )
    level_summary_df = build_level_summary_df(incidents, date_col, level_col, term_days)

    summary_df = pd.DataFrame(
        {
            "Metric": [
                "Total Late Incidents",
                "Unique Late Days",
                "School Impact Rate",
                "Average Lateness Density",
            ],
            "Value": [
                total_late_incidents,
                unique_late_days,
                school_impact_rate,
                average_lateness_density,
            ],
            "Display": [
                f"{total_late_incidents}",
                f"{unique_late_days}",
                f"{school_impact_rate:.1f}%",
                f"{average_lateness_density:.2f}",
            ],
            "Definition": [
                "Individual student late events",
                "Calendar days where at least 1 student was late",
                "Percentage of term affected by lateness",
                "Average students late per affected day",
            ],
        }
    )

    daily_summary_df = (
        incidents.groupby(date_col)
        .size()
        .reset_index(name="Late Incidents")
        .sort_values(date_col)
    )

    available_student_fields = [col for col in [student_id_col, name_col, class_col, level_col] if col in incidents.columns]
    student_summary_df = (
        incidents.groupby(available_student_fields)
        .size()
        .reset_index(name="Total Late Incidents")
        .sort_values("Total Late Incidents", ascending=False)
    )

    top_10_days_df = daily_summary_df.sort_values("Late Incidents", ascending=False).head(10).copy()
    top_10_students_df = student_summary_df.head(10).copy()

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Management Summary", index=False)
        level_summary_df.to_excel(writer, sheet_name="Level Summary", index=False)
        daily_summary_df.to_excel(writer, sheet_name="Daily Breakdown", index=False)
        student_summary_df.to_excel(writer, sheet_name="Student Summary", index=False)
        top_10_days_df.to_excel(writer, sheet_name="Top 10 Days", index=False)
        top_10_students_df.to_excel(writer, sheet_name="Top 10 Students", index=False)
        incidents.to_excel(writer, sheet_name="Filtered Late Incidents", index=False)

        wb = writer.book
        for sheet_name in [
            "Management Summary",
            "Level Summary",
            "Daily Breakdown",
            "Student Summary",
            "Top 10 Days",
            "Top 10 Students",
            "Filtered Late Incidents",
        ]:
            ws = wb[sheet_name]
            apply_table_style(ws)
            autosize_columns(ws)

        dashboard = wb.create_sheet("Executive Dashboard", 0)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        dark_blue_fill = PatternFill("solid", fgColor="1F4E78")
        white_font = Font(color="FFFFFF", bold=True)

        for col in range(1, 13):
            dashboard.column_dimensions[get_column_letter(col)].width = 18

        dashboard.merge_cells("A1:L1")
        dashboard["A1"] = f"{level_value.upper()} Lateness Executive Dashboard"
        dashboard["A1"].fill = dark_blue_fill
        dashboard["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        dashboard["A1"].alignment = Alignment(horizontal="center")

        dashboard.merge_cells("A2:L2")
        dashboard["A2"] = f"Term Length: {term_days} school days | Filtered Level: {level_value.upper()}"
        dashboard["A2"].alignment = Alignment(horizontal="center")

        kpis = [
            ("A4:C6", "Total Late Incidents", total_late_incidents),
            ("D4:F6", "Unique Late Days", unique_late_days),
            ("G4:I6", "School Impact Rate", f"{school_impact_rate:.1f}%"),
            ("J4:L6", "Average Lateness Density", f"{average_lateness_density:.2f}"),
        ]
        for cell_range, label, value in kpis:
            dashboard.merge_cells(cell_range)
            cell = dashboard[cell_range.split(":")[0]]
            cell.value = f"{label}\n{value}"
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, size=15)
            cell.border = thin_border

        dashboard.conditional_formatting.add(
            "A4:L6",
            ColorScaleRule(
                start_type="min",
                start_color="E2F0D9",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFF2CC",
                end_type="max",
                end_color="FCE4D6",
            ),
        )

        dashboard.merge_cells("A8:L10")
        dashboard["A8"] = (
            f"Interpretation: {level_value.upper()} recorded {total_late_incidents} late incidents across "
            f"{unique_late_days} unique late days. This affected {school_impact_rate:.1f}% of the term, "
            f"with an average of {average_lateness_density:.2f} students late on each affected day."
        )
        dashboard["A8"].alignment = Alignment(wrap_text=True, vertical="top")
        dashboard["A8"].border = thin_border

        dashboard["A12"] = "Top 10 Worst Days"
        dashboard["A13"] = "Date"
        dashboard["B13"] = "Late Incidents"
        dashboard["E12"] = "Top 10 Students"
        dashboard["E13"] = "Name"
        dashboard["F13"] = "Late Incidents"
        dashboard["A30"] = "Jenjang Summary"
        dashboard["A31"] = "Level"
        dashboard["B31"] = "Late Incidents"

        for cell_ref in ["A13", "B13", "E13", "F13", "A31", "B31"]:
            dashboard[cell_ref].fill = dark_blue_fill
            dashboard[cell_ref].font = white_font
            dashboard[cell_ref].border = thin_border

        start_row = 14
        for idx, row in enumerate(top_10_days_df.itertuples(index=False), start=start_row):
            dashboard[f"A{idx}"] = str(row[0])
            dashboard[f"B{idx}"] = int(row[1])
            dashboard[f"A{idx}"].border = thin_border
            dashboard[f"B{idx}"].border = thin_border

        name_output_col = name_col if name_col in top_10_students_df.columns else available_student_fields[0]
        for idx, row_dict in enumerate(top_10_students_df.to_dict(orient="records"), start=start_row):
            dashboard[f"E{idx}"] = str(row_dict.get(name_output_col, "Unknown"))
            dashboard[f"F{idx}"] = int(row_dict["Total Late Incidents"])
            dashboard[f"E{idx}"].border = thin_border
            dashboard[f"F{idx}"].border = thin_border

        level_summary_start_row = 32
        for idx, row_dict in enumerate(level_summary_df.to_dict(orient="records"), start=level_summary_start_row):
            dashboard[f"A{idx}"] = str(row_dict["Level"])
            dashboard[f"B{idx}"] = int(row_dict["Total Late Incidents"])
            dashboard[f"A{idx}"].border = thin_border
            dashboard[f"B{idx}"].border = thin_border

        max_top_day_row = max(start_row, start_row + len(top_10_days_df) - 1)
        max_top_student_row = max(start_row, start_row + len(top_10_students_df) - 1)
        max_level_summary_row = max(level_summary_start_row, level_summary_start_row + len(level_summary_df) - 1)
        dashboard.conditional_formatting.add(
            f"B{start_row}:B{max_top_day_row}",
            ColorScaleRule(
                start_type="min",
                start_color="E2F0D9",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFF2CC",
                end_type="max",
                end_color="F8696B",
            ),
        )
        dashboard.conditional_formatting.add(
            f"F{start_row}:F{max_top_student_row}",
            ColorScaleRule(
                start_type="min",
                start_color="E2F0D9",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFF2CC",
                end_type="max",
                end_color="F8696B",
            ),
        )
        dashboard.conditional_formatting.add(
            f"B{level_summary_start_row}:B{max_level_summary_row}",
            ColorScaleRule(
                start_type="min",
                start_color="E2F0D9",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFF2CC",
                end_type="max",
                end_color="F8696B",
            ),
        )

        top_days_sheet = wb["Top 10 Days"]
        top_students_sheet = wb["Top 10 Students"]
        top_days_sheet.conditional_formatting.add(
            f"B2:B{top_days_sheet.max_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=["5"], fill=PatternFill("solid", fgColor="FCE4D6")),
        )
        top_students_sheet.conditional_formatting.add(
            f"{get_column_letter(top_students_sheet.max_column)}2:{get_column_letter(top_students_sheet.max_column)}{top_students_sheet.max_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=["3"], fill=PatternFill("solid", fgColor="FFF2CC")),
        )

        chart1 = BarChart()
        chart1.title = "Top 10 Worst Late Days"
        chart1.y_axis.title = "Late Incidents"
        chart1.x_axis.title = "Date"
        chart1.height = 8
        chart1.width = 14
        chart1.add_data(Reference(dashboard, min_col=2, min_row=13, max_row=max_top_day_row), titles_from_data=True)
        chart1.set_categories(Reference(dashboard, min_col=1, min_row=14, max_row=max_top_day_row))
        chart1.legend = None
        dashboard.add_chart(chart1, "H12")

        chart2 = BarChart()
        chart2.title = "Top 10 Students by Late Incidents"
        chart2.y_axis.title = "Late Incidents"
        chart2.x_axis.title = "Student"
        chart2.height = 8
        chart2.width = 14
        chart2.add_data(Reference(dashboard, min_col=6, min_row=13, max_row=max_top_student_row), titles_from_data=True)
        chart2.set_categories(Reference(dashboard, min_col=5, min_row=14, max_row=max_top_student_row))
        chart2.legend = None
        dashboard.add_chart(chart2, "H30")

        chart3 = BarChart()
        chart3.title = "Late Incidents by Jenjang"
        chart3.y_axis.title = "Late Incidents"
        chart3.x_axis.title = "Jenjang"
        chart3.height = 8
        chart3.width = 14
        chart3.add_data(Reference(dashboard, min_col=2, min_row=31, max_row=max_level_summary_row), titles_from_data=True)
        chart3.set_categories(Reference(dashboard, min_col=1, min_row=32, max_row=max_level_summary_row))
        chart3.legend = None
        dashboard.add_chart(chart3, "H48")

    return {
        "total_late_incidents": total_late_incidents,
        "unique_late_days": unique_late_days,
        "school_impact_rate": school_impact_rate,
        "average_lateness_density": average_lateness_density,
    }


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_file)
    output_path = Path(args.output)

    df = load_source(input_path, args.sheet)
    incidents, columns = build_late_incident_frame(df)
    incidents, level_col = ensure_level(incidents, columns["level"], args.default_level)
    columns["level"] = level_col

    filtered = incidents[
        incidents[level_col].astype(str).str.strip().ne("")
    ].copy()
    requested_level = args.level_value.strip().upper()
    if requested_level != "ALL":
        filtered = filtered[
            filtered[level_col].astype(str).str.strip().str.upper().eq(requested_level)
        ].copy()
    if filtered.empty:
        raise ValueError(f"No late incidents found for level '{args.level_value}'.")

    metrics = build_dashboard(output_path, filtered, columns, args.level_value, args.term_days)
    print(f"Dashboard created: {output_path}")
    print(f"Total Late Incidents     : {metrics['total_late_incidents']}")
    print(f"Unique Late Days         : {metrics['unique_late_days']}")
    print(f"School Impact Rate       : {metrics['school_impact_rate']:.1f}%")
    print(f"Average Lateness Density : {metrics['average_lateness_density']:.2f}")


if __name__ == "__main__":
    main()
